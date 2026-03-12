"""
Rutas de Carteras
Endpoints para gestión de carteras de clientes
"""

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query, Form
from fastapi.responses import FileResponse
from pathlib import Path
from datetime import datetime
import shutil
import openpyxl
import pandas as pd
import json
from typing import Dict, Any, List, Optional
from pydantic import BaseModel

from app.models.user import UserInfo
from app.utils.dependencies import get_current_user
from app.utils.excel_cache import excel_cache

# Crear router
router = APIRouter()

# Rutas de archivos para Carteras
CARTERAS_BASE_DIR = Path(__file__).parent.parent.parent / "data" / "carga_cm" / "Cartera"
CARTERAS_RESPALDO_DIR = CARTERAS_BASE_DIR / "Respaldo"
CARTERAS_CAMBIOS_DIR = CARTERAS_BASE_DIR / "Cambios"
CARTERAS_BASE_FILE = CARTERAS_RESPALDO_DIR / "TBL_Cartera.xlsx"
USUARIOS_SISTEMA_PATH = Path(__file__).parent.parent.parent / "data" / "usuarios_sistema.xlsx"

# Rutas de archivos para Estructura de Venta
ESTRUCTURA_VENTA_DIR = Path(__file__).parent.parent.parent / "data" / "carga_cm" / "EstructuraVenta"
ESTRUCTURA_VENTA_CAMBIOS_DIR = ESTRUCTURA_VENTA_DIR / "Cambios"

# Asegurar que existen las carpetas
CARTERAS_RESPALDO_DIR.mkdir(parents=True, exist_ok=True)
CARTERAS_CAMBIOS_DIR.mkdir(parents=True, exist_ok=True)


def limpiar_carpeta_cambios_carteras() -> list[str]:
    """Elimina TODOS los archivos de la carpeta Cambios de Carteras."""
    eliminados = []
    for archivo in CARTERAS_CAMBIOS_DIR.iterdir():
        if archivo.is_file():
            try:
                archivo.unlink()
                eliminados.append(archivo.name)
                print(f"  Eliminado: {archivo.name}")
            except Exception as e:
                print(f"  Error al eliminar {archivo.name}: {str(e)}")
    return eliminados


def limpiar_backups_respaldo_carteras() -> list[str]:
    """Elimina archivos de backup antiguos en Respaldo, conservando el archivo base actual."""
    eliminados = []
    for archivo in CARTERAS_RESPALDO_DIR.iterdir():
        if archivo.is_file() and archivo != CARTERAS_BASE_FILE:
            try:
                archivo.unlink()
                eliminados.append(archivo.name)
                print(f"  Eliminado backup: {archivo.name}")
            except Exception as e:
                print(f"  Error al eliminar backup {archivo.name}: {str(e)}")
    return eliminados


# Funciones auxiliares para validación
def obtener_archivo_jefe_cartera(usuario: str, zona: Optional[str] = None) -> Path:
    """Obtiene la ruta del archivo individual del jefe para carteras (por zona si se especifica)"""
    if zona:
        return CARTERAS_CAMBIOS_DIR / f"TBL_Cartera_{usuario}_{zona}.xlsx"
    return CARTERAS_CAMBIOS_DIR / f"TBL_Cartera_{usuario}.xlsx"


def obtener_archivo_estado_validacion_cartera(usuario: str, zona: Optional[str] = None) -> Path:
    """Obtiene la ruta del archivo de estado de validación del jefe para carteras (por zona si se especifica)"""
    if zona:
        return CARTERAS_CAMBIOS_DIR / f"VALIDADO_{usuario}_{zona}.json"
    return CARTERAS_CAMBIOS_DIR / f"VALIDADO_{usuario}.json"


def esta_validado_cartera(usuario: str, zona: Optional[str] = None) -> bool:
    """Verifica si un jefe ya validó su cartera (por zona si se especifica).
    Si zona=None, retorna True solo si TODAS las zonas del usuario están validadas."""
    if zona:
        archivo_estado = obtener_archivo_estado_validacion_cartera(usuario, zona)
        if not archivo_estado.exists():
            return False
        try:
            with open(archivo_estado, 'r', encoding='utf-8') as f:
                estado = json.load(f)
                return estado.get('validado', False)
        except:
            return False
    else:
        # Sin zona específica: verificar que TODAS las zonas del usuario estén validadas
        try:
            zonas_usuario = obtener_zonas_usuario(usuario)
            if not zonas_usuario:
                # Sin zonas configuradas: fallback al archivo legacy sin zona
                archivo_legacy = CARTERAS_CAMBIOS_DIR / f"VALIDADO_{usuario}.json"
                if archivo_legacy.exists():
                    try:
                        with open(archivo_legacy, 'r', encoding='utf-8') as f:
                            estado = json.load(f)
                            return estado.get('validado', False)
                    except:
                        pass
                return False
            # Todas las zonas deben estar validadas
            return all(esta_validado_cartera(usuario, z) for z in zonas_usuario)
        except:
            return False


def marcar_como_validado_cartera(usuario: str, zona: Optional[str] = None) -> None:
    """Marca explícitamente que el jefe validó su cartera (por zona si se especifica)"""
    archivo_estado = obtener_archivo_estado_validacion_cartera(usuario, zona)
    
    # Obtener información del usuario
    nombre_usuario = usuario
    zonas_usuario = obtener_zonas_usuario(usuario)
    
    try:
        df_usuarios = excel_cache.read_excel(USUARIOS_SISTEMA_PATH, sheet_name='Sheet1')
        usuario_data = df_usuarios[df_usuarios['Usuario'] == usuario]
        if not usuario_data.empty:
            nombre_usuario = usuario_data.iloc[0].get('Nombre', usuario)
    except:
        pass
    
    estado = {
        "validado": True,
        "usuario": usuario,
        "nombre": nombre_usuario,
        "zona": zona,
        "zonas": zonas_usuario,
        "fecha_validacion": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "origen": "Carteras"
    }
    
    with open(archivo_estado, 'w', encoding='utf-8') as f:
        json.dump(estado, f, indent=2, ensure_ascii=False)
    
    print(f"✓ Validación de cartera marcada para {usuario} zona={zona}")


# Modelos Pydantic
class ClienteCartera(BaseModel):
    CodDistrito: str
    DesDistrito: str
    CodVend: int
    NombreVend: str
    CodCliente: int
    RutCliente: str
    RazonSocial: str
    TipoNeg: str
    Relev: int
    NivPrecio: str
    Direccion: str
    Comuna: str

    class Config:
        from_attributes = True


class CarteraResponse(BaseModel):
    clientes: List[ClienteCartera]
    metadata: Dict[str, Any]


class ValidarZonaCarterasRequest(BaseModel):
    zona: Optional[str] = None


def obtener_zonas_usuario(usuario: str) -> List[str]:
    """
    Obtiene las zonas asignadas a un usuario desde usuarios_sistema.xlsx
    """
    try:
        if not USUARIOS_SISTEMA_PATH.exists():
            print(f"Archivo usuarios_sistema.xlsx no encontrado en {USUARIOS_SISTEMA_PATH}")
            return []
        
        df = excel_cache.read_excel(USUARIOS_SISTEMA_PATH, sheet_name='Sheet1')
        
        # Buscar el usuario (la columna se llama 'Usuario' con mayúscula)
        usuario_data = df[df['Usuario'] == usuario]
        
        if usuario_data.empty:
            print(f"Usuario {usuario} no encontrado en usuarios_sistema.xlsx")
            return []
        
        # Obtener las zonas - las columnas de zonas contienen 'X' cuando están asignadas
        # Excluir las primeras columnas que son datos del usuario
        columnas_sistema = ['Usuario', 'Cargo', 'Nombre', 'Correo', 'Contraseña', 'Cambiar', 'Bloquear']
        zonas = []
        
        for col in df.columns:
            if col not in columnas_sistema:
                zona_value = usuario_data.iloc[0][col]
                # Si tiene 'X' o cualquier valor que no sea NaN, está asignada
                if pd.notna(zona_value):
                    zonas.append(str(col))
        
        print(f"Zonas encontradas para {usuario}: {zonas}")
        return zonas
    
    except Exception as e:
        print(f"Error al obtener zonas del usuario: {str(e)}")
        import traceback
        traceback.print_exc()
        return []


def obtener_estatus_vendedores_estructura_venta(usuario: str, zona_filtro: Optional[str] = None) -> tuple[Dict[int, str], str]:
    """
    Obtiene el Estatus de TODOS los vendedores de la Estructura de Venta del usuario.
    Retorna (dict_codigo_estatus, ruta_archivo_usado)
    Ejemplo: {777: 'Eliminado', 128: 'Antiguo', 999: 'Nuevo'}
    Si zona_filtro se indica, solo devuelve vendedores de esa zona.
    """
    try:
        # Obtener las zonas asignadas al usuario
        user_zonas = obtener_zonas_usuario(usuario)
        
        if not user_zonas:
            print(f"Usuario {usuario} no tiene zonas asignadas")
            return ({}, "")
        
        # Si se especifica una zona concreta, filtrar solo por ella
        if zona_filtro:
            zonas_a_filtrar = [zona_filtro]
        else:
            zonas_a_filtrar = user_zonas
        
        # Buscar archivo individual del usuario
        archivo_usuario = ESTRUCTURA_VENTA_CAMBIOS_DIR / f"TBL_EstructuraVenta_{usuario}.xlsx"
        
        if not archivo_usuario.exists():
            # Buscar de forma flexible
            archivos_ev = list(ESTRUCTURA_VENTA_CAMBIOS_DIR.glob("TBL_EstructuraVenta_*.xlsx"))
            if not archivos_ev:
                return ({}, "")
            archivo_usuario = archivos_ev[0] if len(archivos_ev) == 1 else max(archivos_ev, key=lambda f: f.stat().st_mtime)
        
        # Leer la hoja Dotacion
        try:
            df = excel_cache.read_excel(archivo_usuario, sheet_name='Dotacion')
        except Exception as e:
            print(f"Error al leer hoja 'Dotacion': {str(e)}")
            return ({}, "")
        
        # Verificar columnas necesarias
        if 'CodVenta' not in df.columns:
            return ({}, "")
        
        # Asegurar que existe columna Estatus
        if 'Estatus' not in df.columns:
            df['Estatus'] = 'Antiguo'
        else:
            df['Estatus'] = df['Estatus'].fillna('Antiguo')
        
        # Filtrar por zonas del usuario (o por zona_filtro si se especificó)
        columna_distrito = None
        for col in ['CodDistrito', 'Distrito', 'Zona', 'CodZona', 'DesDistrito']:
            if col in df.columns:
                columna_distrito = col
                break
        
        if columna_distrito:
            df = df[df[columna_distrito].isin(zonas_a_filtrar)]
        
        # Crear diccionario código: estatus
        estatus_dict = {}
        for _, row in df.iterrows():
            cod_venta = row['CodVenta']
            if pd.notna(cod_venta):
                estatus_dict[int(cod_venta)] = str(row['Estatus'])
        
        print(f"✓ Leídos {len(estatus_dict)} vendedores con su Estatus de {archivo_usuario.name}")
        return (estatus_dict, str(archivo_usuario))
    
    except Exception as e:
        print(f"Error al obtener estatus de vendedores: {str(e)}")
        import traceback
        traceback.print_exc()
        return ({}, "")


def obtener_vendedores_con_cargo_estructura_venta(usuario: str, zona_filtro: Optional[str] = None) -> tuple[List[int], Dict[int, str], str]:
    """
    Obtiene los vendedores de la Estructura de Venta del usuario (excluyendo hojas de reemplazos).
    Solo incluye vendedores de los distritos asignados al usuario.
    Si zona_filtro se indica, solo devuelve vendedores de esa zona.
    Retorna (lista_codigos_vendedor, dict_codigos_cargo, ruta_archivo_usado)
    """
    try:
        # PRIMERO: Obtener las zonas asignadas al usuario
        user_zonas = obtener_zonas_usuario(usuario)
        
        if not user_zonas:
            print(f"Usuario {usuario} no tiene zonas asignadas")
            return ([], {}, "")
        
        # Si se especifica una zona concreta, filtrar solo por ella
        if zona_filtro:
            zonas_a_filtrar = [zona_filtro]
        else:
            zonas_a_filtrar = user_zonas
        
        print(f" Zonas a filtrar para {usuario}: {zonas_a_filtrar}")
        # Buscar archivo individual del usuario (intento exacto)
        archivo_usuario = ESTRUCTURA_VENTA_CAMBIOS_DIR / f"TBL_EstructuraVenta_{usuario}.xlsx"
        
        # Si no existe con el nombre exacto, buscar de forma flexible
        if not archivo_usuario.exists():
            print(f"No se encontró archivo exacto: TBL_EstructuraVenta_{usuario}.xlsx")
            print(f" Buscando archivos de Estructura de Venta en la carpeta...")
            
            # Buscar todos los archivos que empiecen con TBL_EstructuraVenta_
            archivos_ev = list(ESTRUCTURA_VENTA_CAMBIOS_DIR.glob("TBL_EstructuraVenta_*.xlsx"))
            
            if not archivos_ev:
                print(f" No se encontró ningún archivo de Estructura de Venta en {ESTRUCTURA_VENTA_CAMBIOS_DIR}")
                return ([], {}, "")
            
            # Si hay exactamente un archivo, usarlo
            if len(archivos_ev) == 1:
                archivo_usuario = archivos_ev[0]
                print(f" Se encontró un archivo de Estructura de Venta: {archivo_usuario.name}")
            else:
                # Si hay múltiples archivos, buscar el más reciente o el que coincida parcialmente
                print(f"Se encontraron {len(archivos_ev)} archivos de Estructura de Venta")
                
                # Intentar buscar por coincidencia parcial del nombre de usuario
                usuario_parts = usuario.lower().replace('.', '_').split('_')
                archivo_encontrado = None
                
                for archivo in archivos_ev:
                    archivo_nombre = archivo.stem.lower()  # nombre sin extensión
                    # Verificar si alguna parte del usuario está en el nombre del archivo
                    if any(part in archivo_nombre for part in usuario_parts if len(part) > 2):
                        archivo_encontrado = archivo
                        print(f" Se encontró coincidencia parcial: {archivo.name}")
                        break
                
                # Si no se encuentra coincidencia, usar el más reciente
                if not archivo_encontrado:
                    archivo_encontrado = max(archivos_ev, key=lambda f: f.stat().st_mtime)
                    print(f"Usando el archivo más reciente: {archivo_encontrado.name}")
                
                archivo_usuario = archivo_encontrado
        else:
            print(f" Archivo exacto encontrado: {archivo_usuario.name}")
        
        print(f" Leyendo Estructura de Venta de: {archivo_usuario.name}")
        
        # Leer directamente la hoja Dotacion usando pandas
        try:
            df = excel_cache.read_excel(archivo_usuario, sheet_name='Dotacion')
            print(f" Hoja 'Dotacion' leída correctamente: {len(df)} registros")
        except Exception as e:
            print(f" Error al leer hoja 'Dotacion': {str(e)}")
            return ([], {}, "")
        
        # IMPORTANTE: Excluir vendedores eliminados
        registros_antes = len(df)
        if 'Estatus' in df.columns:
            df = df[df['Estatus'] != 'Eliminado'].copy()
            registros_despues = len(df)
            eliminados_excluidos = registros_antes - registros_despues
            if eliminados_excluidos > 0:
                print(f"✓ Excluidos {eliminados_excluidos} vendedores con Estatus='Eliminado'")
        else:
            print(f"⚠ Columna 'Estatus' no encontrada, no se pueden excluir eliminados")
        
        # Verificar que tenga las columnas necesarias
        if 'CodVenta' not in df.columns:
            print(f" No se encontró la columna 'CodVenta' en la hoja Dotacion")
            print(f" Columnas disponibles: {list(df.columns)}")
            return ([], {}, "")
        
        vendedores_codigos = set()
        vendedores_cargo = {}  # Diccionario para almacenar código: cargo
        
        # FILTRAR POR DISTRITO/ZONA del usuario
        columna_distrito = None
        posibles_nombres_distrito = ['CodDistrito', 'Distrito', 'Zona', 'CodZona', 'DesDistrito']
        
        for col in posibles_nombres_distrito:
            if col in df.columns:
                columna_distrito = col
                break
        
        if columna_distrito:
            # Filtrar solo los vendedores de la(s) zona(s) correspondiente(s)
            df_filtrado = df[df[columna_distrito].isin(zonas_a_filtrar)]
            print(f" Registros antes del filtro: {len(df)}")
            print(f" Registros después de filtrar por zonas {zonas_a_filtrar}: {len(df_filtrado)}")
        else:
            # Si no hay columna de distrito, usar todos (caso legacy)
            print(f"No se encontró columna de distrito - Usando TODOS los vendedores")
            df_filtrado = df
        
        if len(df_filtrado) == 0:
            print(f"No hay vendedores después de filtrar por zonas {zonas_a_filtrar}")
            return ([], {}, "")
        
        # Obtener códigos de vendedor
        vendedores_codigos = set(df_filtrado['CodVenta'].dropna().astype(int).unique())
        
        # Almacenar el cargo de cada vendedor
        if 'Cargo' not in df_filtrado.columns:
            print(f"No se encontró la columna 'Cargo' en la hoja Dotacion")
            print(f"Columnas disponibles: {list(df_filtrado.columns)}")
            # Continuar sin cargos
            vendedores_cargo = {cod: 'Sin Cargo' for cod in vendedores_codigos}
        else:
            print(f"Columna 'Cargo' encontrada, leyendo cargos de vendedores...")
            for _, row in df_filtrado.iterrows():
                if pd.notna(row['CodVenta']):
                    cod = int(row['CodVenta'])
                    cargo_raw = row['Cargo']
                    cargo = str(cargo_raw).strip() if pd.notna(cargo_raw) else 'Sin Cargo'
                    vendedores_cargo[cod] = cargo
                    # Solo imprimir algunos vendedores para no llenar el log
                    if cod in [1231459, 1236732, 1236733]:  # Los vendedores problemáticos
                        print(f"    Vendedor {cod}: Cargo RAW='{cargo_raw}' | Cargo PROCESADO='{cargo}' | Tipo={type(cargo_raw)}")
        
        vendedores_lista = sorted(list(vendedores_codigos))
        
        print(f"Total vendedores únicos de zonas {zonas_a_filtrar}: {len(vendedores_lista)}")
        print(f"Vendedores con información de cargo: {len(vendedores_cargo)}")
        
        return (vendedores_lista, vendedores_cargo, str(archivo_usuario))
    
    except Exception as e:
        print(f" Error al leer Estructura de Venta: {str(e)}")
        import traceback
        traceback.print_exc()
        return ([], {}, "")


def obtener_vendedores_estructura_venta(usuario: str) -> tuple[List[int], str]:
    """
    Obtiene los vendedores de la Estructura de Venta del usuario (excluyendo reemplazos).
    Solo incluye vendedores de los distritos asignados al usuario.
    Retorna (lista_codigos_vendedor, ruta_archivo_usado)
    """
    try:
        # PRIMERO: Obtener las zonas asignadas al usuario
        user_zonas = obtener_zonas_usuario(usuario)
        
        if not user_zonas:
            print(f"  Usuario {usuario} no tiene zonas asignadas")
            return ([], "")
        
        print(f" Zonas asignadas a {usuario}: {user_zonas}")
        # Buscar archivo individual del usuario (intento exacto)
        archivo_usuario = ESTRUCTURA_VENTA_CAMBIOS_DIR / f"TBL_EstructuraVenta_{usuario}.xlsx"
        
        # Si no existe con el nombre exacto, buscar de forma flexible
        if not archivo_usuario.exists():
            print(f" No se encontró archivo exacto: TBL_EstructuraVenta_{usuario}.xlsx")
            print(f" Buscando archivos de Estructura de Venta en la carpeta...")
            
            # Buscar todos los archivos que empiecen con TBL_EstructuraVenta_
            archivos_ev = list(ESTRUCTURA_VENTA_CAMBIOS_DIR.glob("TBL_EstructuraVenta_*.xlsx"))
            
            if not archivos_ev:
                print(f" No se encontró ningún archivo de Estructura de Venta en {ESTRUCTURA_VENTA_CAMBIOS_DIR}")
                return ([], "")
            
            # Si hay exactamente un archivo, usarlo
            if len(archivos_ev) == 1:
                archivo_usuario = archivos_ev[0]
                print(f" Se encontró un archivo de Estructura de Venta: {archivo_usuario.name}")
            else:
                # Si hay múltiples archivos, buscar el más reciente o el que coincida parcialmente
                print(f" Se encontraron {len(archivos_ev)} archivos de Estructura de Venta")
                
                # Intentar buscar por coincidencia parcial del nombre de usuario
                usuario_parts = usuario.lower().replace('.', '_').split('_')
                archivo_encontrado = None
                
                for archivo in archivos_ev:
                    archivo_nombre = archivo.stem.lower()  # nombre sin extensión
                    # Verificar si alguna parte del usuario está en el nombre del archivo
                    if any(part in archivo_nombre for part in usuario_parts if len(part) > 2):
                        archivo_encontrado = archivo
                        print(f" Se encontró coincidencia parcial: {archivo.name}")
                        break
                
                # Si no se encuentra coincidencia, usar el más reciente
                if not archivo_encontrado:
                    archivo_encontrado = max(archivos_ev, key=lambda f: f.stat().st_mtime)
                    print(f" Usando el archivo más reciente: {archivo_encontrado.name}")
                
                archivo_usuario = archivo_encontrado
        else:
            print(f" Archivo exacto encontrado: {archivo_usuario.name}")
        
        print(f" Leyendo Estructura de Venta de: {archivo_usuario.name}")
        
        # Leer el archivo Excel
        wb = openpyxl.load_workbook(archivo_usuario, read_only=True, data_only=True)
        
        # Mostrar todas las hojas disponibles para diagnóstico
        print(f" Hojas disponibles en el archivo: {wb.sheetnames}")
        
        vendedores_codigos = set()
        
        # Hojas que se deben EXCLUIR
        hojas_excluir = ['Vendedor Reemplazo', 'Dotacion Eliminada', 'Dotación Eliminada']
        
        # Leer TODAS las hojas EXCEPTO las excluidas
        hojas_leidas = 0
        vendedores_totales_antes_filtro = 0
        
        for sheet_name in wb.sheetnames:
            # Saltar hojas excluidas
            if sheet_name in hojas_excluir:
                print(f"    ⏭  Saltando hoja excluida: '{sheet_name}'")
                continue
            
            try:
                df = pd.read_excel(archivo_usuario, sheet_name=sheet_name)
                
                # Verificar que tenga la columna CodVenta
                if 'CodVenta' not in df.columns:
                    print(f"    Hoja '{sheet_name}': No tiene columna 'CodVenta' - Columnas: {list(df.columns)[:5]}")
                    continue
                
                # IMPORTANTE: Excluir vendedores eliminados por si acaso
                registros_antes = len(df)
                if 'Estatus' in df.columns:
                    df = df[df['Estatus'] != 'Eliminado'].copy()
                    registros_despues = len(df)
                    if registros_despues < registros_antes:
                        print(f"    Hoja '{sheet_name}': Excluidos {registros_antes - registros_despues} vendedores eliminados")
                
                vendedores_totales_antes_filtro += len(df['CodVenta'].dropna())
                
                # FILTRAR POR DISTRITO/ZONA del usuario
                # Buscar columna de distrito (puede tener diferentes nombres)
                columna_distrito = None
                posibles_nombres_distrito = ['CodDistrito', 'Distrito', 'Zona', 'CodZona', 'DesDistrito']
                
                for col in posibles_nombres_distrito:
                    if col in df.columns:
                        columna_distrito = col
                        break
                
                if columna_distrito:
                    # Filtrar solo los vendedores de las zonas asignadas al usuario
                    df_filtrado = df[df[columna_distrito].isin(user_zonas)]
                    total_antes = len(df)
                    total_despues = len(df_filtrado)
                    
                    if total_despues > 0:
                        # Obtener códigos de vendedor después del filtro
                        codigos = df_filtrado['CodVenta'].dropna().astype(int).unique().tolist()
                        vendedores_codigos.update(codigos)
                        print(f"     Hoja '{sheet_name}': {len(codigos)} vendedores (filtrado de {total_antes} por distritos {user_zonas})")
                        hojas_leidas += 1
                    else:
                        print(f"     Hoja '{sheet_name}': Sin vendedores después de filtrar por distritos {user_zonas}")
                else:
                    # Si no hay columna de distrito, tomar todos (caso legacy)
                    print(f"    Hoja '{sheet_name}': No se encontró columna de distrito - Usando TODOS los vendedores")
                    codigos = df['CodVenta'].dropna().astype(int).unique().tolist()
                    if len(codigos) > 0:
                        vendedores_codigos.update(codigos)
                        print(f"    Hoja '{sheet_name}': {len(codigos)} vendedores (sin filtro de distrito)")
                        hojas_leidas += 1
                    
            except Exception as e:
                print(f"    Error al leer hoja '{sheet_name}': {str(e)}")
        
        wb.close()
        
        vendedores_lista = sorted(list(vendedores_codigos))
        
        if hojas_leidas == 0:
            print(f" No se pudo leer vendedores de ninguna hoja")
            return ([], "")
        
        print(f" Total vendedores únicos FILTRADOS por distritos {user_zonas}: {len(vendedores_lista)} (de {vendedores_totales_antes_filtro} totales)")
        
        return (vendedores_lista, str(archivo_usuario))
    
    except Exception as e:
        print(f" Error al leer Estructura de Venta: {str(e)}")
        import traceback
        traceback.print_exc()
        return ([], "")


def obtener_codigos_jefes_venta() -> List[int]:
    """
    Obtiene los códigos de vendedor de los Jefes de Venta.
    Los Jefes pueden estar en la hoja "Dotación" del archivo base de Estructura de Venta
    con cargos como "JEFE DE VENTA", "JEFE DISTRITO", etc.
    """
    try:
        # Archivo base de Estructura de Venta
        archivo_base_ev = ESTRUCTURA_VENTA_DIR / "Respaldo" / "TBL EstructuraVentaCM.xlsx"
        
        if not archivo_base_ev.exists():
            print(f" No se encontró archivo base de Estructura de Venta: {archivo_base_ev}")
            return []
        
        print(f" Leyendo Estructura de Venta base para identificar Jefes: {archivo_base_ev.name}")
        
        # Primero verificar qué hojas tiene el archivo
        wb = openpyxl.load_workbook(archivo_base_ev, read_only=True, data_only=True)
        hojas_disponibles = wb.sheetnames
        wb.close()
        
        print(f" Hojas disponibles en archivo base: {hojas_disponibles}")
        
        # Buscar la hoja de Dotación (puede tener diferentes nombres)
        hoja_dotacion = None
        posibles_nombres = ['Dotación', 'Dotacion', 'Dotación ', 'Dotacion ']
        
        for nombre in posibles_nombres:
            if nombre in hojas_disponibles:
                hoja_dotacion = nombre
                break
        
        # Si no se encuentra con nombres exactos, buscar parcialmente
        if not hoja_dotacion:
            for hoja in hojas_disponibles:
                if 'dotacion' in hoja.lower() or 'dotación' in hoja.lower():
                    hoja_dotacion = hoja
                    break
        
        if not hoja_dotacion:
            print(f" No se encontró hoja de Dotación en el archivo base")
            print(f" La validación de Jefes se omitirá (no se encontró archivo base válido)")
            return []
        
        print(f" Usando hoja: '{hoja_dotacion}'")
        
        # Leer la hoja de Dotación
        df = pd.read_excel(archivo_base_ev, sheet_name=hoja_dotacion)
        
        # IMPORTANTE: Excluir registros eliminados
        if 'Estatus' in df.columns:
            registros_antes = len(df)
            df = df[df['Estatus'] != 'Eliminado'].copy()
            registros_despues = len(df)
            if registros_despues < registros_antes:
                print(f" Excluidos {registros_antes - registros_despues} registros eliminados")
        
        # Buscar registros donde el Cargo contenga "JEFE"
        if 'Cargo' not in df.columns:
            print(f" No se encontró columna 'Cargo' en la hoja {hoja_dotacion}")
            return []
        
        df_jefes = df[df['Cargo'].str.upper().str.contains('JEFE', na=False)]
        
        # Obtener códigos de vendedor de los Jefes
        if 'CodVenta' in df_jefes.columns:
            codigos_jefes = df_jefes['CodVenta'].dropna().astype(int).unique().tolist()
            if len(codigos_jefes) > 0:
                print(f" Se encontraron {len(codigos_jefes)} Jefe(s) de Venta con códigos: {codigos_jefes}")
            else:
                print(f"  No se encontraron Jefes de Venta en la hoja {hoja_dotacion}")
            return codigos_jefes
        else:
            print(f"  No se encontró columna 'CodVenta' en la hoja {hoja_dotacion}")
            return []
        
    except Exception as e:
        print(f"  Error al obtener códigos de Jefes de Venta: {str(e)}")
        print(f"  La validación de Jefes se omitirá")
        return []


def obtener_info_jefes_venta() -> Dict[int, str]:
    """
    Obtiene los códigos y nombres de los Jefes de Venta.
    Retorna un diccionario: {codigo: "Nombre Completo"}
    """
    try:
        archivo_base_ev = ESTRUCTURA_VENTA_DIR / "Respaldo" / "TBL EstructuraVentaCM.xlsx"
        
        if not archivo_base_ev.exists():
            return {}
        
        # Buscar hoja de Dotación
        wb = openpyxl.load_workbook(archivo_base_ev, read_only=True, data_only=True)
        hojas_disponibles = wb.sheetnames
        wb.close()
        
        hoja_dotacion = None
        posibles_nombres = ['Dotación', 'Dotacion', 'Dotación ', 'Dotacion ']
        
        for nombre in posibles_nombres:
            if nombre in hojas_disponibles:
                hoja_dotacion = nombre
                break
        
        if not hoja_dotacion:
            for hoja in hojas_disponibles:
                if 'dotacion' in hoja.lower() or 'dotación' in hoja.lower():
                    hoja_dotacion = hoja
                    break
        
        if not hoja_dotacion:
            return {}
        
        # Leer hoja
        df = pd.read_excel(archivo_base_ev, sheet_name=hoja_dotacion)
        
        # IMPORTANTE: Excluir registros eliminados
        if 'Estatus' in df.columns:
            df = df[df['Estatus'] != 'Eliminado'].copy()
        
        if 'Cargo' not in df.columns or 'CodVenta' not in df.columns:
            return {}
        
        df_jefes = df[df['Cargo'].str.upper().str.contains('JEFE', na=False)]
        
        # Crear diccionario con código: nombre
        info_jefes = {}
        for _, row in df_jefes.iterrows():
            if pd.notna(row['CodVenta']):
                cod = int(row['CodVenta'])
                
                # Intentar construir nombre completo
                nombre_parts = []
                if 'Nombre' in df.columns and pd.notna(row['Nombre']):
                    nombre_parts.append(str(row['Nombre']))
                if 'APaterno' in df.columns and pd.notna(row['APaterno']):
                    nombre_parts.append(str(row['APaterno']))
                if 'AMaterno' in df.columns and pd.notna(row['AMaterno']):
                    nombre_parts.append(str(row['AMaterno']))
                
                nombre_completo = ' '.join(nombre_parts) if nombre_parts else f"Jefe {cod}"
                info_jefes[cod] = nombre_completo
        
        return info_jefes
        
    except Exception as e:
        print(f" Error al obtener info de Jefes: {str(e)}")
        return {}


@router.post("/upload-base-file")
async def upload_base_file(
    file: UploadFile = File(...),
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Sube el Maestro Cliente de Carteras que será usado como plantilla.
    Solo accesible para usuarios ADMIN.
    
    Args:
        file: Archivo Excel con la información de carteras
        current_user: Usuario actual (debe ser ADMIN)
        
    Returns:
        Información sobre el archivo subido
    """
    # Verificar permisos de ADMIN
    if current_user.cargo.upper() != "ADMIN":
        raise HTTPException(
            status_code=403,
            detail="Solo usuarios ADMIN pueden subir archivos base"
        )
    
    # Validar que sea un archivo Excel
    if not file.filename:
        raise HTTPException(status_code=400, detail="No se proporcionó un archivo")
    
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        raise HTTPException(
            status_code=400,
            detail="El archivo debe ser un Excel (.xlsx o .xls)"
        )
    
    try:
        # Crear backup del archivo anterior si existe
        if CARTERAS_BASE_FILE.exists():
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_file = CARTERAS_RESPALDO_DIR / f"TBL_Cartera_backup_{timestamp}.xlsx"
            shutil.copy2(CARTERAS_BASE_FILE, backup_file)
        
        # Guardar el nuevo archivo
        with open(CARTERAS_BASE_FILE, 'wb') as f:
            contents = await file.read()
            f.write(contents)
        
        # Validar que tenga las columnas obligatorias
        try:
            df_validacion = pd.read_excel(CARTERAS_BASE_FILE, nrows=0)
            columnas_obligatorias = [
                'CodDistrito', 'DesDistrito', 'CodVend', 'NombreVend', 'CodCliente',
                'RutCliente', 'RazonSocial', 'TipoNeg', 'Relev', 'NivPrecio',
                'Direccion', 'Comuna'
            ]
            columnas_archivo = list(df_validacion.columns)
            columnas_faltantes = [col for col in columnas_obligatorias if col not in columnas_archivo]
            
            if columnas_faltantes:
                if CARTERAS_BASE_FILE.exists():
                    CARTERAS_BASE_FILE.unlink()
                raise HTTPException(
                    status_code=400,
                    detail={
                        "mensaje": f"Al archivo le faltan {len(columnas_faltantes)} columna(s) obligatoria(s).",
                        "columnas_faltantes": columnas_faltantes,
                        "sugerencia": f"Agrega las columnas: {', '.join(columnas_faltantes)}"
                    }
                )
        except HTTPException:
            raise
        except Exception as e:
            if CARTERAS_BASE_FILE.exists():
                CARTERAS_BASE_FILE.unlink()
            raise HTTPException(
                status_code=400,
                detail=f"No se pudo leer el archivo Excel: {str(e)}"
            )
        
        # Leer metadatos del archivo
        wb = openpyxl.load_workbook(CARTERAS_BASE_FILE, read_only=True, data_only=True)
        sheets = wb.sheetnames
        file_size = CARTERAS_BASE_FILE.stat().st_size
        wb.close()
        
        # PASO CRÍTICO: Eliminar todos los archivos de Cambios y backups de Respaldo
        # Esto fuerza a todos a comenzar desde el nuevo archivo base
        print(f"\n🔄 INICIANDO LIMPIEZA DE ARCHIVOS DE CARTERAS...")
        
        archivos_eliminados = limpiar_carpeta_cambios_carteras()
        backups_eliminados = limpiar_backups_respaldo_carteras()
        archivos_eliminados.extend(backups_eliminados)
        
        print(f"✓ Limpieza completada: {len(archivos_eliminados)} archivos eliminados")
        
        return {
            "success": True,
            "message": f"Maestro Cliente subido exitosamente. {len(archivos_eliminados)} archivos anteriores eliminados.",
            "filename": CARTERAS_BASE_FILE.name,
            "sheets": sheets,
            "size": file_size,
            "archivos_eliminados": len(archivos_eliminados),
            "upload_date": datetime.now().isoformat()
        }
        
    except HTTPException:
        raise
    except Exception as e:
        # Si hay error, intentar restaurar el backup
        if CARTERAS_BASE_FILE.exists():
            CARTERAS_BASE_FILE.unlink()
        
        raise HTTPException(
            status_code=500,
            detail=f"Error al procesar el archivo: {str(e)}"
        )


@router.get("/status")
async def obtener_status_carteras(current_user: UserInfo = Depends(get_current_user)):
    """
    Endpoint liviano para obtener solo el estado de validación de carteras.
    No lee el Excel completo — solo verifica archivos JSON y existencia de archivos.
    """
    try:
        is_admin = current_user.cargo.upper() in ('ADMIN', 'ADMINISTRADOR')
        if is_admin:
            return {"es_pendiente_validacion": False, "esta_validado": False, "is_admin": True}

        es_pendiente = False
        esta_validado = False
        if 'JEFE' in (current_user.cargo or '').upper():
            zonas_jefe = obtener_zonas_usuario(current_user.usuario)  # usa excel_cache
            if zonas_jefe:
                validaciones = [esta_validado_cartera(current_user.usuario, z) for z in zonas_jefe]
                esta_validado = all(validaciones)
                es_pendiente = not esta_validado
            else:
                es_pendiente = True

        return {
            "es_pendiente_validacion": es_pendiente,
            "esta_validado": esta_validado,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/info")
async def get_carteras_info(
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Obtiene información sobre el estado del módulo de carteras.
    
    Args:
        current_user: Usuario actual
        
    Returns:
        Información del módulo
    """
    base_file_exists = CARTERAS_BASE_FILE.exists()
    
    info: Dict[str, Any] = {
        "module": "carteras",
        "base_file_exists": base_file_exists,
        "base_file_path": str(CARTERAS_BASE_FILE) if base_file_exists else None,
    }
    
    if base_file_exists:
        try:
            file_stats = CARTERAS_BASE_FILE.stat()
            wb = openpyxl.load_workbook(CARTERAS_BASE_FILE, read_only=True, data_only=True)
            
            info.update({
                "file_size": file_stats.st_size,
                "last_modified": datetime.fromtimestamp(file_stats.st_mtime).isoformat(),
                "sheets": wb.sheetnames,
                "total_sheets": len(wb.sheetnames)
            })
            
            wb.close()
        except Exception as e:
            info["error"] = f"Error al leer archivo base: {str(e)}"
    
    return info


@router.get("/cargar", response_model=CarteraResponse)
async def cargar_carteras(
    current_user: UserInfo = Depends(get_current_user),
    zona: Optional[str] = Query(None)
):
    """
    Carga y retorna datos de carteras filtrados por zona del usuario.
    - ADMIN: Ve todos los clientes
    - Jefe de Venta: Ve solo clientes de sus distritos asignados
    - Si se provee el parámetro 'zona', filtra solo esa zona específica
    """
    try:
        if not CARTERAS_BASE_FILE.exists():
            raise HTTPException(
                status_code=404,
                detail=f"Maestro Cliente no encontrado. Por favor, sube el Maestro Cliente desde el panel de administración."
            )
        
        # Determinar si es admin
        is_admin = current_user.cargo.upper() == 'ADMIN'
        
        # Si no es admin, obtener sus zonas
        user_zonas = []
        if not is_admin:
            user_zonas = obtener_zonas_usuario(current_user.usuario)
            if not user_zonas:
                raise HTTPException(
                    status_code=403,
                    detail="No tienes distritos asignados. Contacta al administrador."
                )
        
        # Verificar estado de validación y determinar qué archivo leer (para jefes de venta)
        es_pendiente_validacion = False
        archivo_individual = None
        archivo_a_leer = CARTERAS_BASE_FILE
        
        if not is_admin:
            # Usar archivo de zona específica si hay zona seleccionada
            archivo_jefe = obtener_archivo_jefe_cartera(current_user.usuario, zona)
            
            # Si tiene archivo individual (de zona), leer ese; si no, usar el base
            if archivo_jefe.exists():
                archivo_individual = archivo_jefe.name
                archivo_a_leer = archivo_jefe
                print(f"✓ Cargando archivo individual de cartera para {current_user.usuario} zona={zona}: {archivo_jefe.name}")
            else:
                print(f"✓ Cargando archivo base de cartera para {current_user.usuario} zona={zona}")
            
            # Verificar si es jefe de venta y si ha validado la zona
            if current_user.cargo and 'JEFE' in current_user.cargo.upper():
                es_pendiente_validacion = not esta_validado_cartera(current_user.usuario, zona)
                print(f"🔍 Estado de validación para {current_user.usuario} zona={zona}: pendiente={es_pendiente_validacion}")
        
        # Leer archivo de carteras (individual o base según corresponda)
        try:
            df = excel_cache.read_excel(archivo_a_leer, sheet_name='Carga Admin')
        except ValueError as e:
            if 'Worksheet named' in str(e) and 'not found' in str(e):
                try:
                    df = excel_cache.read_excel(archivo_a_leer, sheet_name='Cartera')
                    print(f"✓ Archivo leído con hoja 'Cartera'")
                except ValueError:
                    # Leer la primera hoja disponible como fallback
                    try:
                        df = excel_cache.read_excel(archivo_a_leer, sheet_name=0)
                        wb = openpyxl.load_workbook(archivo_a_leer, read_only=True, data_only=True)
                        hoja_usada = wb.sheetnames[0]
                        wb.close()
                        print(f"⚠️  Hojas esperadas no encontradas, leyendo primera hoja: '{hoja_usada}'")
                    except Exception:
                        wb = openpyxl.load_workbook(archivo_a_leer, read_only=True, data_only=True)
                        hojas_disponibles = wb.sheetnames
                        wb.close()
                        raise HTTPException(
                            status_code=400,
                            detail=f"El archivo no contiene las hojas esperadas ('Carga Admin' o 'Cartera'). Hojas disponibles: {', '.join(hojas_disponibles)}"
                        )
            else:
                raise
        
        # Filtrar por zona(s)
        if not is_admin:
            zona_filter = [zona] if zona else user_zonas
            if zona_filter:
                df = df[df['CodDistrito'].isin(zona_filter)]
        
        # Verificar que hay datos
        if df.empty:
            return CarteraResponse(
                clientes=[],
                metadata={
                    "total_clientes": 0,
                    "total_vendedores": 0,
                    "total_gestores": 0,
                    "total_distritos": 0,
                    "user_role": current_user.cargo,
                    "user_zonas": user_zonas if not is_admin else [],
                    "zona_seleccionada": zona
                }
            )
        
        clientes_list = df.to_dict('records')
        
        # Calcular vendedores y gestores por separado usando cargo de Estructura de Venta
        total_vendedores_unicos = df['CodVend'].nunique()
        total_gestores = 0
        total_vendedores_sin_gestores = total_vendedores_unicos
        
        try:
            usuario_ev = current_user.usuario if not is_admin else None
            if usuario_ev:
                _, cargo_dict, _ = obtener_vendedores_con_cargo_estructura_venta(usuario_ev, zona_filtro=zona)
            else:
                # Para admin, intentar leer cargos del archivo base de EV
                archivo_base_ev = ESTRUCTURA_VENTA_DIR / "Respaldo" / "TBL EstructuraVentaCM.xlsx"
                cargo_dict = {}
                if archivo_base_ev.exists():
                    try:
                        df_ev = pd.read_excel(archivo_base_ev, sheet_name='Dotacion')
                        if 'CodVenta' in df_ev.columns and 'Cargo' in df_ev.columns:
                            for _, row in df_ev.iterrows():
                                if pd.notna(row['CodVenta']):
                                    cargo_dict[int(row['CodVenta'])] = str(row['Cargo']).strip() if pd.notna(row['Cargo']) else ''
                    except:
                        pass
            
            if cargo_dict:
                codigos_cartera = df['CodVend'].dropna().astype(int).unique()
                gestores = [c for c in codigos_cartera if 'GESTOR' in cargo_dict.get(c, '').upper()]
                total_gestores = len(gestores)
                total_vendedores_sin_gestores = total_vendedores_unicos - total_gestores
        except Exception as e:
            print(f"No se pudo calcular gestores: {str(e)}")
        
        metadata = {
            "total_clientes": len(df),
            "total_vendedores": total_vendedores_sin_gestores,
            "total_gestores": total_gestores,
            "total_distritos": df['CodDistrito'].nunique(),
            "total_comunas": df['Comuna'].nunique(),
            "user_role": current_user.cargo,
            "user_zonas": user_zonas if not is_admin else [],
            "zona_seleccionada": zona,
            "tipos_negocio": df['TipoNeg'].value_counts().to_dict(),
            "niveles_precio": df['NivPrecio'].value_counts().to_dict(),
            "filename": archivo_individual if archivo_individual else CARTERAS_BASE_FILE.name,
            "es_pendiente_validacion": es_pendiente_validacion,
            "esta_validado": esta_validado_cartera(current_user.usuario, zona) if not is_admin else False
        }
        
        return CarteraResponse(
            clientes=clientes_list,
            metadata=metadata
        )
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error al cargar carteras: {str(e)}"
        )


@router.get("/estado-validacion-zonas")
async def estado_validacion_zonas(current_user: UserInfo = Depends(get_current_user)):
    """
    Retorna el estado de validación de cartera por cada zona asignada al jefe.
    """
    try:
        if not (current_user.cargo and 'JEFE' in current_user.cargo.upper()):
            return {"estados": {}}
        
        zonas = obtener_zonas_usuario(current_user.usuario)
        estados = {}
        for z in zonas:
            archivo_estado = obtener_archivo_estado_validacion_cartera(current_user.usuario, z)
            if archivo_estado.exists():
                try:
                    with open(archivo_estado, 'r', encoding='utf-8') as f:
                        estado = json.load(f)
                    estados[z] = {
                        "validado": estado.get('validado', False),
                        "fecha_validacion": estado.get('fecha_validacion', None)
                    }
                except:
                    estados[z] = {"validado": False, "fecha_validacion": None}
            else:
                estados[z] = {"validado": False, "fecha_validacion": None}
        
        return {"estados": estados}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/")
async def get_carteras_root(
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Endpoint raíz del módulo de carteras.
    
    Args:
        current_user: Usuario actual
        
    Returns:
        Información básica del módulo
    """
    return {
        "module": "Carteras",
        "description": "Gestión y asignación de carteras de clientes",
        "user": current_user.usuario,
        "cargo": current_user.cargo
    }


@router.get("/descargar-formato")
async def descargar_formato(
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Descarga un archivo Excel vacío con solo las columnas para usar como plantilla.
    """
    try:
        print(f" Descargando formato vacío para {current_user.usuario}")
        
        # Definir las columnas del formato
        columnas = [
            'CodDistrito',
            'DesDistrito',
            'CodVend',
            'NombreVend',
            'CodCliente',
            'RutCliente',
            'RazonSocial',
            'TipoNeg',
            'Relev',
            'NivPrecio',
            'Direccion',
            'Comuna'
        ]
        
        # Crear DataFrame vacío con las columnas
        df = pd.DataFrame(columns=columnas)
        
        # Crear archivo temporal
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_file = CARTERAS_CAMBIOS_DIR / f"temp_formato_{current_user.usuario}_{timestamp}.xlsx"
        
        # Guardar Excel vacío
        df.to_excel(temp_file, sheet_name='Cartera', index=False, engine='openpyxl')
        
        print(f" Formato generado exitosamente: {temp_file.name}")
        
        # Descargar el archivo
        return FileResponse(
            path=str(temp_file),
            filename=f"Formato_Cartera_{datetime.now().strftime('%Y%m%d')}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except Exception as e:
        print(f" ERROR en descargar-formato: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error al descargar formato: {str(e)}"
        )


@router.post("/upload-jefe-venta")
async def upload_jefe_venta(
    file: UploadFile = File(...),
    zona: Optional[str] = Form(None),
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Permite a Jefes de Venta subir archivos de carteras con los campos del formato.
    El archivo se valida y se guarda en la carpeta de Cambios.
    
    Args:
        file: Archivo Excel con la información de carteras
        current_user: Usuario actual (Jefe de Venta o Admin)
        
    Returns:
        Información sobre el archivo subido y validación
    """
    # Verificar permisos
    cargo_upper = current_user.cargo.upper()
    if not (cargo_upper == "ADMIN" or "JEFE" in cargo_upper):
        raise HTTPException(
            status_code=403,
            detail="Solo usuarios ADMIN o Jefe de Venta pueden subir archivos"
        )
    
    # Validar que sea un archivo Excel
    if not file.filename:
        raise HTTPException(status_code=400, detail="No se proporcionó un archivo")
    
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        raise HTTPException(
            status_code=400,
            detail="El archivo debe ser un Excel (.xlsx o .xls)"
        )
    
    try:
        # Columnas requeridas del formato
        columnas_requeridas = [
            'CodDistrito',
            'DesDistrito',
            'CodVend',
            'NombreVend',
            'CodCliente',
            'RutCliente',
            'RazonSocial',
            'TipoNeg',
            'Relev',
            'NivPrecio',
            'Direccion',
            'Comuna'
        ]
        
        # Leer el archivo subido
        contents = await file.read()
        
        # Obtener la extensión real del archivo
        file_extension = '.xlsx' if file.filename.endswith('.xlsx') else '.xls'
        
        # ASEGURAR que existe la carpeta de Cambios ANTES de intentar guardar
        try:
            CARTERAS_CAMBIOS_DIR.mkdir(parents=True, exist_ok=True)
            print(f" Carpeta de cambios verificada/creada: {CARTERAS_CAMBIOS_DIR}")
        except Exception as mkdir_error:
            print(f" ERROR al crear carpeta: {str(mkdir_error)}")
            raise HTTPException(
                status_code=500,
                detail={
                    "message": "Error al crear carpeta de trabajo",
                    "validaciones": [
                        {
                            "nombre": "Carpeta de trabajo",
                            "estado": "failed",
                            "mensaje": f"No se pudo crear la carpeta para guardar archivos. Contacta al administrador del sistema."
                        },
                        {
                            "nombre": "Sugerencia",
                            "estado": "warning",
                            "mensaje": "Verifica que el servidor tenga permisos de escritura o contacta al administrador."
                        }
                    ]
                }
            )
        
        # Crear archivo temporal para validación con la extensión correcta
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_file = CARTERAS_CAMBIOS_DIR / f"temp_upload_{current_user.usuario}_{timestamp}{file_extension}"
        
        # Guardar el archivo
        try:
            with open(temp_file, 'wb') as f:
                f.write(contents)
            
            # Verificar que el archivo se guardó correctamente
            if not temp_file.exists():
                raise HTTPException(
                    status_code=500,
                    detail={
                        "message": "Error al guardar el archivo",
                        "validaciones": [
                            {
                                "nombre": "Guardado de archivo",
                                "estado": "failed",
                                "mensaje": "No se pudo guardar el archivo en el servidor. Inténtalo de nuevo o contacta al administrador."
                            }
                        ]
                    }
                )
            
            print(f" Archivo guardado temporalmente: {temp_file}")
            print(f"   Tamaño: {temp_file.stat().st_size} bytes")
            
        except IOError as io_error:
            raise HTTPException(
                status_code=500,
                detail={
                    "message": "Error al escribir el archivo",
                    "validaciones": [
                        {
                            "nombre": "Escritura de archivo",
                            "estado": "failed",
                            "mensaje": f"No se pudo escribir el archivo en el servidor. Inténtalo de nuevo."
                        }
                    ]
                }
            )
        
        # Validar estructura del archivo
        try:
            # Intentar leer el archivo con el motor apropiado
            try:
                print(f" Intentando leer archivo con motor: {'openpyxl' if file_extension == '.xlsx' else 'xlrd'}")
                if file_extension == '.xlsx':
                    df = pd.read_excel(temp_file, engine='openpyxl')
                else:
                    df = pd.read_excel(temp_file, engine='xlrd')
                print(f" Archivo leído exitosamente: {len(df)} filas, {len(df.columns)} columnas")
            except FileNotFoundError as fnf_error:
                # Archivo no encontrado - esto no debería pasar
                raise HTTPException(
                    status_code=500,
                    detail={
                        "message": "Error: archivo no encontrado",
                        "validaciones": [
                            {
                                "nombre": "Lectura de archivo",
                                "estado": "failed",
                                "mensaje": "El archivo no se encontró después de guardarlo. Esto es un error del sistema."
                            },
                            {
                                "nombre": "Sugerencia",
                                "estado": "warning",
                                "mensaje": "Por favor, contacta al administrador si el problema persiste."
                            }
                        ]
                    }
                )
            except Exception as read_error:
                # Error al leer el archivo - probablemente formato incorrecto
                temp_file.unlink(missing_ok=True)
                
                error_msg = str(read_error).lower()
                if 'xlrd' in error_msg or 'openpyxl' in error_msg or 'format' in error_msg:
                    raise HTTPException(
                        status_code=400,
                        detail={
                            "message": "Error en formato del archivo",
                            "validaciones": [
                                {
                                    "nombre": "Formato de archivo",
                                    "estado": "failed",
                                    "mensaje": "El archivo no se puede leer como Excel válido. Asegúrate de que sea un archivo .xlsx o .xls y que no esté dañado."
                                },
                                {
                                    "nombre": "Sugerencia",
                                    "estado": "warning",
                                    "mensaje": "Abre el archivo en Excel, verifica que se abra correctamente y guárdalo nuevamente como .xlsx"
                                }
                            ]
                        }
                    )
                else:
                    raise HTTPException(
                        status_code=400,
                        detail={
                            "message": f" Error al leer el archivo: {str(read_error)}",
                            "validaciones": [
                                {
                                    "nombre": "Lectura de archivo",
                                    "estado": "failed",
                                    "mensaje": f"No se pudo leer el archivo. Detalle técnico: {str(read_error)}"
                                }
                            ]
                        }
                    )
            
            # Validar que tenga columnas
            if df.empty:
                temp_file.unlink(missing_ok=True)
                raise HTTPException(
                    status_code=400,
                    detail={
                        "message": "Error: archivo vacío",
                        "validaciones": [
                            {
                                "nombre": "Contenido del archivo",
                                "estado": "failed",
                                "mensaje": "El archivo está vacío o no contiene datos. Verifica que el archivo tenga información de clientes."
                            }
                        ]
                    }
                )
            
            # Validar que tenga todas las columnas requeridas
            columnas_archivo = df.columns.tolist()
            columnas_faltantes = [col for col in columnas_requeridas if col not in columnas_archivo]
            
            if columnas_faltantes:
                temp_file.unlink(missing_ok=True)
                raise HTTPException(
                    status_code=400,
                    detail={
                        "message": "Error: faltan columnas en el archivo",
                        "validaciones": [
                            {
                                "nombre": "Columnas requeridas",
                                "estado": "failed",
                                "mensaje": f"El archivo no tiene todas las columnas necesarias. Faltan: {', '.join(columnas_faltantes)}"
                            },
                            {
                                "nombre": "Columnas encontradas",
                                "estado": "warning",
                                "mensaje": f"Columnas en tu archivo: {', '.join(columnas_archivo)}"
                            }
                        ]
                    }
                )
            
            # Validar que tenga datos
            if len(df) == 0:
                temp_file.unlink(missing_ok=True)
                raise HTTPException(
                    status_code=400,
                    detail={
                        "message": "Error: archivo sin datos",
                        "validaciones": [
                            {
                                "nombre": "Filas de datos",
                                "estado": "failed",
                                "mensaje": "El archivo no contiene filas de datos. Asegúrate de que tenga al menos un registro de cliente."
                            }
                        ]
                    }
                )
            
            # ==========================================
            # VALIDACIONES DE DATOS
            # ==========================================
            
            # Lista para almacenar advertencias (no bloquean la subida)
            advertencias = []
            
            # Lista para almacenar validaciones importantes
            validaciones_detalle = []
            
            # ==========================================
            # VALIDACIÓN CRÍTICA: VENDEDORES ELIMINADOS Y CLIENTES HUÉRFANOS
            # ==========================================
            print(f"\n🔍 Verificando vendedores eliminados y clientes huérfanos...")
            
            # Verificar si existe archivo anterior del jefe (por zona)
            archivo_anterior = obtener_archivo_jefe_cartera(current_user.usuario, zona)
            
            if archivo_anterior.exists():
                try:
                    # Leer archivo anterior
                    df_anterior = pd.read_excel(archivo_anterior, sheet_name='Cartera')
                    
                    # Obtener vendedores del archivo anterior y nuevo
                    vendedores_anteriores = set(df_anterior['CodVend'].dropna().astype(int).unique())
                    vendedores_nuevos = set(df['CodVend'].dropna().astype(int).unique())
                    
                    # Detectar vendedores eliminados
                    vendedores_eliminados = vendedores_anteriores - vendedores_nuevos
                    
                    if len(vendedores_eliminados) > 0:
                        print(f"⚠️  Se detectaron {len(vendedores_eliminados)} vendedor(es) eliminado(s): {sorted(vendedores_eliminados)}")
                        
                        # Por cada vendedor eliminado, verificar qué pasó con sus clientes
                        clientes_huerfanos = []
                        clientes_reasignados = []
                        clientes_eliminados = []
                        
                        for cod_vend in vendedores_eliminados:
                            # Obtener clientes de este vendedor en archivo anterior
                            clientes_vendedor_anterior = df_anterior[df_anterior['CodVend'] == cod_vend]
                            
                            for _, cliente_row in clientes_vendedor_anterior.iterrows():
                                cod_cliente = cliente_row['CodCliente']
                                razon_social = cliente_row.get('RazonSocial', f'Cliente {cod_cliente}')
                                
                                # Buscar si el cliente está en el nuevo archivo
                                cliente_nuevo = df[df['CodCliente'] == cod_cliente]
                                
                                if cliente_nuevo.empty:
                                    # Cliente fue eliminado completamente - OK (no es huérfano)
                                    clientes_eliminados.append({
                                        'codigo_cliente': str(cod_cliente),
                                        'razon_social': str(razon_social),
                                        'vendedor_anterior': int(cod_vend)
                                    })
                                else:
                                    # Cliente existe en nuevo archivo
                                    nuevo_vendedor = cliente_nuevo.iloc[0]['CodVend']
                                    
                                    if pd.isna(nuevo_vendedor) or nuevo_vendedor == '':
                                        # Cliente SIN vendedor asignado - HUÉRFANO (ERROR CRÍTICO)
                                        clientes_huerfanos.append({
                                            'codigo_cliente': str(cod_cliente),
                                            'razon_social': str(razon_social),
                                            'vendedor_anterior': int(cod_vend),
                                            'rut': str(cliente_nuevo.iloc[0].get('RutCliente', 'N/A')),
                                            'comuna': str(cliente_nuevo.iloc[0].get('Comuna', 'N/A'))
                                        })
                                    else:
                                        # Cliente reasignado a otro vendedor - OK
                                        clientes_reasignados.append({
                                            'codigo_cliente': str(cod_cliente),
                                            'razon_social': str(razon_social),
                                            'vendedor_anterior': int(cod_vend),
                                            'vendedor_nuevo': int(nuevo_vendedor)
                                        })
                        
                        print(f"   - Clientes huérfanos (SIN vendedor): {len(clientes_huerfanos)}")
                        print(f"   - Clientes reasignados: {len(clientes_reasignados)}")
                        print(f"   - Clientes eliminados: {len(clientes_eliminados)}")
                        
                        # Si hay clientes huérfanos - ERROR
                        if len(clientes_huerfanos) > 0:
                            
                            # Agrupar clientes huérfanos por vendedor eliminado
                            huerfanos_por_vendedor = {}
                            for cliente in clientes_huerfanos:
                                vend = cliente['vendedor_anterior']
                                if vend not in huerfanos_por_vendedor:
                                    huerfanos_por_vendedor[vend] = []
                                huerfanos_por_vendedor[vend].append(cliente)
                            
                            # Construir mensaje detallado
                            mensaje_detalle = []
                            for vend, clientes in sorted(huerfanos_por_vendedor.items()):
                                mensaje_detalle.append(f"Vendedor {vend}: {len(clientes)} cliente(s) sin reasignar")
                            
                            clientes_ejemplos = clientes_huerfanos[:5]
                            codigos_ejemplos = [c['codigo_cliente'] for c in clientes_ejemplos]
                            
                            validaciones_detalle.append({
                                "nombre": "Clientes sin vendedor asignado",
                                "estado": "failed",
                                "mensaje": f"Se eliminaron {len(vendedores_eliminados)} vendedor(es) y {len(clientes_huerfanos)} cliente(s) quedaron sin vendedor. Debes reasignarlos antes de continuar.",
                                "detalle_completo": clientes_huerfanos
                            })
                        
                        # Si NO hay clientes huérfanos - mostrar advertencia informativa
                        else:
                            print(f"✅ Todos los clientes de vendedores eliminados fueron correctamente reasignados o eliminados")
                            
                            # Construir mensaje informativo
                            mensaje_info = f"Se eliminaron {len(vendedores_eliminados)} vendedor(es): {', '.join(map(str, sorted(vendedores_eliminados)))}. "
                            
                            if len(clientes_reasignados) > 0:
                                mensaje_info += f"{len(clientes_reasignados)} cliente(s) fueron reasignados correctamente"
                            if len(clientes_eliminados) > 0:
                                if len(clientes_reasignados) > 0:
                                    mensaje_info += f" y "
                                mensaje_info += f"{len(clientes_eliminados)} cliente(s) fueron eliminados"
                            
                            validaciones_detalle.append({
                                "nombre": "Vendedores eliminados",
                                "estado": "warning",
                                "mensaje": mensaje_info,
                                "detalle_completo": {
                                    "vendedores_eliminados": sorted(list(vendedores_eliminados)),
                                    "clientes_reasignados": clientes_reasignados,
                                    "clientes_eliminados": clientes_eliminados
                                }
                            })
                            advertencias.append(mensaje_info)
                    else:
                        print(f"✅ No se eliminaron vendedores")
                
                except Exception as e:
                    print(f"⚠️  No se pudo verificar vendedores eliminados: {str(e)}")
                    # No bloqueamos la subida por error en esta validación
            else:
                print(f"ℹ️  Primera subida de cartera para {current_user.usuario} - no hay archivo anterior para comparar")
            
            # NO MOSTRAR estas validaciones cuando están OK:
            # - Formato de archivo Excel
            # - Columnas requeridas 
            # - Archivo contiene datos
            # - Campos opcionales cuando están completos
            
            # SÍ MOSTRAR siempre:
            # - CodVend obligatorio (OK o error)
            # - CodCliente obligatorio (OK o error)
            # - Sin clientes duplicados (OK o error)
            # - Validación contra archivo BASE (OK o error)
            # - Validación de zonas (OK o error)
            # - Campos opcionales solo cuando faltan (warning)
            
            # 1. OBLIGATORIO: Validar que todos los registros tengan código de vendedor
            registros_sin_codvend = df[df['CodVend'].isna() | (df['CodVend'] == '')]
            if len(registros_sin_codvend) > 0:
                filas_error = registros_sin_codvend.index.tolist()[:5]
                
                # Crear detalle estructurado
                detalle_sin_codvend = []
                for idx in registros_sin_codvend.index[:20]:  # Primeros 20
                    registro = df.iloc[idx]
                    detalle_sin_codvend.append({
                        "fila": int(idx + 2),  # +2 porque Excel empieza en 1 y tiene header
                        "codigo_cliente": str(registro.get('CodCliente', 'N/A')),
                        "razon_social": str(registro.get('RazonSocial', 'Sin nombre')),
                        "mensaje": "Falta código de vendedor"
                    })
                
                validaciones_detalle.append({
                    "nombre": "Código de vendedor",
                    "estado": "failed",
                    "mensaje": f"Hay {len(registros_sin_codvend)} registro(s) sin código de vendedor. Completa este campo obligatorio para continuar.",
                    "detalle_completo": detalle_sin_codvend
                })
            else:
                validaciones_detalle.append({
                    "nombre": "Código de vendedor",
                    "estado": "passed",
                    "mensaje": f"Todos los {len(df)} registros tienen código de vendedor"
                })
            
            # 2. OBLIGATORIO: Validar que todos los registros tengan código de cliente
            registros_sin_codcliente = df[df['CodCliente'].isna() | (df['CodCliente'] == '')]
            if len(registros_sin_codcliente) > 0:
                filas_error = registros_sin_codcliente.index.tolist()[:5]
                
                # Crear detalle estructurado
                detalle_sin_codcliente = []
                for idx in registros_sin_codcliente.index[:20]:
                    registro = df.iloc[idx]
                    detalle_sin_codcliente.append({
                        "fila": int(idx + 2),
                        "codigo_vendedor": str(registro.get('CodVend', 'N/A')),
                        "razon_social": str(registro.get('RazonSocial', 'Sin nombre')),
                        "mensaje": "Falta código de cliente"
                    })
                
                validaciones_detalle.append({
                    "nombre": "Código de cliente",
                    "estado": "failed",
                    "mensaje": f"Hay {len(registros_sin_codcliente)} registro(s) sin código de cliente. Completa este campo obligatorio para continuar.",
                    "detalle_completo": detalle_sin_codcliente
                })
            else:
                validaciones_detalle.append({
                    "nombre": "Código de cliente",
                    "estado": "passed",
                    "mensaje": f"Todos los {len(df)} registros tienen código de cliente"
                })
            
            # 3. Validar que no haya clientes duplicados
            clientes_duplicados = df[df.duplicated(subset=['CodCliente'], keep=False)]
            if len(clientes_duplicados) > 0:
                temp_file.unlink(missing_ok=True)
                codigos_duplicados = clientes_duplicados['CodCliente'].unique().tolist()[:5]
                
                # Crear detalle estructurado de duplicados
                detalle_duplicados = []
                for cod_cliente in clientes_duplicados['CodCliente'].unique():
                    registros_cliente = df[df['CodCliente'] == cod_cliente]
                    vendedores = registros_cliente['CodVend'].tolist()
                    razon_social = registros_cliente['RazonSocial'].iloc[0] if 'RazonSocial' in registros_cliente.columns and pd.notna(registros_cliente['RazonSocial'].iloc[0]) else 'N/A'
                    detalle_duplicados.append({
                        "codigo_cliente": str(cod_cliente),
                        "razon_social": str(razon_social),
                        "cantidad_registros": len(registros_cliente),
                        "vendedores": [int(v) for v in vendedores]
                    })
                
                validaciones_detalle.append({
                    "nombre": "Clientes duplicados",
                    "estado": "failed",
                    "mensaje": f"Se encontraron {len(clientes_duplicados)} registro(s) con clientes repetidos. Cada cliente debe aparecer solo una vez en el archivo.",
                    "detalle_completo": detalle_duplicados
                })
            else:
                validaciones_detalle.append({
                    "nombre": "Clientes duplicados",
                    "estado": "passed",
                    "mensaje": f"No hay clientes duplicados en los {len(df)} registros"
                })
            # Campos opcionales - generar advertencias si están vacíos
            campos_opcionales = {
                'RutCliente': 'RUT de cliente',
                'RazonSocial': 'Razón social',
                'NombreVend': 'Nombre de vendedor',
                'DesDistrito': 'Descripción de distrito',
                'TipoNeg': 'Tipo de negocio',
                'NivPrecio': 'Nivel de precio',
                'Direccion': 'Dirección',
                'Comuna': 'Comuna'
            }
            
            for campo, descripcion in campos_opcionales.items():
                if campo in df.columns:
                    registros_vacios = df[df[campo].isna() | (df[campo] == '')]
                    if len(registros_vacios) > 0:
                        mensaje_adv = f"{len(registros_vacios)} registro(s) sin {descripcion}"
                        advertencias.append(f" {mensaje_adv}")
                        # Solo agregar al desglose cuando HAY advertencia
                        validaciones_detalle.append({
                            "nombre": f"{campo} (opcional)",
                            "estado": "warning",
                            "mensaje": mensaje_adv
                        })
                    # No agregar si está completo - solo mostrar cuando falta
            
            # Validar campo Relev si existe
            if 'Relev' in df.columns:
                registros_sin_relev = df[df['Relev'].isna()]
                if len(registros_sin_relev) > 0:
                    mensaje_adv = f"{len(registros_sin_relev)} registro(s) sin Relevancia"
                    advertencias.append(f" {mensaje_adv}")
                    # Solo agregar cuando HAY advertencia
                    validaciones_detalle.append({
                        "nombre": "Relev (opcional)",
                        "estado": "warning",
                        "mensaje": mensaje_adv
                    })
                # No agregar si está completo
            
            if advertencias:
                print(f" Advertencias encontradas:")
                for adv in advertencias:
                    print(f"   {adv}")
            
            # ==========================================
            # 4. VALIDAR VENDEDORES ELIMINADOS CON CLIENTES ASIGNADOS
            # ==========================================
            print(f"\n🔍 Verificando vendedores eliminados con clientes asignados...")
            
            # Obtener el Estatus de los vendedores de Estructura de Venta (solo de la zona subida)
            estatus_vendedores, archivo_ev_estatus = obtener_estatus_vendedores_estructura_venta(current_user.usuario, zona_filtro=zona)
            
            if estatus_vendedores:
                # Obtener vendedores en Carteras
                vendedores_en_cartera = df['CodVend'].dropna().astype(int).unique()
                
                # Buscar vendedores que tienen clientes asignados PERO están eliminados en Estructura de Venta
                vendedores_eliminados_con_clientes = []
                
                for cod_vend in vendedores_en_cartera:
                    cod_vend_int = int(cod_vend)
                    estatus = estatus_vendedores.get(cod_vend_int, None)
                    
                    if estatus == 'Eliminado':
                        # Este vendedor tiene clientes pero está eliminado
                        clientes_vendedor = df[df['CodVend'] == cod_vend_int]
                        num_clientes = len(clientes_vendedor)
                        
                        # Obtener nombres de todos los clientes para el detalle
                        clientes_detalle = []
                        for _, cliente in clientes_vendedor.iterrows():
                            clientes_detalle.append({
                                'codigo_cliente': str(cliente['CodCliente']),
                                'razon_social': str(cliente.get('RazonSocial', f"Cliente {cliente['CodCliente']}")),
                                'comuna': str(cliente.get('Comuna', 'N/A')),
                                'vendedor_eliminado': int(cod_vend_int)
                            })
                        
                        vendedores_eliminados_con_clientes.append({
                            'codigo_vendedor': cod_vend_int,
                            'num_clientes': num_clientes,
                            'clientes': clientes_detalle
                        })
                
                if len(vendedores_eliminados_con_clientes) > 0:
                    # ERROR: Hay vendedores eliminados con clientes asignados
                    
                    total_clientes_huerfanos = sum(v['num_clientes'] for v in vendedores_eliminados_con_clientes)
                    codigos_vendedores_eliminados = [v['codigo_vendedor'] for v in vendedores_eliminados_con_clientes]
                    
                    # Crear detalle para el modal
                    detalle_clientes_huerfanos = []
                    for v in vendedores_eliminados_con_clientes:
                        for cliente in v['clientes']:
                            detalle_clientes_huerfanos.append(cliente)
                    
                    mensaje_error = f"Hay {len(vendedores_eliminados_con_clientes)} vendedor(es) eliminados en la Estructura de Venta, pero {total_clientes_huerfanos} cliente(s) siguen asignados a ellos. Reasigna estos clientes a vendedores activos."
                    
                    validaciones_detalle.append({
                        "nombre": "Clientes de vendedores eliminados",
                        "estado": "failed",
                        "mensaje": mensaje_error,
                        "detalle_completo": detalle_clientes_huerfanos
                    })
                else:
                    print(f"✓ No hay vendedores eliminados con clientes asignados")
            else:
                print(f"⚠ No se pudo obtener el Estatus de vendedores de Estructura de Venta")
            
            # ==========================================
            # 5. VALIDAR VENDEDORES DE ESTRUCTURA DE VENTA
            # ==========================================
            # Los vendedores de Estructura de Venta (sin reemplazos) de los distritos del usuario deben estar 100% en Carteras
            print(f"\n Iniciando validación de vendedores de Estructura de Venta...")
            
            # Obtener zonas del usuario para mostrar en mensajes
            user_zonas = obtener_zonas_usuario(current_user.usuario)
            
            # USAR LA NUEVA FUNCIÓN que incluye información de cargo (filtrada por la zona subida)
            vendedores_ev, vendedores_cargo_dict, archivo_ev = obtener_vendedores_con_cargo_estructura_venta(current_user.usuario, zona_filtro=zona)
            
            if vendedores_ev and archivo_ev:
                # Obtener vendedores del archivo de Carteras
                vendedores_cartera = set(df['CodVend'].dropna().astype(int).unique())
                vendedores_ev_set = set(vendedores_ev)
                
                # Verificar que TODOS los vendedores de EV estén en Carteras
                vendedores_faltantes = vendedores_ev_set - vendedores_cartera
                
                if len(vendedores_faltantes) > 0:
                    # CLASIFICAR vendedores faltantes en: reemplazo, jefes, y normales
                    vendedores_reemplazo_faltantes = []
                    vendedores_jefe_faltantes = []
                    vendedores_normales_faltantes = []
                    
                    for cod in vendedores_faltantes:
                        cargo_original = vendedores_cargo_dict.get(cod, '')
                        cargo = cargo_original.upper().strip()
                        
                        print(f"       Vendedor faltante {cod}:")
                        print(f"         • Cargo ORIGINAL: '{cargo_original}'")
                        print(f"         • Cargo NORMALIZADO: '{cargo}'")
                        print(f"         • Tipo: {type(cargo_original)}")
                        print(f"         • Longitud: {len(cargo_original)} caracteres")
                        print(f"         • Bytes: {cargo_original.encode('utf-8')}")
                        
                        # Verificar si es REEMPLAZO
                        if 'REEMPLAZO' in cargo or 'REEMPLAZ' in cargo:
                            vendedores_reemplazo_faltantes.append(cod)
                            print(f"          Clasificado como REEMPLAZO")
                        # Verificar si es JEFE DE VENTA
                        elif 'JEFE' in cargo:
                            vendedores_jefe_faltantes.append(cod)
                            print(f"          Clasificado como JEFE DE VENTA")
                        else:
                            vendedores_normales_faltantes.append(cod)
                            print(f"          Clasificado como NORMAL (no es reemplazo ni jefe)")
                    
                    print(f" Vendedores faltantes: {len(vendedores_faltantes)} total")
                    print(f"   - Reemplazo: {len(vendedores_reemplazo_faltantes)}")
                    print(f"   - Jefes de Venta: {len(vendedores_jefe_faltantes)}")
                    print(f"   - Normales: {len(vendedores_normales_faltantes)}")
                    
                    # Si HAY vendedores NORMALES faltantes -> ERROR CRÍTICO
                    if len(vendedores_normales_faltantes) > 0:
                        vendedores_error = sorted(vendedores_normales_faltantes)[:10]
                        
                        # Obtener nombres de los vendedores faltantes desde archivo EV
                        detalle_vendedores_faltantes = []
                        try:
                            df_ev = pd.read_excel(archivo_ev, sheet_name='Dotacion')
                            for cod in vendedores_normales_faltantes:
                                vendedor_data = df_ev[df_ev['CodVenta'] == cod]
                                if not vendedor_data.empty:
                                    nombre_parts = []
                                    if 'Nombre' in df_ev.columns and pd.notna(vendedor_data.iloc[0]['Nombre']):
                                        nombre_parts.append(str(vendedor_data.iloc[0]['Nombre']))
                                    if 'APaterno' in df_ev.columns and pd.notna(vendedor_data.iloc[0]['APaterno']):
                                        nombre_parts.append(str(vendedor_data.iloc[0]['APaterno']))
                                    if 'AMaterno' in df_ev.columns and pd.notna(vendedor_data.iloc[0]['AMaterno']):
                                        nombre_parts.append(str(vendedor_data.iloc[0]['AMaterno']))
                                    
                                    nombre = ' '.join(nombre_parts) if nombre_parts else f"Vendedor {cod}"
                                    cargo = vendedor_data.iloc[0].get('Cargo', 'TITULAR')
                                else:
                                    nombre = f"Vendedor {cod}"
                                    cargo = 'TITULAR'
                                
                                detalle_vendedores_faltantes.append({
                                    "codigo": int(cod),
                                    "nombre": str(nombre),
                                    "cargo": str(cargo)
                                })
                        except:
                            # Si hay error, usar info básica
                            for cod in vendedores_normales_faltantes:
                                detalle_vendedores_faltantes.append({
                                    "codigo": int(cod),
                                    "nombre": f"Vendedor {cod}",
                                    "cargo": "TITULAR"
                                })
                        
                        mensaje_error = f"{len(vendedores_normales_faltantes)} vendedor(es) de tu Estructura de Venta no tienen cartera asignada. Todos los vendedores activos deben tener al menos un cliente."
                        
                        validaciones_detalle.append({
                            "nombre": "Vendedores de Estructura de Venta",
                            "estado": "failed",
                            "mensaje": mensaje_error,
                            "detalle_completo": detalle_vendedores_faltantes
                        })
                    
                    # Si SOLO faltan vendedores de REEMPLAZO y/o JEFES -> VALIDACIÓN OK con advertencia
                    else:
                        print(f" Validación OK: Solo faltan vendedores de reemplazo y/o jefes de venta")
                        
                        # Obtener nombres de los vendedores de reemplazo faltantes desde archivo EV
                        detalle_reemplazos = []
                        
                        try:
                            df_ev = pd.read_excel(archivo_ev, sheet_name='Dotacion')
                            
                            # Procesar vendedores de reemplazo
                            for cod in vendedores_reemplazo_faltantes:
                                vendedor_data = df_ev[df_ev['CodVenta'] == cod]
                                if not vendedor_data.empty:
                                    nombre_parts = []
                                    if 'Nombre' in df_ev.columns and pd.notna(vendedor_data.iloc[0]['Nombre']):
                                        nombre_parts.append(str(vendedor_data.iloc[0]['Nombre']))
                                    if 'APaterno' in df_ev.columns and pd.notna(vendedor_data.iloc[0]['APaterno']):
                                        nombre_parts.append(str(vendedor_data.iloc[0]['APaterno']))
                                    if 'AMaterno' in df_ev.columns and pd.notna(vendedor_data.iloc[0]['AMaterno']):
                                        nombre_parts.append(str(vendedor_data.iloc[0]['AMaterno']))
                                    
                                    nombre = ' '.join(nombre_parts) if nombre_parts else f"Vendedor {cod}"
                                    cargo = vendedor_data.iloc[0].get('Cargo', 'REEMPLAZO')
                                else:
                                    nombre = f"Vendedor {cod}"
                                    cargo = 'REEMPLAZO'
                                
                                detalle_reemplazos.append({
                                    "codigo": int(cod),
                                    "nombre": str(nombre),
                                    "cargo": str(cargo)
                                })
                        except:
                            # Si hay error, usar info básica
                            for cod in vendedores_reemplazo_faltantes:
                                detalle_reemplazos.append({
                                    "codigo": int(cod),
                                    "nombre": f"Vendedor {cod}",
                                    "cargo": "REEMPLAZO"
                                })
                        
                        # Calcular cuántos vendedores REALES están en carteras
                        vendedores_en_carteras = len(vendedores_ev) - len(vendedores_reemplazo_faltantes) - len(vendedores_jefe_faltantes)
                        
                        mensaje_ok = f"Los {vendedores_en_carteras} vendedores activos tienen cartera asignada"
                        
                        validaciones_detalle.append({
                            "nombre": "Vendedores de Estructura de Venta",
                            "estado": "passed",
                            "mensaje": mensaje_ok
                        })
                        
                        # Agregar advertencia de reemplazos si hay
                        if len(vendedores_reemplazo_faltantes) > 0:
                            mensaje_advertencia_reemplazos = f"{len(vendedores_reemplazo_faltantes)} vendedor(es) de reemplazo no tienen cartera asignada. Esto es normal y no requiere acción."
                            validaciones_detalle.append({
                                "nombre": "Vendedores de Reemplazo",
                                "estado": "warning",
                                "mensaje": mensaje_advertencia_reemplazos,
                                "detalle_completo": detalle_reemplazos
                            })
                            advertencias.append(mensaje_advertencia_reemplazos)
                
                else:
                    # Validación OK - Todos los vendedores están presentes
                    print(f" Validación de vendedores: Los {len(vendedores_ev)} vendedores están en Carteras")
                    validaciones_detalle.append({
                        "nombre": "Vendedores de Estructura de Venta",
                        "estado": "passed",
                        "mensaje": f"Los {len(vendedores_ev)} vendedores activos tienen cartera asignada"
                    })
            else:
                # No se encontró archivo de Estructura de Venta o no se pudieron leer vendedores
                print(f"  No se encontró archivo de Estructura de Venta para {current_user.usuario} o el archivo no contiene vendedores")
                validaciones_detalle.append({
                    "nombre": "Vendedores de Estructura de Venta",
                    "estado": "warning",
                    "mensaje": "No se pudo verificar contra la Estructura de Venta porque no se encontró el archivo. Se recomienda subir o revisar la Estructura de Venta."
                })
            
            # ==========================================
            # 6. VALIDAR QUE VENDEDORES NORMALES TENGAN AL MENOS 1 CLIENTE
            # ==========================================
            print(f"\n🔍 Validando asignación de clientes a vendedores...")
            
            # Solo hacer esta validación si se pudieron obtener vendedores de Estructura de Venta
            if vendedores_ev and archivo_ev:
                # Obtener cantidad de clientes por vendedor en archivo Carteras
                clientes_por_vendedor = df.groupby('CodVend').size().to_dict()
                
                # Obtener vendedores del archivo Carteras
                vendedores_cartera = set(df['CodVend'].dropna().astype(int).unique())
                
                # Clasificar vendedores en: NORMALES con 0 clientes, REEMPLAZO con 0 clientes, REEMPLAZO con clientes
                vendedores_normales_sin_clientes = []
                vendedores_reemplazo_sin_clientes = []
                vendedores_reemplazo_con_clientes = []
                
                for cod_vend in vendedores_cartera:
                    # Obtener cantidad de clientes
                    num_clientes = clientes_por_vendedor.get(cod_vend, 0)
                    
                    # Obtener cargo del vendedor
                    cargo_original = vendedores_cargo_dict.get(cod_vend, '')
                    cargo = cargo_original.upper().strip()
                    
                    # Verificar si es REEMPLAZO
                    es_reemplazo = 'REEMPLAZO' in cargo or 'REEMPLAZ' in cargo
                    # Verificar si es JEFE
                    es_jefe = 'JEFE' in cargo
                    
                    # Saltar jefes (ya validado en otra sección)
                    if es_jefe:
                        continue
                    
                    if num_clientes == 0:
                        # Vendedor sin clientes
                        if es_reemplazo:
                            vendedores_reemplazo_sin_clientes.append({
                                "codigo": int(cod_vend),
                                "cargo": str(cargo_original),
                                "clientes": 0
                            })
                        else:
                            vendedores_normales_sin_clientes.append({
                                "codigo": int(cod_vend),
                                "cargo": str(cargo_original),
                                "clientes": 0
                            })
                    else:
                        # Vendedor con clientes
                        if es_reemplazo:
                            vendedores_reemplazo_con_clientes.append({
                                "codigo": int(cod_vend),
                                "cargo": str(cargo_original),
                                "clientes": int(num_clientes)
                            })
                
                print(f"📊 Resultado de asignación de clientes:")
                print(f"   - Vendedores NORMALES sin clientes: {len(vendedores_normales_sin_clientes)}")
                print(f"   - Vendedores REEMPLAZO sin clientes: {len(vendedores_reemplazo_sin_clientes)}")
                print(f"   - Vendedores REEMPLAZO con clientes: {len(vendedores_reemplazo_con_clientes)}")
                
                # ========== ERROR CRÍTICO: Vendedores NORMALES sin clientes ==========
                if len(vendedores_normales_sin_clientes) > 0:
                    
                    codigos_sin_clientes = [v['codigo'] for v in vendedores_normales_sin_clientes]
                    codigos_ejemplos = codigos_sin_clientes[:10]
                    
                    # Obtener nombres desde Estructura de Venta si es posible
                    try:
                        df_ev = pd.read_excel(archivo_ev, sheet_name='Dotacion')
                        for v in vendedores_normales_sin_clientes:
                            vendedor_data = df_ev[df_ev['CodVenta'] == v['codigo']]
                            if not vendedor_data.empty:
                                nombre_parts = []
                                if 'Nombre' in df_ev.columns and pd.notna(vendedor_data.iloc[0]['Nombre']):
                                    nombre_parts.append(str(vendedor_data.iloc[0]['Nombre']))
                                if 'APaterno' in df_ev.columns and pd.notna(vendedor_data.iloc[0]['APaterno']):
                                    nombre_parts.append(str(vendedor_data.iloc[0]['APaterno']))
                                v['nombre'] = ' '.join(nombre_parts) if nombre_parts else f"Vendedor {v['codigo']}"
                            else:
                                v['nombre'] = f"Vendedor {v['codigo']}"
                    except:
                        for v in vendedores_normales_sin_clientes:
                            v['nombre'] = f"Vendedor {v['codigo']}"
                    
                    mensaje_error = f"{len(vendedores_normales_sin_clientes)} vendedor(es) no tienen ningún cliente asignado. Cada vendedor activo debe tener al menos un cliente."
                    
                    validaciones_detalle.append({
                        "nombre": "Vendedores sin clientes",
                        "estado": "failed",
                        "mensaje": mensaje_error,
                        "detalle_completo": vendedores_normales_sin_clientes
                    })
                
                # ========== ADVERTENCIA: Vendedores REEMPLAZO sin clientes ==========
                if len(vendedores_reemplazo_sin_clientes) > 0:
                    # Obtener nombres desde Estructura de Venta si es posible
                    try:
                        df_ev = pd.read_excel(archivo_ev, sheet_name='Dotacion')
                        for v in vendedores_reemplazo_sin_clientes:
                            vendedor_data = df_ev[df_ev['CodVenta'] == v['codigo']]
                            if not vendedor_data.empty:
                                nombre_parts = []
                                if 'Nombre' in df_ev.columns and pd.notna(vendedor_data.iloc[0]['Nombre']):
                                    nombre_parts.append(str(vendedor_data.iloc[0]['Nombre']))
                                if 'APaterno' in df_ev.columns and pd.notna(vendedor_data.iloc[0]['APaterno']):
                                    nombre_parts.append(str(vendedor_data.iloc[0]['APaterno']))
                                v['nombre'] = ' '.join(nombre_parts) if nombre_parts else f"Vendedor {v['codigo']}"
                            else:
                                v['nombre'] = f"Vendedor {v['codigo']}"
                    except:
                        for v in vendedores_reemplazo_sin_clientes:
                            v['nombre'] = f"Vendedor {v['codigo']}"
                    
                    codigos_reemplazo = [v['codigo'] for v in vendedores_reemplazo_sin_clientes]
                    mensaje_warning = f"{len(vendedores_reemplazo_sin_clientes)} vendedor(es) de reemplazo no tienen clientes asignados. Esto es normal."
                    
                    validaciones_detalle.append({
                        "nombre": "Vendedores de Reemplazo sin clientes",
                        "estado": "warning",
                        "mensaje": mensaje_warning,
                        "detalle_completo": vendedores_reemplazo_sin_clientes
                    })
                    advertencias.append(mensaje_warning)
                    print(f"⚠️  {mensaje_warning}")
                
                # ========== ADVERTENCIA: Vendedores REEMPLAZO con clientes ==========
                if len(vendedores_reemplazo_con_clientes) > 0:
                    # Obtener nombres desde Estructura de Venta si es posible
                    try:
                        df_ev = pd.read_excel(archivo_ev, sheet_name='Dotacion')
                        for v in vendedores_reemplazo_con_clientes:
                            vendedor_data = df_ev[df_ev['CodVenta'] == v['codigo']]
                            if not vendedor_data.empty:
                                nombre_parts = []
                                if 'Nombre' in df_ev.columns and pd.notna(vendedor_data.iloc[0]['Nombre']):
                                    nombre_parts.append(str(vendedor_data.iloc[0]['Nombre']))
                                if 'APaterno' in df_ev.columns and pd.notna(vendedor_data.iloc[0]['APaterno']):
                                    nombre_parts.append(str(vendedor_data.iloc[0]['APaterno']))
                                v['nombre'] = ' '.join(nombre_parts) if nombre_parts else f"Vendedor {v['codigo']}"
                            else:
                                v['nombre'] = f"Vendedor {v['codigo']}"
                    except:
                        for v in vendedores_reemplazo_con_clientes:
                            v['nombre'] = f"Vendedor {v['codigo']}"
                    
                    codigos_con_clientes = [f"{v['codigo']} ({v['clientes']} clientes)" for v in vendedores_reemplazo_con_clientes[:5]]
                    total_clientes_reemplazo = sum(v['clientes'] for v in vendedores_reemplazo_con_clientes)
                    
                    mensaje_warning = f"{len(vendedores_reemplazo_con_clientes)} vendedor(es) de reemplazo tienen clientes asignados ({total_clientes_reemplazo} clientes en total). Verifica si esto es correcto."
                    
                    validaciones_detalle.append({
                        "nombre": "Vendedores de Reemplazo con clientes",
                        "estado": "warning",
                        "mensaje": mensaje_warning,
                        "detalle_completo": vendedores_reemplazo_con_clientes
                    })
                    advertencias.append(mensaje_warning)
                    print(f"⚠️  {mensaje_warning}")
                
                # Si no hay problemas, mensaje de éxito
                if len(vendedores_normales_sin_clientes) == 0 and len(vendedores_reemplazo_sin_clientes) == 0 and len(vendedores_reemplazo_con_clientes) == 0:
                    print(f"✅ Todos los vendedores normales tienen clientes asignados correctamente")
            else:
                print(f"ℹ️  No se pudo validar asignación de clientes (no hay archivo de Estructura de Venta)")
            
            # ==========================================
            # 7. VALIDAR QUE NO HAYA CLIENTES ASIGNADOS A JEFES DE VENTA
            # ==========================================
            print(f"\n🔍 Validando que no haya clientes asignados a Jefes de Venta...")
            codigos_jefes = obtener_codigos_jefes_venta()
            info_jefes = obtener_info_jefes_venta()
            
            if codigos_jefes:
                # Obtener vendedores del archivo de Carteras
                vendedores_en_carteras = df['CodVend'].dropna().astype(int).unique()
                
                # Verificar si hay Jefes con clientes asignados
                jefes_con_clientes = [j for j in codigos_jefes if j in vendedores_en_carteras]
                
                if len(jefes_con_clientes) > 0:
                    
                    # Obtener información COMPLETA de TODOS los jefes y sus clientes
                    detalle_jefes_completo = []
                    resumen_jefes = []
                    total_clientes_afectados = 0
                    
                    for cod_jefe in jefes_con_clientes:
                        # Obtener TODOS los clientes de este jefe
                        clientes_jefe = df[df['CodVend'] == cod_jefe]
                        cantidad = len(clientes_jefe)
                        total_clientes_afectados += cantidad
                        
                        # Nombre del jefe
                        nombre_jefe = info_jefes.get(cod_jefe, f"Jefe {cod_jefe}")
                        
                        # Resumen para mensaje corto
                        resumen_jefes.append(f"{nombre_jefe} (Cód: {cod_jefe}) - {cantidad} clientes")
                        
                        # Lista completa de clientes
                        clientes_list = []
                        for _, cliente in clientes_jefe.iterrows():
                            clientes_list.append({
                                "codigo_cliente": str(cliente.get('CodCliente', 'N/A')),
                                "razon_social": str(cliente.get('RazonSocial', 'Sin nombre')),
                                "rut": str(cliente.get('RutCliente', 'N/A')),
                                "comuna": str(cliente.get('Comuna', 'N/A'))
                            })
                        
                        detalle_jefes_completo.append({
                            "codigo_jefe": int(cod_jefe),
                            "nombre_jefe": str(nombre_jefe),
                            "cantidad_clientes": int(cantidad),
                            "clientes": clientes_list
                        })
                    
                    advertencias.append(f"Se encontraron {len(jefes_con_clientes)} Jefe(s) con {total_clientes_afectados} clientes asignados. Se recomienda reasignar estos clientes a vendedores.")
                    validaciones_detalle.append({
                        "nombre": "Clientes asignados a Jefes",
                        "estado": "warning",
                        "mensaje": f"Se encontraron {len(jefes_con_clientes)} Jefe(s) de Venta con {total_clientes_afectados} clientes asignados. Se recomienda reasignar estos clientes a vendedores.",
                        "detalle_completo": detalle_jefes_completo
                    })
                    print(f"⚠️ Advertencia: {len(jefes_con_clientes)} Jefe(s) con clientes asignados. Continuando como advertencia...")
                
                else:
                    # Validación OK
                    print(f" Validación de Jefes: Ningún Jefe tiene clientes asignados")
                    validaciones_detalle.append({
                        "nombre": "Clientes asignados a Jefes",
                        "estado": "passed",
                        "mensaje": "Ningún Jefe de Venta tiene clientes asignados directamente"
                    })
            else:
                # No se encontraron Jefes - solo informativo, no es error
                print(f"  No se identificaron Jefes de Venta en el sistema")
                validaciones_detalle.append({
                    "nombre": "Clientes asignados a Jefes",
                    "estado": "warning",
                    "mensaje": "No se encontró información de Jefes de Venta en el sistema para realizar esta validación."
                })
            
            # ==========================================
            # 8. VALIDAR CONTRA ARCHIVO BASE DEL ADMIN
            # ==========================================
            # Verificar que todos los clientes del archivo existan en el archivo BASE
            if not CARTERAS_BASE_FILE.exists():
                temp_file.unlink(missing_ok=True)
                raise HTTPException(
                    status_code=400,
                    detail={
                        "message": "Maestro Cliente no encontrado",
                        "validaciones": [
                            {
                                "nombre": "Maestro Cliente",
                                "estado": "failed",
                                "mensaje": "No existe el Maestro Cliente en el sistema. El administrador debe subirlo primero desde el panel de administración."
                            }
                        ]
                    }
                )
            
            # Leer archivo BASE del administrador
            try:
                df_base = excel_cache.read_excel(CARTERAS_BASE_FILE, sheet_name='Carga Admin')
                
                # SI NO ES ADMIN: Filtrar el Maestro por zona
                if cargo_upper != "ADMIN":
                    user_zonas = obtener_zonas_usuario(current_user.usuario)
                    if not user_zonas:
                        temp_file.unlink(missing_ok=True)
                        raise HTTPException(
                            status_code=403,
                            detail={
                                "message": "Sin zonas asignadas",
                                "validaciones": [
                                    {
                                        "nombre": "Zonas del usuario",
                                        "estado": "failed",
                                        "mensaje": "No tienes zonas asignadas en el sistema. Contacta al administrador para que te asigne zonas."
                                    }
                                ]
                            }
                        )
                    
                    # Si se sube para una zona específica, filtrar solo por esa zona;
                    # si no, usar todas las zonas del usuario
                    zonas_maestro = [zona] if zona else user_zonas
                    df_base_filtrado = df_base[df_base['CodDistrito'].isin(zonas_maestro)]
                    print(f" Maestro completo: {len(df_base)} clientes | Maestro filtrado para zonas {zonas_maestro}: {len(df_base_filtrado)} clientes")
                else:
                    # Si es ADMIN, usar todo el Maestro
                    df_base_filtrado = df_base
                    zona_maestro_label = zona if zona else "todas"
                    print(f" Usuario ADMIN - Maestro completo: {len(df_base)} clientes (zona={zona_maestro_label})")
                
                # Obtener los códigos de cliente del archivo base FILTRADO
                clientes_base = set(df_base_filtrado['CodCliente'].astype(str).str.strip())
                
                # Obtener los códigos de cliente del archivo subido
                clientes_subidos = set(df['CodCliente'].astype(str).str.strip())
                
                # Validación exacta: el archivo debe tener EXACTAMENTE los mismos CodCliente que el Maestro
                # (ni más ni menos)
                clientes_no_existen = clientes_subidos - clientes_base   # están en cartera pero NO en Maestro
                clientes_faltantes  = clientes_base - clientes_subidos   # están en Maestro pero NO en cartera
                hay_error = False

                # --- Error 1: clientes que no existen en el Maestro ---
                if len(clientes_no_existen) > 0:
                    hay_error = True
                    clientes_error = list(clientes_no_existen)[:10]
                    detalle_no_existentes = []
                    for cod_cliente in clientes_no_existen:
                        cliente_data = df[df['CodCliente'].astype(str).str.strip() == cod_cliente]
                        if not cliente_data.empty:
                            razon_social = cliente_data['RazonSocial'].iloc[0] if 'RazonSocial' in cliente_data.columns and pd.notna(cliente_data['RazonSocial'].iloc[0]) else 'Sin nombre'
                            cod_vend = cliente_data['CodVend'].iloc[0] if 'CodVend' in cliente_data.columns and pd.notna(cliente_data['CodVend'].iloc[0]) else 'N/A'
                            detalle_no_existentes.append({
                                "codigo_cliente": str(cod_cliente),
                                "razon_social": str(razon_social),
                                "cod_vendedor": str(cod_vend),
                                "mensaje": "Este cliente no existe en el Maestro Cliente"
                            })
                    validaciones_detalle.append({
                        "nombre": "Clientes no registrados",
                        "estado": "failed",
                        "mensaje": f"{len(clientes_no_existen)} cliente(s) de tu archivo no están registrados en el Maestro Cliente. Solo puedes incluir clientes que existan en el Maestro.",
                        "detalle_completo": detalle_no_existentes
                    })
                    print(f"❌ {len(clientes_no_existen)} clientes no existen en el Maestro")

                # --- Error 2: clientes del Maestro que faltan en la cartera ---
                if len(clientes_faltantes) > 0:
                    hay_error = True
                    clientes_ejemplos = list(clientes_faltantes)[:10]
                    detalle_faltantes = []
                    for cod_cliente in list(clientes_faltantes)[:20]:
                        cliente_info = df_base_filtrado[df_base_filtrado['CodCliente'].astype(str).str.strip() == cod_cliente]
                        if not cliente_info.empty:
                            detalle_faltantes.append({
                                "codigo_cliente": str(cod_cliente),
                                "razon_social": str(cliente_info.iloc[0].get('RazonSocial', 'Sin nombre')),
                                "distrito": str(cliente_info.iloc[0].get('CodDistrito', 'N/A')),
                                "mensaje": "Cliente del Maestro no incluido en el archivo"
                            })
                    if cargo_upper == "ADMIN":
                        msg_faltantes = f"Faltan {len(clientes_faltantes)} cliente(s) del Maestro Cliente que no están en tu archivo. Todos los clientes del Maestro deben estar incluidos."
                    else:
                        msg_faltantes = f"Faltan {len(clientes_faltantes)} cliente(s) del Maestro Cliente (zonas {zonas_maestro}) que no están en tu archivo. Todos los clientes de tus zonas deben estar incluidos."
                    validaciones_detalle.append({
                        "nombre": "Clientes faltantes del Maestro",
                        "estado": "failed",
                        "mensaje": msg_faltantes,
                        "detalle_completo": detalle_faltantes
                    })
                    print(f"❌ {len(clientes_faltantes)} clientes del Maestro faltan en el archivo")

                # --- Éxito: coincidencia exacta ---
                if not hay_error:
                    if cargo_upper == "ADMIN":
                        mensaje_validacion = f"Los {len(clientes_subidos)} clientes coinciden con el Maestro Cliente"
                    else:
                        mensaje_validacion = f"Los {len(clientes_subidos)} clientes coinciden con el Maestro Cliente de tus zonas ({', '.join(zonas_maestro)})"
                    validaciones_detalle.append({
                        "nombre": "Validación contra Maestro Cliente",
                        "estado": "passed",
                        "mensaje": mensaje_validacion
                    })
                    print(f"✅ Validación contra Maestro Cliente: coincidencia exacta ({len(clientes_subidos)} clientes)")
                
            except KeyError as e:
                temp_file.unlink(missing_ok=True)
                raise HTTPException(
                    status_code=400,
                    detail={
                        "message": "Error en formato del archivo",
                        "validaciones": [
                            {
                                "nombre": "Columnas del archivo",
                                "estado": "failed",
                                "mensaje": f"El archivo no tiene el formato correcto. Falta la columna: {str(e)}"
                            }
                        ]
                    }
                )
            except Exception as e:
                temp_file.unlink(missing_ok=True)
                raise HTTPException(
                    status_code=500,
                    detail={
                        "message": "Error inesperado al leer Maestro Cliente",
                        "validaciones": [
                            {
                                "nombre": "Error del sistema",
                                "estado": "failed",
                                "mensaje": "Ocurrió un error inesperado al procesar el Maestro Cliente. Inténtalo de nuevo o contacta al administrador."
                            }
                        ]
                    }
                )
            
            # VALIDACIÓN ADICIONAL: Verificar que los distritos del archivo coincidan con las zonas del usuario
            if cargo_upper != "ADMIN":
                # Ya obtuvimos user_zonas en la validación anterior
                user_zonas = obtener_zonas_usuario(current_user.usuario)
                
                # Verificar que todos los distritos del archivo pertenezcan al usuario
                distritos_archivo = df['CodDistrito'].unique().tolist()
                distritos_invalidos = [d for d in distritos_archivo if d not in user_zonas]
                
                if distritos_invalidos:
                    temp_file.unlink(missing_ok=True)
                    validaciones_detalle.append({
                        "nombre": "Validación de zonas asignadas",
                        "estado": "failed",
                        "mensaje": f"El archivo contiene distritos que no están asignados a tu usuario: {', '.join(map(str, distritos_invalidos))}. Solo puedes subir datos de tus zonas."
                    })
                    raise HTTPException(
                        status_code=403,
                        detail={
                            "message": f"El archivo contiene distritos no asignados a tu usuario: {', '.join(map(str, distritos_invalidos))}",
                            "validaciones": validaciones_detalle
                        }
                    )
                # SÍ mostrar esta validación cuando pasa
                validaciones_detalle.append({
                    "nombre": "Validación de zonas asignadas",
                    "estado": "passed",
                    "mensaje": "Todos los distritos del archivo pertenecen a tus zonas asignadas"
                })
            else:
                # Para ADMIN también mostrar
                validaciones_detalle.append({
                    "nombre": "Validación de zonas asignadas",
                    "estado": "passed",
                    "mensaje": "Usuario administrador, sin restricción de zonas"
                })
            
            # ==========================================
            # VERIFICACIÓN FINAL: Si hay errores, no guardar el archivo
            # ==========================================
            errores_encontrados = [v for v in validaciones_detalle if v['estado'] == 'failed']
            if errores_encontrados:
                temp_file.unlink(missing_ok=True)
                raise HTTPException(
                    status_code=400,
                    detail={
                        "message": f"Se encontraron {len(errores_encontrados)} error(es) en el archivo. Corrige los problemas indicados y vuelve a intentarlo.",
                        "validaciones": validaciones_detalle
                    }
                )

            # Guardar con nombre permanente
            final_filename = f"Cartera_JefeVenta_{current_user.usuario}_{timestamp}.xlsx"
            final_file = CARTERAS_CAMBIOS_DIR / final_filename
            shutil.move(str(temp_file), str(final_file))
            
            # IMPORTANTE: También guardar/copiar con nombre estándar para el sistema (por zona)
            archivo_estandar = obtener_archivo_jefe_cartera(current_user.usuario, zona)
            shutil.copy2(str(final_file), str(archivo_estandar))
            print(f"✓ Archivo copiado a: {archivo_estandar.name} (zona={zona})")
            
            # NOTA: NO marcar como validado automáticamente
            # El usuario debe revisar los resultados y decidir si validar presionando "Validar Cartera"
            
            # Calcular estadísticas del archivo
            total_clientes = len(df)
            total_vendedores = df['CodVend'].nunique()
            total_distritos = df['CodDistrito'].nunique()
            
            print(f"✓ Archivo subido exitosamente por {current_user.usuario}: {final_filename}")
            print(f"   Total clientes: {total_clientes}")
            print(f"   Total vendedores: {total_vendedores}")
            print(f"   Total distritos: {total_distritos}")
            
            mensaje = "Archivo subido exitosamente. Revisa las validaciones y confirma con 'Validar Cartera'"
            if advertencias:
                mensaje += f" (con {len(advertencias)} advertencia(s))"
            
            return {
                "success": True,
                "message": mensaje,
                "filename": final_filename,
                "stats": {
                    "total_clientes": total_clientes,
                    "total_vendedores": total_vendedores,
                    "total_distritos": total_distritos,
                    "distritos": df['CodDistrito'].unique().tolist()
                },
                "warnings": advertencias if advertencias else [],
                "validaciones": validaciones_detalle,  # Desglose completo de validaciones
                "upload_date": datetime.now().isoformat(),
                "uploaded_by": current_user.usuario
            }
            
        except pd.errors.EmptyDataError:
            if temp_file.exists():
                temp_file.unlink(missing_ok=True)
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "Error: archivo vacío o inválido",
                    "validaciones": [
                        {
                            "nombre": "Contenido del archivo",
                            "estado": "failed",
                            "mensaje": "El archivo Excel está vacío o tiene un formato inválido. Verifica que el archivo tenga datos."
                        }
                    ]
                }
            )
        except Exception as e:
            if temp_file.exists():
                temp_file.unlink(missing_ok=True)
            raise
        
    except HTTPException:
        raise
    except Exception as e:
        print(f" ERROR en upload-jefe-venta: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail={
                "message": "Error inesperado al procesar el archivo",
                "validaciones": [
                    {
                        "nombre": "Error del sistema",
                        "estado": "failed",
                        "mensaje": f"Ocurrió un error inesperado. Inténtalo de nuevo."
                    },
                    {
                        "nombre": "Sugerencia",
                        "estado": "warning",
                        "mensaje": "Si el problema persiste, contacta al administrador del sistema."
                    }
                ]
            }
        )


@router.post("/descargar")
async def descargar_cartera(
    clientes_data: dict,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Descarga el archivo Excel con los clientes que el usuario ve en la tabla.
    Recibe los datos filtrados desde el frontend.
    """
    try:
        # Extraer los clientes del body
        clientes = clientes_data.get('clientes', [])
        
        if not clientes:
            raise HTTPException(
                status_code=404,
                detail="No hay clientes para descargar"
            )
        
        print(f"📥 Descargando cartera para {current_user.usuario}")
        print(f"   Total clientes a exportar: {len(clientes)}")
        
        # Crear DataFrame con los datos recibidos
        df = pd.DataFrame(clientes)
        
        # Crear archivo temporal
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_file = CARTERAS_CAMBIOS_DIR / f"temp_cartera_{current_user.usuario}_{timestamp}.xlsx"
        
        # Guardar Excel
        df.to_excel(temp_file, sheet_name='Cartera', index=False, engine='openpyxl')
        
        print(f" Archivo generado exitosamente: {temp_file.name}")
        
        # Descargar el archivo
        return FileResponse(
            path=str(temp_file),
            filename=f"Cartera_Clientes_{current_user.usuario}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except HTTPException:
        raise
    except Exception as e:
        print(f" ERROR en descargar-cartera: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error al descargar archivo: {str(e)}"
        )


# ==========================================
# ENDPOINTS DE VALIDACIÓN (migrados desde carteras_validacion.py)
# ==========================================

@router.post("/re-validar")
async def re_validar_cartera_actual(
    request: ValidarZonaCarterasRequest,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Valida la cartera ya guardada sin necesidad de subir un nuevo archivo.
    Corre exactamente las mismas validaciones que el upload y, si pasan,
    marca la cartera como validada automáticamente.
    """
    import io as _io

    zona = request.zona if request else None

    cargo_upper = current_user.cargo.upper()
    if not (cargo_upper == "ADMIN" or "JEFE" in cargo_upper):
        raise HTTPException(status_code=403, detail="Solo usuarios ADMIN o Jefe de Venta pueden validar carteras")

    # Localizar el archivo guardado del jefe para esta zona
    archivo_jefe = obtener_archivo_jefe_cartera(current_user.usuario, zona)

    if not archivo_jefe.exists():
        # Intentar con el archivo base
        if not CARTERAS_BASE_FILE.exists():
            raise HTTPException(
                status_code=404,
                detail="No hay archivo de cartera guardado para esta zona. Sube un archivo primero."
            )
        # Copiar / filtrar el archivo base
        try:
            df_base = excel_cache.read_excel(CARTERAS_BASE_FILE, sheet_name='Carga Admin')
        except (ValueError, KeyError):
            df_base = excel_cache.read_excel(CARTERAS_BASE_FILE, sheet_name=0)

        df_filtered = df_base[df_base['CodDistrito'] == zona] if zona else df_base
        df_filtered.to_excel(archivo_jefe, index=False, engine='openpyxl')
        print(f"📄 Archivo base copiado/filtrado para re-validar: {archivo_jefe}")

    # Leer el archivo y crear un UploadFile virtual
    with open(archivo_jefe, 'rb') as f:
        file_bytes = f.read()

    mock_file = UploadFile(
        filename=archivo_jefe.name,
        file=_io.BytesIO(file_bytes)
    )

    # Llamar al handler de upload (corre TODAS las validaciones).
    # Si hay errores de negocio, lanzará HTTPException que el cliente recibirá igual que en upload.
    result = await upload_jefe_venta(mock_file, zona, current_user)

    # Si llegamos aquí, todas las validaciones pasaron.
    # NO marcamos como validado aquí; el usuario debe confirmar con el botón "Validar Cartera".
    return {
        "success": True,
        "auto_validado": False,
        "message": f"Validación de cartera zona {zona} completada. Revisa los resultados y confirma.",
        "stats": result.get("stats", {}),
        "validaciones": result.get("validaciones", []),
        "warnings": result.get("warnings", []),
        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }


@router.post("/validar-sin-cambios")
async def validar_sin_cambios(
    request: ValidarZonaCarterasRequest,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Permite a un jefe validar su cartera sin hacer cambios (por zona).
    Crea una copia del archivo base filtrada para la zona del jefe.
    """
    try:
        zona = request.zona if request else None
        
        # Solo jefes de venta pueden validar
        if not current_user.cargo or 'JEFE' not in current_user.cargo.upper():
            raise HTTPException(
                status_code=403,
                detail="Solo jefes de venta pueden validar carteras"
            )
        
        archivo_jefe = obtener_archivo_jefe_cartera(current_user.usuario, zona)
        
        # Verificar si ya fue validado explícitamente (para esta zona)
        archivo_estado = obtener_archivo_estado_validacion_cartera(current_user.usuario, zona)
        if archivo_estado.exists():
            try:
                with open(archivo_estado, 'r', encoding='utf-8') as f:
                    estado = json.load(f)
                    return {
                        "success": True,
                        "message": f"Cartera zona {zona} ya validada previamente",
                        "usuario": current_user.usuario,
                        "zona": zona,
                        "fecha": estado.get('fecha_validacion', 'fecha desconocida')
                    }
            except:
                pass
        
        # Si no existe archivo individual de zona, copiar el base filtrado
        if not archivo_jefe.exists():
            if not CARTERAS_BASE_FILE.exists():
                raise HTTPException(status_code=404, detail="Archivo base de carteras no encontrado")
            
            # Si hay zona, filtrar el archivo base por esa zona
            if zona:
                try:
                    df_base = excel_cache.read_excel(CARTERAS_BASE_FILE, sheet_name='Carga Admin')
                except ValueError:
                    df_base = excel_cache.read_excel(CARTERAS_BASE_FILE, sheet_name=0)
                df_zona = df_base[df_base['CodDistrito'] == zona]
                df_zona.to_excel(archivo_jefe, sheet_name='Cartera', index=False, engine='openpyxl')
                print(f"📄 Archivo zona {zona} creado para {current_user.usuario}: {archivo_jefe}")
            else:
                shutil.copy2(CARTERAS_BASE_FILE, archivo_jefe)
                print(f"📄 Archivo creado para {current_user.usuario}: {archivo_jefe}")
        else:
            print(f"📄 Archivo ya existe para {current_user.usuario} zona={zona}, validando el existente")
        
        # Marcar explícitamente como validado (para esta zona)
        marcar_como_validado_cartera(current_user.usuario, zona)
        
        return {
            "success": True,
            "message": f"Cartera zona {zona} validada exitosamente",
            "usuario": current_user.usuario,
            "zona": zona,
            "archivo": archivo_jefe.name,
            "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al validar cartera: {str(e)}"
        )


@router.post("/limpiar-validaciones")
async def limpiar_validaciones(current_user: UserInfo = Depends(get_current_user)):
    """
    Elimina todos los archivos de la carpeta Cambios y backups antiguos de Respaldo
    para iniciar un nuevo ciclo. Solo accesible para administradores.
    """
    try:
        # Verificar que sea admin
        if current_user.cargo.upper() != 'ADMIN':
            raise HTTPException(status_code=403, detail="Solo administradores pueden limpiar validaciones")
        
        # Limpiar TODOS los archivos de Cambios
        archivos_eliminados = limpiar_carpeta_cambios_carteras()
        
        # Limpiar backups antiguos de Respaldo (conserva el archivo base)
        backups_eliminados = limpiar_backups_respaldo_carteras()
        archivos_eliminados.extend(backups_eliminados)
        
        return {
            "success": True,
            "message": f"Se eliminaron {len(archivos_eliminados)} archivo(s) de carteras",
            "archivos_eliminados": len(archivos_eliminados)
        }
    
    except HTTPException:
        raise
    except Exception as e:
        print(f"ERROR en limpiar-validaciones: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al limpiar validaciones: {str(e)}"
        )


@router.get("/descargar-consolidado")
async def descargar_consolidado(current_user: UserInfo = Depends(get_current_user)):
    """
    Descarga el último archivo consolidado.
    Solo accesible para administradores.
    """
    try:
        # Verificar que sea admin
        if current_user.cargo.upper() != 'ADMIN':
            raise HTTPException(status_code=403, detail="Solo administradores pueden descargar consolidado")
        
        # Buscar el archivo consolidado más reciente
        archivos_consolidados = list(CARTERAS_CAMBIOS_DIR.glob("TBL_Cartera_consolidado_*.xlsx"))
        
        if not archivos_consolidados:
            raise HTTPException(status_code=404, detail="No hay archivo consolidado disponible. Debe consolidar primero.")
        
        # Obtener el más reciente
        archivo_consolidado = max(archivos_consolidados, key=lambda p: p.stat().st_mtime)
        
        return FileResponse(
            path=str(archivo_consolidado),
            filename=f"Cartera_Consolidada_{datetime.now().strftime('%Y%m%d')}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except HTTPException:
        raise
    except Exception as e:
        print(f"ERROR en descargar-consolidado: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al descargar archivo consolidado: {str(e)}"
        )
