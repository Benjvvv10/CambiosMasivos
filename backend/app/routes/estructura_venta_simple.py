"""
Rutas simplificadas para gestión de estructura de ventas
Usa el archivo Excel base como fuente única de datos
"""

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from fastapi.responses import FileResponse
from typing import List, Dict, Any, Optional, Union
from pydantic import BaseModel
from app.utils.dependencies import get_current_user
from app.models.user import UserInfo
import pandas as pd
import numpy as np
import os
import json
from datetime import datetime
from pathlib import Path
import openpyxl
import shutil
from app.utils.excel_cache import excel_cache

router = APIRouter()


def _sanitize_for_json(obj):
    """Convierte tipos numpy/pandas a tipos nativos de Python para serialización JSON."""
    if isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize_for_json(v) for v in obj]
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating,)):
        if np.isnan(obj):
            return None
        return float(obj)
    if isinstance(obj, np.bool_):
        return bool(obj)
    if isinstance(obj, (np.ndarray,)):
        return obj.tolist()
    if pd.isna(obj):
        return None
    return obj


class CambioVendedor(BaseModel):
    """Representa un cambio en un vendedor"""
    codVend: str
    nombreVendedor: str
    field: str
    oldValue: Optional[Union[str, int, float, bool]] = None
    newValue: Optional[Union[str, int, float, bool]] = None


class GuardarCambiosRequest(BaseModel):
    """Request para guardar cambios"""
    cambios: List[CambioVendedor]
    editedData: Dict[str, Dict[str, Any]]


class ValidarZonaRequest(BaseModel):
    """Request para validar una zona específica"""
    zona: str


class SheetInfo(BaseModel):
    """Información de una hoja del Excel"""
    model_config = {"arbitrary_types_allowed": True}
    
    name: str
    rows: int
    columns: List[str]
    data: List[Dict[str, Any]]


class EstructuraVentaResponse(BaseModel):
    """Respuesta con toda la estructura de ventas"""
    model_config = {"arbitrary_types_allowed": True}
    
    sheets: List[SheetInfo]
    metadata: Dict[str, Any]


# Rutas de archivos
EXCEL_ORIGINAL_PATH = Path(__file__).parent.parent.parent / "data" / "carga_cm" / "EstructuraVenta" / "Respaldo" / "TBL EstructuraVentaCM.xlsx"
CAMBIOS_DIR = Path(__file__).parent.parent.parent / "data" / "carga_cm" / "EstructuraVenta" / "Cambios"
USUARIOS_SISTEMA_PATH = Path(__file__).parent.parent.parent / "data" / "usuarios_sistema.xlsx"

# Rutas legacy (para endpoints antiguos que aún las usan)
EXCEL_BASE_PATH = EXCEL_ORIGINAL_PATH  # Alias para compatibilidad

# Directorio Respaldo de Estructura
ESTRUCTURA_RESPALDO_DIR = EXCEL_ORIGINAL_PATH.parent

# Asegurar que existe la carpeta Cambios
CAMBIOS_DIR.mkdir(parents=True, exist_ok=True)


def limpiar_carpeta_cambios_estructura() -> list[str]:
    """Elimina TODOS los archivos de la carpeta Cambios de Estructura de Venta."""
    eliminados = []
    for archivo in CAMBIOS_DIR.iterdir():
        if archivo.is_file():
            try:
                archivo.unlink()
                eliminados.append(archivo.name)
                print(f"  Eliminado: {archivo.name}")
            except Exception as e:
                print(f"  Error al eliminar {archivo.name}: {str(e)}")
    return eliminados


def limpiar_backups_respaldo_estructura() -> list[str]:
    """Elimina archivos de backup antiguos en Respaldo, conservando el archivo base actual."""
    eliminados = []
    for archivo in ESTRUCTURA_RESPALDO_DIR.iterdir():
        if archivo.is_file() and archivo != EXCEL_ORIGINAL_PATH:
            try:
                archivo.unlink()
                eliminados.append(archivo.name)
                print(f"  Eliminado backup: {archivo.name}")
            except Exception as e:
                print(f"  Error al eliminar backup {archivo.name}: {str(e)}")
    return eliminados


def validar_vendedor_nuevo(datos: Dict[str, Any]) -> tuple[bool, list[str]]:
    """
    Valida que un vendedor nuevo tenga todos los campos obligatorios
    Retorna (es_valido, lista_de_errores)
    """
    campos_obligatorios = {
        'DesOficina': 'Oficina',
        'CodVenta': 'Código de Venta',
        'Cargo': 'Cargo',
        'Rut': 'RUT',
        'Nombre': 'Nombre',
        'APaterno': 'Apellido Paterno',
        'AMaterno': 'Apellido Materno',
        'Telefono': 'Teléfono',
        'Correo': 'Correo',
        'Genero': 'Género',
        'TallaPantalon': 'Talla Pantalón',
        'TallaCamisa': 'Talla Camisa'
    }
    
    errores = []
    
    # Verificar que todos los campos obligatorios estén presentes y no vacíos
    for campo, nombre_campo in campos_obligatorios.items():
        valor = datos.get(campo)
        if valor is None or valor == '' or (isinstance(valor, str) and valor.strip() == ''):
            errores.append(f"{nombre_campo} es obligatorio")
    
    # Validaciones de formato si los campos están presentes
    if 'Correo' in datos and datos['Correo']:
        correo = datos['Correo']
        if not correo.endswith('@cial.cl'):
            errores.append("El correo debe terminar en @cial.cl")
    
    if 'Telefono' in datos and datos['Telefono']:
        telefono = str(datos['Telefono'])
        if not telefono.startswith('9') or len(telefono) != 9:
            errores.append("El teléfono debe comenzar con 9 y tener 9 dígitos")
    
    if 'Cargo' in datos and datos['Cargo']:
        cargos_validos = ['TITULAR', 'RESIDENTE', 'GESTOR', 'REEMPLAZO', 
                         'VENDEDOR TITULAR', 'VENDEDOR RESIDENTE', 
                         'GESTOR SUPERMERCADO', 'VENDEDOR REEMPLAZO']
        if datos['Cargo'] not in cargos_validos:
            errores.append(f"Cargo inválido: {datos['Cargo']}")
    
    return (len(errores) == 0, errores)


def obtener_archivo_jefe(usuario: str) -> Path:
    """Obtiene la ruta del archivo individual del jefe"""
    return CAMBIOS_DIR / f"TBL_EstructuraVenta_{usuario}.xlsx"


def obtener_archivo_estado_validacion(usuario: str, zona: Optional[str] = None) -> Path:
    """Obtiene la ruta del archivo de estado de validación del jefe.
    Si se proporciona zona, crea un archivo específico por zona."""
    if zona:
        return CAMBIOS_DIR / f"VALIDADO_{usuario}_{zona}.json"
    return CAMBIOS_DIR / f"VALIDADO_{usuario}.json"


def marcar_como_validado(usuario: str, zona: Optional[str] = None) -> None:
    """Marca explícitamente que el jefe validó su estructura desde Estructura de Venta.
    Si se proporciona zona, marca solo esa zona como validada."""
    archivo_estado = obtener_archivo_estado_validacion(usuario, zona)
    estado = {
        "validado": True,
        "usuario": usuario,
        "zona": zona,
        "fecha_validacion": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "origen": "Estructura de Venta"
    }
    with open(archivo_estado, 'w', encoding='utf-8') as f:
        json.dump(estado, f, indent=2, ensure_ascii=False)
    zona_texto = f" - Zona: {zona}" if zona else ""
    print(f"✓ Validación marcada para {usuario}{zona_texto}")


def verificar_validacion_jefe_por_archivo(usuario: str, zona: Optional[str] = None) -> bool:
    """Verifica si un jefe ha validado EXPLÍCITAMENTE su estructura desde Estructura de Venta.
    Si se proporciona zona, verifica solo esa zona."""
    # Verificar si existe el archivo de estado de validación
    archivo_estado = obtener_archivo_estado_validacion(usuario, zona)
    validado = archivo_estado.exists()
    
    zona_texto = f" - Zona: {zona}" if zona else ""
    
    if validado:
        try:
            with open(archivo_estado, 'r', encoding='utf-8') as f:
                estado = json.load(f)
                print(f" Verificando {usuario}{zona_texto}: VALIDADO desde {estado.get('origen', 'desconocido')} el {estado.get('fecha_validacion', 'fecha desconocida')}")
        except:
            print(f" Verificando {usuario}{zona_texto}: VALIDADO (archivo estado existe)")
    else:
        print(f" Verificando {usuario}{zona_texto}: NO VALIDADO - Pendiente de validación")
    
    return validado


def limpiar_estado_validacion(usuario: str, zona: Optional[str] = None) -> None:
    """Elimina el estado de validación de un jefe.
    Si se proporciona zona, elimina solo el estado de esa zona."""
    archivo_estado = obtener_archivo_estado_validacion(usuario, zona)
    if archivo_estado.exists():
        archivo_estado.unlink()
        zona_texto = f" - Zona: {zona}" if zona else ""
        print(f"✓ Estado de validación eliminado para {usuario}{zona_texto}")


def convertir_valor_segun_tipo(valor: Any, dtype) -> Any:
    """
    Convierte un valor al tipo correcto según el dtype de pandas
    Maneja conversiones de string a int/float y valores vacíos
    """
    # Si el valor es None, vacío o NA, retornar pd.NA
    if pd.isna(valor) or valor == '' or valor is None:
        return pd.NA
    
    try:
        if pd.api.types.is_integer_dtype(dtype):
            # Convertir a integer
            return int(float(valor))
        elif pd.api.types.is_float_dtype(dtype):
            # Convertir a float
            return float(valor)
        else:
            # Para strings y otros tipos, retornar como está
            return valor
    except (ValueError, TypeError):
        # Si hay error de conversión, retornar pd.NA
        return pd.NA


def obtener_zonas_usuario(usuario: str) -> List[str]:
    """
    Obtiene las zonas asignadas a un usuario desde usuarios_sistema.xlsx
    Las zonas son las columnas después de 'Bloquear' que tienen valor 1
    """
    try:
        if not USUARIOS_SISTEMA_PATH.exists():
            print(f"DEBUG obtener_zonas: Archivo no existe: {USUARIOS_SISTEMA_PATH}")
            return []
        
        df_usuarios = excel_cache.read_excel(USUARIOS_SISTEMA_PATH, sheet_name=0)
        
        print(f"DEBUG obtener_zonas: Buscando usuario '{usuario}'")
        print(f"DEBUG obtener_zonas: Columnas disponibles: {df_usuarios.columns.tolist()}")
        
        # Buscar el usuario
        user_row = df_usuarios[df_usuarios['Usuario'].str.lower() == usuario.lower()]
        
        if user_row.empty:
            print(f"DEBUG obtener_zonas: Usuario '{usuario}' no encontrado")
            print(f"DEBUG obtener_zonas: Usuarios disponibles: {df_usuarios['Usuario'].tolist()}")
            return []
        
        print(f"DEBUG obtener_zonas: Usuario encontrado")
        
        # Obtener columnas de zonas (después de 'Bloquear')
        zonas_columns = df_usuarios.columns[7:]  # Columnas desde índice 7 en adelante son zonas
        
        print(f"DEBUG obtener_zonas: Columnas de zonas: {zonas_columns.tolist()}")
        
        # Filtrar zonas donde el valor es 1
        user_data = user_row.iloc[0]
        zonas_asignadas = [col for col in zonas_columns if pd.notna(user_data.get(col)) and user_data.get(col) == 1]
        
        print(f"DEBUG obtener_zonas: Zonas asignadas a {usuario}: {zonas_asignadas}")
        
        return zonas_asignadas
    
    except Exception as e:
        print(f"Error al obtener zonas del usuario: {str(e)}")
        import traceback
        traceback.print_exc()
        return []


def obtener_nombres_zonas(codigos_zonas: List[str]) -> List[dict]:
    """
    Obtiene los nombres completos de las zonas desde el archivo de estructura de ventas
    Retorna una lista de diccionarios con 'codigo' y 'nombre'
    """
    try:
        if not EXCEL_ORIGINAL_PATH.exists():
            return [{"codigo": cod, "nombre": cod} for cod in codigos_zonas]
        
        # Usar el archivo base original
        archivo_excel = EXCEL_ORIGINAL_PATH
        
        # Leer la hoja Dotacion
        df = excel_cache.read_excel(archivo_excel, sheet_name='Dotacion')
        
        zonas_con_nombres = []
        for codigo in codigos_zonas:
            # Buscar el nombre de la zona en el dataframe
            zona_data = df[df['CodDistrito'] == codigo]
            if not zona_data.empty and 'DesDistrito' in df.columns:
                nombre = zona_data.iloc[0]['DesDistrito']
                zonas_con_nombres.append({"codigo": codigo, "nombre": nombre})
            else:
                # Si no se encuentra, usar solo el código
                zonas_con_nombres.append({"codigo": codigo, "nombre": codigo})
        
        return zonas_con_nombres
    
    except Exception as e:
        print(f"Error al obtener nombres de zonas: {str(e)}")
        # En caso de error, retornar solo los códigos
        return [{"codigo": cod, "nombre": cod} for cod in codigos_zonas]


@router.get("/status")
async def obtener_status_estructura_venta(current_user: UserInfo = Depends(get_current_user)):
    """
    Endpoint liviano para obtener solo el estado de validación del usuario.
    No lee el Excel completo — solo verifica archivos JSON y existencia de archivos.
    """
    try:
        is_admin = current_user.cargo.upper() in ('ADMIN', 'ADMINISTRADOR')
        if is_admin:
            return {"es_pendiente_validacion": False, "esta_validado": False, "tiene_archivo": False, "is_admin": True}

        archivo_jefe = obtener_archivo_jefe(current_user.usuario)
        tiene_archivo = archivo_jefe.exists()
        filename = archivo_jefe.name if tiene_archivo else EXCEL_ORIGINAL_PATH.name

        es_pendiente = False
        esta_validado = False
        if 'JEFE' in (current_user.cargo or '').upper():
            zonas_jefe = obtener_zonas_usuario(current_user.usuario)  # usa excel_cache
            if zonas_jefe:
                validaciones = [verificar_validacion_jefe_por_archivo(current_user.usuario, z) for z in zonas_jefe]
                esta_validado = all(validaciones)
                es_pendiente = not esta_validado
            else:
                es_pendiente = True

        return {
            "es_pendiente_validacion": es_pendiente,
            "esta_validado": esta_validado,
            "tiene_archivo": tiene_archivo,
            "filename": filename,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/estado-validacion-zonas")
async def obtener_estado_validacion_zonas(current_user: UserInfo = Depends(get_current_user)):
    """
    Obtiene el estado de validación de cada zona asignada al usuario actual.
    Retorna un diccionario con el código de zona como key y el estado de validación.
    """
    try:
        # Obtener zonas del usuario
        zonas = obtener_zonas_usuario(current_user.usuario)
        
        estados = {}
        for zona in zonas:
            # Verificar si esta zona ha sido validada
            validado = verificar_validacion_jefe_por_archivo(current_user.usuario, zona)
            
            fecha_validacion = None
            if validado:
                archivo_estado = obtener_archivo_estado_validacion(current_user.usuario, zona)
                try:
                    with open(archivo_estado, 'r', encoding='utf-8') as f:
                        estado = json.load(f)
                        fecha_validacion = estado.get('fecha_validacion')
                except:
                    pass
            
            estados[zona] = {
                "validado": validado,
                "fecha_validacion": fecha_validacion
            }
        
        return {
            "estados": estados
        }
    
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al obtener estado de validación: {str(e)}"
        )


@router.get("/cargar")
async def cargar_estructura_venta(current_user: UserInfo = Depends(get_current_user)):
    """
    Carga y retorna datos del archivo Excel.
    - ADMIN: Ve el archivo base (Respaldo)
    - Jefe de Venta: Ve su archivo individual si existe, sino el base filtrado por sus zonas
    """
    try:
        # Determinar qué archivo cargar
        is_admin = current_user.cargo.upper() == 'ADMIN'
        
        if is_admin:
            # Admin siempre ve el archivo base
            archivo_a_cargar = EXCEL_ORIGINAL_PATH
        else:
            # Jefe: buscar su archivo individual
            archivo_jefe = obtener_archivo_jefe(current_user.usuario)
            if archivo_jefe.exists():
                archivo_a_cargar = archivo_jefe
                print(f" Cargando archivo individual de {current_user.usuario}: {archivo_jefe}")
            else:
                # Si no existe, cargar el base
                archivo_a_cargar = EXCEL_ORIGINAL_PATH
                print(f" Cargando archivo base para {current_user.usuario} (aún no validó)")
        
        if not archivo_a_cargar.exists():
            raise HTTPException(
                status_code=404,
                detail=f"Archivo no encontrado: {archivo_a_cargar}"
            )
        
        # Si no es admin, obtener sus zonas desde usuarios_sistema.xlsx
        user_zonas = []
        if not is_admin:
            user_zonas = obtener_zonas_usuario(current_user.usuario)
        
        # Cargar datos de la hoja Dotacion
        df_dotacion = excel_cache.read_excel(archivo_a_cargar, sheet_name='Dotacion')
        
        # Asegurarse de que existe la columna Estatus y que todos los registros tengan un valor
        if 'Estatus' not in df_dotacion.columns:
            df_dotacion['Estatus'] = 'Antiguo'
        else:
            # Si la columna existe pero tiene valores nulos, asignar 'Antiguo'
            df_dotacion['Estatus'] = df_dotacion['Estatus'].fillna('Antiguo')
            # Si tiene valores vacíos, reemplazarlos por 'Antiguo'
            df_dotacion.loc[df_dotacion['Estatus'] == '', 'Estatus'] = 'Antiguo'
        
        # *** DETECCIÓN DE VENDEDORES ELIMINADOS ***
        # Si es un jefe con archivo individual, comparar con BASE para detectar eliminaciones
        if not is_admin and archivo_jefe.exists() and archivo_a_cargar == archivo_jefe:
            # Cargar archivo BASE para comparación
            if EXCEL_ORIGINAL_PATH.exists():
                df_base = excel_cache.read_excel(EXCEL_ORIGINAL_PATH, sheet_name='Dotacion')
                
                # Filtrar BASE por las zonas del jefe
                if len(user_zonas) > 0 and 'CodDistrito' in df_base.columns:
                    df_base_jefe = df_base[df_base['CodDistrito'].isin(user_zonas)]
                    
                    # Obtener códigos de vendedores en BASE y en archivo actual
                    # Incluir TODOS los vendedores del actual (incluso los que ya están marcados como Eliminado)
                    vendedores_base = set(df_base_jefe['CodVenta'].dropna().astype(str))
                    vendedores_actuales_todos = set(df_dotacion['CodVenta'].dropna().astype(str))
                    
                    # Detectar vendedores eliminados (están en BASE pero no en actual)
                    vendedores_eliminados = vendedores_base - vendedores_actuales_todos
                    
                    if vendedores_eliminados:
                        print(f"⚠ Detectados {len(vendedores_eliminados)} vendedores eliminados completamente (no están en archivo jefe): {vendedores_eliminados}")
                        
                        # Agregar filas de vendedores eliminados con Estatus = 'Eliminado'
                        for cod_venta in vendedores_eliminados:
                            # Buscar vendedor en BASE
                            vendedor_base_rows = df_base_jefe[df_base_jefe['CodVenta'].astype(str) == cod_venta]
                            
                            if not vendedor_base_rows.empty:
                                vendedor_base = vendedor_base_rows.iloc[0].copy()
                                vendedor_base['Estatus'] = 'Eliminado'
                                
                                # Agregar al DataFrame actual usando concat
                                df_dotacion = pd.concat([df_dotacion, vendedor_base.to_frame().T], ignore_index=True)
                                print(f"   ✓ Agregado vendedor {cod_venta} como Eliminado")
                        
                        print(f"✓ Total: {len(vendedores_eliminados)} vendedores eliminados agregados al DataFrame")
        
        # Filtrar datos según usuario (solo si no es admin y tiene zonas asignadas)
        if not is_admin and len(user_zonas) > 0:
            if 'CodDistrito' in df_dotacion.columns:
                df_dotacion = df_dotacion[df_dotacion['CodDistrito'].isin(user_zonas)]
        
        # Columnas de nombre: vectorización rápida en lugar de apply() fila por fila
        _n  = df_dotacion['Nombre'].fillna('').astype(str).str.strip()
        _ap = df_dotacion['APaterno'].fillna('').astype(str).str.strip()
        _am = df_dotacion['AMaterno'].fillna('').astype(str).str.strip()

        # Limpiar Nombre: remover apellidos si ya están incluidos en el campo Nombre
        @np.vectorize
        def _limpiar_nombre(n, ap, am):
            if ap and ap in n:
                n = n.replace(ap, '').strip()
            if am and am in n:
                n = n.replace(am, '').strip()
            return n

        df_dotacion['Nombre'] = _limpiar_nombre(_n.values, _ap.values, _am.values)

        # Actualizar _n con el Nombre ya limpio para NombreCompleto
        _n = df_dotacion['Nombre'].astype(str).str.strip()
        df_dotacion['NombreCompleto'] = (
            (_n + ' ' + _ap + ' ' + _am).str.strip().str.replace(r'\s+', ' ', regex=True)
        )
        
        # Verificar datos faltantes en TODAS las columnas
        # Excluimos TallaPantalon y TallaCamisa que pueden ser opcionales
        columnas_requeridas = [
            'Año', 'Mes', 'CodDistrito', 'DesDistrito', 'CodOficina', 'DesOficina',
            'CodVenta', 'Cargo', 'Rut', 'Nombre', 'APaterno', 'AMaterno',
            'Telefono', 'Correo', 'ZonaEstival', 'Genero'
        ]
        
        # Crear columna que indica si tiene datos faltantes
        df_dotacion['TieneDatosFaltantes'] = df_dotacion[columnas_requeridas].isnull().any(axis=1)
        
        # Preparar columnas base (incluir TODOS los campos originales más el flag de datos faltantes y Estatus)
        columnas_base = [
            'Año', 'Mes', 'CodDistrito', 'DesDistrito', 'CodOficina', 'DesOficina',
            'CodVenta', 'Cargo', 'Rut', 'Nombre', 'APaterno', 'AMaterno',
            'Telefono', 'Correo', 'ZonaEstival', 'Genero', 'TallaPantalon', 'TallaCamisa',
            'NombreCompleto', 'TieneDatosFaltantes', 'Estatus'
        ]
        
        # Detectar columnas extra del archivo original y agregarlas para preservarlas
        columnas_extra = [col for col in df_dotacion.columns if col not in columnas_base]
        if columnas_extra:
            columnas_base = columnas_base + columnas_extra
            print(f"ℹ Columnas adicionales detectadas y preservadas: {columnas_extra}")
        
        # 1. Vendedor Titular: DesOficina == DesDistrito Y Cargo == "TITULAR" (NO Eliminado)
        df_vendedor_titular = df_dotacion[
            (df_dotacion['DesOficina'] == df_dotacion['DesDistrito']) & 
            (df_dotacion['Cargo'].str.upper() == 'TITULAR') &
            (df_dotacion['Estatus'] != 'Eliminado')
        ][columnas_base].copy()
        
        # 2. Gestor Supermercado (Titular): DesOficina == DesDistrito Y Cargo == "GESTOR" (NO Eliminado)
        df_gestor_supermercado = df_dotacion[
            (df_dotacion['DesOficina'] == df_dotacion['DesDistrito']) & 
            (df_dotacion['Cargo'].str.upper() == 'GESTOR') &
            (df_dotacion['Estatus'] != 'Eliminado')
        ][columnas_base].copy()
        
        # 3. Vendedor Titular - Residentes: DesOficina != DesDistrito Y Cargo IN ("TITULAR", "GESTOR") (NO Eliminado)
        df_vendedor_residentes = df_dotacion[
            (df_dotacion['DesOficina'] != df_dotacion['DesDistrito']) & 
            (df_dotacion['Cargo'].str.upper().isin(['TITULAR', 'GESTOR'])) &
            (df_dotacion['Estatus'] != 'Eliminado')
        ][columnas_base].copy()
        
        # 4. Vendedor Reemplazo: Cargo == "REEMPLAZO" (NO Eliminado)
        df_vendedor_reemplazo = df_dotacion[
            (df_dotacion['Cargo'].str.upper() == 'REEMPLAZO') &
            (df_dotacion['Estatus'] != 'Eliminado')
        ][columnas_base].copy()
        
        # 5. Dotación Eliminada: Estatus == 'Eliminado'
        df_eliminados = df_dotacion[
            df_dotacion['Estatus'] == 'Eliminado'
        ][columnas_base].copy()
        
        # Crear sheets para cada tabla
        sheets_data = []
        
        def _df_to_clean_records(df_src):
            """Convierte DataFrame a lista de dicts con tipos JSON-safe."""
            clean = df_src.where(pd.notnull(df_src), None)
            return _sanitize_for_json(clean.to_dict(orient='records'))
        
        # Tabla 1: Vendedor Titular
        sheets_data.append({
            "name": 'Vendedor Titular',
            "rows": len(df_vendedor_titular),
            "columns": list(df_vendedor_titular.columns),
            "data": _df_to_clean_records(df_vendedor_titular)
        })
        
        # Tabla 2: Gestor Supermercado (Titular)
        sheets_data.append({
            "name": 'Gestor Supermercado (Titular)',
            "rows": len(df_gestor_supermercado),
            "columns": list(df_gestor_supermercado.columns),
            "data": _df_to_clean_records(df_gestor_supermercado)
        })
        
        # Tabla 3: Vendedor Titular - Residentes
        sheets_data.append({
            "name": 'Vendedor Titular - Residentes',
            "rows": len(df_vendedor_residentes),
            "columns": list(df_vendedor_residentes.columns),
            "data": _df_to_clean_records(df_vendedor_residentes)
        })
        
        # Tabla 4: Vendedor Reemplazo
        sheets_data.append({
            "name": 'Vendedor Reemplazo',
            "rows": len(df_vendedor_reemplazo),
            "columns": list(df_vendedor_reemplazo.columns),
            "data": _df_to_clean_records(df_vendedor_reemplazo)
        })
        
        # Tabla 5: Dotación Eliminada
        sheets_data.append({
            "name": 'Dotacion Eliminada',
            "rows": len(df_eliminados),
            "columns": list(df_eliminados.columns),
            "data": _df_to_clean_records(df_eliminados)
        })
        
        # Contar estadísticas
        total_vendedores = len(df_dotacion)
        total_distritos = df_dotacion['CodDistrito'].nunique() if 'CodDistrito' in df_dotacion.columns else 0
        total_oficinas = df_dotacion['DesOficina'].nunique() if 'DesOficina' in df_dotacion.columns else 0
        
        # Obtener fecha de última modificación
        last_modified = datetime.fromtimestamp(archivo_a_cargar.stat().st_mtime)
        
        # Verificar si el usuario actual (jefe) completó la validación
        es_pendiente = False
        
        # Si es jefe de venta, verificar si ha validado TODAS sus zonas
        if current_user.cargo and 'JEFE' in current_user.cargo.upper():
            zonas_jefe = obtener_zonas_usuario(current_user.usuario)
            if not zonas_jefe:
                es_pendiente = True
            else:
                # Pendiente si ALGUNA zona no ha sido validada
                es_pendiente = any(
                    not verificar_validacion_jefe_por_archivo(current_user.usuario, zona)
                    for zona in zonas_jefe
                )
        
        return {
            "sheets": sheets_data,
            "metadata": {
                "filename": archivo_a_cargar.name,
                "last_update": last_modified.strftime("%Y-%m-%d %H:%M:%S"),
                "total_vendedores": total_vendedores,
                "total_distritos": total_distritos,
                "total_oficinas": total_oficinas,
                "total_sheets": len(sheets_data),
                "user_role": current_user.cargo,
                "user_distritos": user_zonas if not is_admin else [],
                "es_pendiente_validacion": es_pendiente
            }
        }
    
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al cargar estructura de ventas: {str(e)}"
        )


@router.get("/vendedores")
async def get_vendedores(
    distrito: str = None,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Obtiene la lista de vendedores, opcionalmente filtrada por distrito
    """
    try:
        if not EXCEL_ORIGINAL_PATH.exists():
            raise HTTPException(status_code=404, detail="Archivo base no encontrado")
        
        df = pd.read_excel(EXCEL_ORIGINAL_PATH, sheet_name='TBL Vendedor')
        
        if distrito:
            df = df[df['CodDistrito'] == distrito]
        
        df = df.where(pd.notnull(df), None)
        return df.to_dict(orient='records')
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/distritos")
async def get_distritos(current_user: UserInfo = Depends(get_current_user)):
    """
    Obtiene la lista de distritos
    """
    try:
        if not EXCEL_BASE_PATH.exists():
            raise HTTPException(status_code=404, detail="Archivo base no encontrado")
        
        df = pd.read_excel(EXCEL_BASE_PATH, sheet_name='TBL Distrito')
        df = df.where(pd.notnull(df), None)
        
        return df.to_dict(orient='records')
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/oficinas")
async def get_oficinas(
    distrito: str = None,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Obtiene la lista de oficinas, opcionalmente filtrada por distrito
    """
    try:
        if not EXCEL_BASE_PATH.exists():
            raise HTTPException(status_code=404, detail="Archivo base no encontrado")
        
        df = pd.read_excel(EXCEL_BASE_PATH, sheet_name='TBL Oficinas')
        
        if distrito:
            df = df[df['CodDistrito'] == distrito]
        
        df = df.where(pd.notnull(df), None)
        return df.to_dict(orient='records')
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/comunas")
async def get_comunas(current_user: UserInfo = Depends(get_current_user)):
    """
    Obtiene la lista de comunas
    """
    try:
        if not EXCEL_BASE_PATH.exists():
            raise HTTPException(status_code=404, detail="Archivo base no encontrado")
        
        df = pd.read_excel(EXCEL_BASE_PATH, sheet_name='TBL Comunas')
        df = df.where(pd.notnull(df), None)
        
        return df.to_dict(orient='records')
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/guardar-cambios")
async def guardar_cambios(
    request: GuardarCambiosRequest,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Guarda los cambios realizados en la estructura de ventas.
    Crea una copia de respaldo con timestamp y guarda los cambios en un nuevo archivo.
    """
    try:
        if not EXCEL_BASE_PATH.exists():
            raise HTTPException(status_code=404, detail="Archivo base no encontrado")
        
        # Crear estructura de carpetas
        respaldo_dir = EXCEL_BASE_PATH.parent / "respaldo"
        cambios_dir = EXCEL_BASE_PATH.parent / "cambios"
        respaldo_dir.mkdir(exist_ok=True)
        cambios_dir.mkdir(exist_ok=True)
        
        # Ruta del archivo original en respaldo
        archivo_original = respaldo_dir / "TBL_EstructuraVenta_ORIGINAL.xlsm"
        
        # Si no existe el original en respaldo, copiarlo
        if not archivo_original.exists():
            shutil.copy2(EXCEL_BASE_PATH, archivo_original)
        
        # Crear nombre de archivo de cambio con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        cambio_filename = f"TBL_EstructuraVenta_cambio_{timestamp}.xlsm"
        cambio_path = cambios_dir / cambio_filename
        
        # Leer el archivo Excel con pandas
        df_vendedor = pd.read_excel(EXCEL_BASE_PATH, sheet_name='TBL Vendedor')
        df_dotacion = pd.read_excel(EXCEL_BASE_PATH, sheet_name='TBL Dotacion')
        
        # Verificar que existe la columna CodVend
        if 'CodVend' not in df_vendedor.columns:
            raise HTTPException(status_code=500, detail="No se encontró la columna CodVend en el Excel")
        
        # Aplicar cada cambio
        cambios_aplicados = 0
        vendedores_nuevos = []
        
        for codVend, cambios_vendedor in request.editedData.items():
            # Verificar si es un vendedor nuevo
            es_nuevo = any(c.codVend == codVend and c.field == 'NUEVO_VENDEDOR' for c in request.cambios)
            
            if es_nuevo:
                # Validar que el vendedor nuevo tenga todos los campos obligatorios
                es_valido, errores = validar_vendedor_nuevo(cambios_vendedor)
                if not es_valido:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Vendedor {codVend} tiene campos incompletos o inválidos: {'; '.join(errores)}"
                    )
                
                # Crear nueva fila para vendedor nuevo
                nueva_fila = cambios_vendedor.copy()
                df_vendedor = pd.concat([df_vendedor, pd.DataFrame([nueva_fila])], ignore_index=True)
                
                # Agregar también a TBL Dotacion si tiene campos correspondientes
                if 'NombreVendedor' in cambios_vendedor or 'ApellidoPaterno' in cambios_vendedor:
                    nueva_fila_dotacion = {
                        'CodVta': cambios_vendedor.get('CodVend'),
                        'Nombre': cambios_vendedor.get('NombreVendedor'),
                        'APaterno': cambios_vendedor.get('ApellidoPaterno'),
                        'AMaterno': cambios_vendedor.get('ApellidoMaterno')
                    }
                    df_dotacion = pd.concat([df_dotacion, pd.DataFrame([nueva_fila_dotacion])], ignore_index=True)
                
                vendedores_nuevos.append(codVend)
                cambios_aplicados += 1
            else:
                # Buscar el vendedor existente por CodVend
                mask = df_vendedor['CodVend'].astype(str).str.strip() == str(codVend).strip()
                indices = df_vendedor[mask].index
                
                if len(indices) > 0:
                    idx = indices[0]
                    
                    # Aplicar cambios en cada campo en TBL Vendedor
                    for field, new_value in cambios_vendedor.items():
                        # Mapear nombres de campos del frontend a columnas del Excel
                        campo_excel = field
                        
                        # Mapeo de campos
                        if field == 'CorreoDotacion' and 'CorreoVendedor' in df_vendedor.columns:
                            campo_excel = 'CorreoVendedor'
                        
                        if campo_excel in df_vendedor.columns:
                            # Verificar si este campo fue explícitamente cambiado
                            cambio_explicito = any(c.codVend == codVend and c.field == field for c in request.cambios)
                            
                            if cambio_explicito or new_value:
                                df_vendedor.at[idx, campo_excel] = new_value
                        else:
                            pass  # Campo no encontrado en TBL Vendedor
                    
                    # También actualizar TBL Dotacion si cambiaron Nombre/Apellidos
                    if 'NombreVendedor' in cambios_vendedor or 'ApellidoPaterno' in cambios_vendedor or 'ApellidoMaterno' in cambios_vendedor:
                        mask_dotacion = df_dotacion['CodVta'].astype(str).str.strip() == str(codVend).strip()
                        indices_dotacion = df_dotacion[mask_dotacion].index
                        
                        if len(indices_dotacion) > 0:
                            idx_dotacion = indices_dotacion[0]
                            if 'NombreVendedor' in cambios_vendedor and 'Nombre' in df_dotacion.columns:
                                df_dotacion.at[idx_dotacion, 'Nombre'] = cambios_vendedor['NombreVendedor']
                            if 'ApellidoPaterno' in cambios_vendedor and 'APaterno' in df_dotacion.columns:
                                df_dotacion.at[idx_dotacion, 'APaterno'] = cambios_vendedor['ApellidoPaterno']
                            if 'ApellidoMaterno' in cambios_vendedor and 'AMaterno' in df_dotacion.columns:
                                df_dotacion.at[idx_dotacion, 'AMaterno'] = cambios_vendedor['ApellidoMaterno']
                    
                    cambios_aplicados += 1
        
        # Guardar usando ExcelWriter para mantener el formato
        modified_filename = f"TBL_EstructuraVenta_modificado_{timestamp}.xlsm"
        modified_path = EXCEL_BASE_PATH.parent / modified_filename
        
        # Leer todas las hojas del archivo original
        with pd.ExcelFile(EXCEL_BASE_PATH) as xls:
            with pd.ExcelWriter(modified_path, engine='openpyxl') as writer:
                # Guardar hojas modificadas
                df_vendedor.to_excel(writer, sheet_name='TBL Vendedor', index=False)
                df_dotacion.to_excel(writer, sheet_name='TBL Dotacion', index=False)
                
                # Copiar las demás hojas sin modificar
                for sheet_name in xls.sheet_names:
                    if sheet_name not in ['TBL Vendedor', 'TBL Dotacion']:
                        df = pd.read_excel(xls, sheet_name=sheet_name)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Guardar copia en carpeta de cambios
        shutil.copy2(modified_path, cambio_path)
        
        # Actualizar el archivo principal con el modificado
        shutil.copy2(modified_path, EXCEL_BASE_PATH)
        
        return {
            "success": True,
            "message": f"Se guardaron {len(request.cambios)} cambios exitosamente",
            "cambios_aplicados": cambios_aplicados,
            "vendedores_nuevos": len(vendedores_nuevos),
            "archivo_cambio": cambio_filename,
            "archivo_backup": f"Backup_{timestamp}.xlsx"
        }
    
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al guardar cambios: {str(e)}"
        )


@router.post("/restablecer")
async def restablecer_archivo(
    body: Dict[str, Any] = {},
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Restaura los datos de una zona específica desde el archivo base original.
    Si se envía 'zona', solo restablece esa zona preservando el resto.
    Si no se envía 'zona', restablece todas las zonas del jefe.
    """
    try:
        zona = body.get('zona', None)
        archivo_jefe = obtener_archivo_jefe(current_user.usuario)
        
        # Si no existe archivo individual, ya está usando el base
        if not archivo_jefe.exists():
            if zona:
                limpiar_estado_validacion(current_user.usuario, zona)
            else:
                limpiar_estado_validacion(current_user.usuario)
            return {
                "success": True,
                "message": "Ya estás usando el archivo base original (no hay cambios guardados)",
                "archivo_original": EXCEL_ORIGINAL_PATH.name
            }
        
        # Crear backup antes de modificar
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = archivo_jefe.parent / f"Backup_antes_restablecer_{current_user.usuario}_{timestamp}.xlsx"
        shutil.copy2(archivo_jefe, backup_path)
        
        if zona:
            # === RESTABLECER SOLO UNA ZONA ===
            # Leer archivo individual del jefe
            df_jefe = pd.read_excel(archivo_jefe, sheet_name='Dotacion')
            
            # Leer archivo base original
            if not EXCEL_ORIGINAL_PATH.exists():
                raise HTTPException(status_code=404, detail="Archivo base no encontrado")
            df_base = excel_cache.read_excel(EXCEL_ORIGINAL_PATH, sheet_name='Dotacion')
            
            # Quitar los datos de la zona a restablecer del archivo del jefe
            df_jefe_sin_zona = df_jefe[df_jefe['CodDistrito'] != zona]
            
            # Obtener los datos originales de esa zona desde el base
            df_base_zona = df_base[df_base['CodDistrito'] == zona].copy()
            
            # Combinar: datos del jefe (sin la zona) + datos base (de la zona)
            df_resultado = pd.concat([df_jefe_sin_zona, df_base_zona], ignore_index=True)
            
            # Leer otras hojas del archivo del jefe para preservarlas
            with pd.ExcelFile(archivo_jefe) as xls_file:
                otras_hojas_nombres = [s for s in xls_file.sheet_names if s != 'Dotacion']
            existing_sheets = {}
            for sheet_name in otras_hojas_nombres:
                existing_sheets[sheet_name] = pd.read_excel(archivo_jefe, sheet_name=sheet_name)
            
            # Guardar el archivo actualizado
            with pd.ExcelWriter(archivo_jefe, engine='openpyxl', mode='w') as writer:
                df_resultado.to_excel(writer, sheet_name='Dotacion', index=False)
                for sheet_name, sheet_df in existing_sheets.items():
                    sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Invalidar caché
            excel_cache.invalidate(archivo_jefe)
            
            # Limpiar validación solo de esa zona
            limpiar_estado_validacion(current_user.usuario, zona)
            print(f"✓ Zona {zona} restablecida para {current_user.usuario}")
            
            # Verificar si quedan cambios en otras zonas; si no, eliminar archivo
            user_zonas = obtener_zonas_usuario(current_user.usuario)
            df_base_full = excel_cache.read_excel(EXCEL_ORIGINAL_PATH, sheet_name='Dotacion')
            df_base_user = df_base_full[df_base_full['CodDistrito'].isin(user_zonas)]
            
            # Comparar tamaños: si el archivo del jefe es igual al base filtrado, eliminar
            if len(df_resultado[df_resultado['CodDistrito'].isin(user_zonas)]) == len(df_base_user):
                # Comparación simple por tamaño; puede haber diferencias en contenido
                # pero es una buena heurística para limpiar archivos innecesarios
                pass  # Mantener el archivo por si hay diferencias en otras hojas
            
            return {
                "success": True,
                "message": f"Zona {zona} restablecida exitosamente desde el archivo base.",
                "zona": zona,
                "backup_creado": backup_path.name
            }
        else:
            # === RESTABLECER TODAS LAS ZONAS ===
            # Eliminar el archivo individual completo
            archivo_jefe.unlink()
            
            # Limpiar validación de todas las zonas del usuario
            user_zonas = obtener_zonas_usuario(current_user.usuario)
            for z in user_zonas:
                limpiar_estado_validacion(current_user.usuario, z)
            limpiar_estado_validacion(current_user.usuario)
            
            # Invalidar caché
            excel_cache.invalidate(archivo_jefe)
            
            print(f"✓ Todas las zonas restablecidas para {current_user.usuario}")
            
            return {
                "success": True,
                "message": "Archivo restablecido exitosamente. Ahora estás usando el archivo base original.",
                "archivo_eliminado": archivo_jefe.name,
                "backup_creado": backup_path.name
            }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al restablecer archivo: {str(e)}"
        )


@router.post("/agregar-vendedor")
async def agregar_vendedor(
    vendedor: Dict[str, Any],
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Agrega un nuevo vendedor al archivo Excel del jefe
    """
    try:
        # Determinar qué archivo usar (individual del jefe o el base)
        archivo_jefe = obtener_archivo_jefe(current_user.usuario)
        
        if archivo_jefe.exists():
            archivo_fuente = archivo_jefe
        else:
            archivo_fuente = EXCEL_ORIGINAL_PATH
        
        if not archivo_fuente.exists():
            raise HTTPException(status_code=404, detail=f"Archivo fuente no encontrado: {archivo_fuente}")
        
        # Leer el archivo Excel desde Cambios o desde el base
        df = pd.read_excel(archivo_fuente, sheet_name='Dotacion')
        
        # Crear nuevo registro con conversión de tipos según las columnas del DataFrame
        campos_vendedor = {
            'Año': vendedor.get('Año'),
            'Mes': vendedor.get('Mes'),
            'CodDistrito': vendedor.get('CodDistrito'),
            'DesDistrito': vendedor.get('DesDistrito'),
            'CodOficina': vendedor.get('CodOficina'),
            'DesOficina': vendedor.get('DesOficina'),
            'CodVenta': vendedor.get('CodVenta'),
            'Cargo': vendedor.get('Cargo'),
            'Rut': vendedor.get('Rut'),
            'Nombre': vendedor.get('Nombre'),
            'APaterno': vendedor.get('APaterno'),
            'AMaterno': vendedor.get('AMaterno'),
            'Telefono': vendedor.get('Telefono'),
            'Correo': vendedor.get('Correo'),
            'ZonaEstival': vendedor.get('ZonaEstival'),
            'Genero': vendedor.get('Genero'),
            'TallaPantalon': vendedor.get('TallaPantalon'),
            'TallaCamisa': vendedor.get('TallaCamisa'),
            'Estatus': 'Nuevo'
        }
        
        # Convertir cada campo según el tipo de la columna existente
        nuevo_registro = {}
        for campo, valor in campos_vendedor.items():
            if campo in df.columns:
                nuevo_registro[campo] = convertir_valor_segun_tipo(valor, df[campo].dtype)
            else:
                nuevo_registro[campo] = valor
        
        # Agregar el nuevo registro al DataFrame
        df = pd.concat([df, pd.DataFrame([nuevo_registro])], ignore_index=True)
        
        # Crear backup antes de modificar en la carpeta Cambios
        if archivo_jefe.exists():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = archivo_jefe.parent / f"Backup_antes_agregar_{current_user.usuario}_{timestamp}.xlsx"
            shutil.copy2(archivo_jefe, backup_path)
        
        # Guardar el archivo actualizado en el archivo individual del jefe
        # Primero leer todas las hojas existentes
        existing_sheets = {}
        with pd.ExcelFile(archivo_fuente) as xls:
            for sheet_name in xls.sheet_names:
                if sheet_name != 'Dotacion':
                    existing_sheets[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
        
        # Escribir el archivo completo en el archivo individual
        with pd.ExcelWriter(archivo_jefe, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name='Dotacion', index=False)
            # Restaurar otras hojas
            for sheet_name, sheet_df in existing_sheets.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return {
            "success": True,
            "message": "Vendedor agregado exitosamente",
            "vendedor": nuevo_registro,
            "archivo_actualizado": archivo_jefe.name
        }
    
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al agregar vendedor: {str(e)}"
        )


@router.post("/actualizar-vendedor")
async def actualizar_vendedor(
    vendedor: Dict[str, Any],
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Actualiza un vendedor existente en el archivo Excel del jefe
    """
    try:
        # Determinar qu\u00e9 archivo usar (individual del jefe o el base)
        archivo_jefe = obtener_archivo_jefe(current_user.usuario)
        
        if archivo_jefe.exists():
            archivo_fuente = archivo_jefe
        else:
            archivo_fuente = EXCEL_ORIGINAL_PATH
        
        if not archivo_fuente.exists():
            raise HTTPException(status_code=404, detail=f"Archivo fuente no encontrado: {archivo_fuente}")
        
        # Leer el archivo Excel
        df = pd.read_excel(archivo_fuente, sheet_name='Dotacion')
        
        # Buscar el vendedor por CodVenta y Rut
        cod_venta = vendedor.get('CodVenta')
        rut = vendedor.get('Rut')
        
        if not cod_venta or not rut:
            raise HTTPException(status_code=400, detail="CodVenta y Rut son requeridos para actualizar")
        
        # Encontrar el \u00edndice del vendedor
        mask = (df['CodVenta'] == cod_venta) & (df['Rut'] == rut)
        
        if not mask.any():
            raise HTTPException(status_code=404, detail=f"Vendedor con CodVenta {cod_venta} y Rut {rut} no encontrado")
        
        # Actualizar los campos con conversi\u00f3n de tipos
        for campo, valor in vendedor.items():
            if campo in df.columns:
                valor_convertido = convertir_valor_segun_tipo(valor, df[campo].dtype)
                df.loc[mask, campo] = valor_convertido
        
        # Crear backup antes de modificar
        if archivo_jefe.exists():
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = archivo_jefe.parent / f"Backup_antes_actualizar_{current_user.usuario}_{timestamp}.xlsx"
            shutil.copy2(archivo_jefe, backup_path)
        
        # Guardar el archivo actualizado en el archivo individual del jefe
        # Primero leer todas las hojas existentes
        existing_sheets = {}
        with pd.ExcelFile(archivo_fuente) as xls:
            for sheet_name in xls.sheet_names:
                if sheet_name != 'Dotacion':
                    existing_sheets[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
        
        # Escribir el archivo completo en el archivo individual
        with pd.ExcelWriter(archivo_jefe, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name='Dotacion', index=False)
            # Restaurar otras hojas
            for sheet_name, sheet_df in existing_sheets.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return {
            "success": True,
            "message": "Vendedor actualizado exitosamente",
            "vendedor": vendedor
        }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al actualizar vendedor: {str(e)}"
        )


@router.post("/guardar-todos-cambios")
async def guardar_todos_cambios(
    cambios: Dict[str, Any],
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Guarda todos los cambios en el archivo individual del jefe
    Parámetro 'marcar_validado' indica si debe marcar como validado (solo desde Estructura de Venta)
    """
    try:
        # Determinar el archivo fuente (de dónde leer)
        archivo_jefe = obtener_archivo_jefe(current_user.usuario)
        
        if archivo_jefe.exists():
            # Ya existe su archivo, leer de ahí
            archivo_fuente = archivo_jefe
        else:
            # No existe, leer del base
            archivo_fuente = EXCEL_ORIGINAL_PATH
        
        if not archivo_fuente.exists():
            raise HTTPException(status_code=404, detail=f"Archivo fuente no encontrado: {archivo_fuente}")
        
        # Leer el archivo Excel
        df = pd.read_excel(archivo_fuente, sheet_name='Dotacion')
        
        # Asegurarse de que existe la columna Estatus y que todos los registros tengan un valor
        if 'Estatus' not in df.columns:
            df['Estatus'] = 'Antiguo'
        else:
            # Si la columna existe pero tiene valores nulos, asignar 'Antiguo'
            df['Estatus'] = df['Estatus'].fillna('Antiguo')
            # Si tiene valores vacíos, reemplazarlos por 'Antiguo'
            df.loc[df['Estatus'] == '', 'Estatus'] = 'Antiguo'
        
        editados = cambios.get('editados', [])
        nuevos = cambios.get('nuevos', [])
        eliminados = cambios.get('eliminados', [])
        marcar_validado = cambios.get('marcar_validado', False)  # Por defecto NO marcar como validado
        zona = cambios.get('zona', None)  # Zona a validar
        
        print(f"\n📝 Procesando cambios:")
        print(f"   • {len(editados)} editados")
        print(f"   • {len(nuevos)} nuevos")
        print(f"   • {len(eliminados)} eliminados")
        
        # DEBUG: Ver contenido de editados
        if len(editados) > 0:
            print(f"\n🔍 DEBUG - Vendedores editados recibidos:")
            for v in editados:
                print(f"   - {v.get('NombreCompleto', 'SIN NOMBRE')} | CodVenta: {v.get('CodVenta', 'N/A')} | Rut: {v.get('Rut', 'N/A')}")
        
        # Aplicar cambios de vendedores eliminados (marcarlos como Estatus='Eliminado')
        for vendedor in eliminados:
            cod_venta = vendedor.get('CodVenta')
            rut = vendedor.get('Rut')
            
            if cod_venta and rut:
                mask = (df['CodVenta'] == cod_venta) & (df['Rut'] == rut)
                
                if mask.any():
                    df.loc[mask, 'Estatus'] = 'Eliminado'
                    print(f"   ✓ Eliminado: {vendedor.get('NombreCompleto', '')} (CodVenta: {cod_venta})")
        
        # Aplicar cambios de vendedores editados
        for vendedor in editados:
            cod_venta = vendedor.get('CodVenta')
            rut = vendedor.get('Rut')
            
            if cod_venta and rut:
                mask = (df['CodVenta'] == cod_venta) & (df['Rut'] == rut)
                
                if mask.any():
                    # Verificar estatus antes de actualizar
                    estatus_antes = df.loc[mask, 'Estatus'].iloc[0]
                    estatus_frontend = vendedor.get('Estatus', '')
                    print(f"   🔍 Procesando editado: {vendedor.get('NombreCompleto', '')} (Estatus antes: {estatus_antes}, Estatus del frontend: {estatus_frontend})")
                    
                    # Actualizar TODOS los campos con conversión de tipos
                    for campo, valor in vendedor.items():
                        if campo in df.columns:
                            valor_convertido = convertir_valor_segun_tipo(valor, df[campo].dtype)
                            df.loc[mask, campo] = valor_convertido
                    
                    # Si el vendedor era 'Antiguo' y ahora está siendo editado, cambiar a 'Modificado'
                    # (solo si el frontend no envió explícitamente 'Nuevo' o 'Modificado')
                    estatus_actual = df.loc[mask, 'Estatus'].iloc[0]
                    if estatus_actual == 'Antiguo' or (estatus_antes == 'Antiguo' and estatus_frontend not in ['Nuevo', 'Modificado']):
                        df.loc[mask, 'Estatus'] = 'Modificado'
                        print(f"   📝 Estatus cambiado automáticamente de '{estatus_antes}' a 'Modificado'")
                    
                    # Verificar el estatus final
                    estatus_final = df.loc[mask, 'Estatus'].iloc[0]
                    print(f"   ✓ Editado: {vendedor.get('NombreCompleto', '')} (CodVenta: {cod_venta}) -> Estatus: {estatus_final}")
                else:
                    print(f"   ⚠️ NO ENCONTRADO: CodVenta={cod_venta}, Rut={rut}")
        
        # Agregar vendedores nuevos
        for vendedor in nuevos:
            # Validar que el vendedor nuevo tenga todos los campos obligatorios
            es_valido, errores = validar_vendedor_nuevo(vendedor)
            if not es_valido:
                cod_venta = vendedor.get('CodVenta', 'desconocido')
                raise HTTPException(
                    status_code=400,
                    detail=f"Vendedor {cod_venta} tiene campos incompletos o inválidos: {'; '.join(errores)}"
                )
            
            # Crear registro con conversión de tipos según las columnas del DataFrame
            nuevo_registro = {}
            campos_vendedor = {
                'Año': vendedor.get('Año'),
                'Mes': vendedor.get('Mes'),
                'CodDistrito': vendedor.get('CodDistrito'),
                'DesDistrito': vendedor.get('DesDistrito'),
                'CodOficina': vendedor.get('CodOficina'),
                'DesOficina': vendedor.get('DesOficina'),
                'CodVenta': vendedor.get('CodVenta'),
                'Cargo': vendedor.get('Cargo'),
                'Rut': vendedor.get('Rut'),
                'Nombre': vendedor.get('Nombre'),
                'APaterno': vendedor.get('APaterno'),
                'AMaterno': vendedor.get('AMaterno'),
                'Telefono': vendedor.get('Telefono'),
                'Correo': vendedor.get('Correo'),
                'ZonaEstival': vendedor.get('ZonaEstival'),
                'Genero': vendedor.get('Genero'),
                'TallaPantalon': vendedor.get('TallaPantalon'),
                'TallaCamisa': vendedor.get('TallaCamisa'),
                'Estatus': 'Nuevo'
            }
            
            # Convertir cada campo según el tipo de la columna existente
            for campo, valor in campos_vendedor.items():
                if campo in df.columns:
                    nuevo_registro[campo] = convertir_valor_segun_tipo(valor, df[campo].dtype)
                else:
                    nuevo_registro[campo] = valor
            
            df = pd.concat([df, pd.DataFrame([nuevo_registro])], ignore_index=True)
            print(f"   ✓ Nuevo: {vendedor.get('NombreCompleto', '')} (CodVenta: {vendedor.get('CodVenta', '')})")
        
        # Obtener nombres de hojas y cerrar inmediatamente para evitar bloqueo en Windows
        with pd.ExcelFile(archivo_fuente) as xls_file:
            otras_hojas = [s for s in xls_file.sheet_names if s != 'Dotacion']
        
        # Leer las otras hojas usando caché (archivo ya cerrado)
        existing_sheets = {}
        for sheet_name in otras_hojas:
            existing_sheets[sheet_name] = excel_cache.read_excel(archivo_fuente, sheet_name=sheet_name)
        
        # Guardar en el archivo individual del jefe
        with pd.ExcelWriter(archivo_jefe, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, sheet_name='Dotacion', index=False)
            # Restaurar otras hojas
            for sheet_name, sheet_df in existing_sheets.items():
                sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"💾 Archivo guardado para {current_user.usuario}: {archivo_jefe}")
        
        # Invalidar caché del archivo del jefe para que la próxima lectura sea fresca
        excel_cache.invalidate(archivo_jefe)
        
        # DEBUG: Verificar conteo de estatus después de guardar
        estatus_counts = df['Estatus'].value_counts().to_dict()
        print(f"📊 Resumen de Estatus guardado:")
        for estatus, count in estatus_counts.items():
            print(f"   - {estatus}: {count}")
        
        # IMPORTANTE: Solo marcar como validado si viene desde Estructura de Venta
        if marcar_validado and current_user.cargo and 'JEFE' in current_user.cargo.upper():
            marcar_como_validado(current_user.usuario, zona)
            zona_texto = f" - Zona: {zona}" if zona else ""
            print(f"✓ Validación marcada para {current_user.usuario}{zona_texto} (desde Estructura de Venta)")
        else:
            print(f"ℹ Cambios guardados para {current_user.usuario} - NO validado (desde Administrar Dotación)")
        
        return {
            "success": True,
            "message": f"Cambios guardados exitosamente. {len(editados)} editados, {len(nuevos)} nuevos, {len(eliminados)} eliminados.",
            "archivo_generado": archivo_jefe.name,
            "total_cambios": len(editados) + len(nuevos) + len(eliminados),
            "validado": marcar_validado
        }
    
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al guardar todos los cambios: {str(e)}"
        )


@router.post("/validar-sin-cambios")
async def validar_sin_cambios(
    request: ValidarZonaRequest,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Permite a un jefe validar su estructura sin hacer cambios.
    Valida una zona específica del jefe.
    """
    try:
        # Solo jefes de venta pueden validar
        if not current_user.cargo or 'JEFE' not in current_user.cargo.upper():
            raise HTTPException(
                status_code=403,
                detail="Solo jefes de venta pueden validar estructura"
            )
        
        zona = request.zona
        archivo_jefe = obtener_archivo_jefe(current_user.usuario)
        
        # Verificar si ya fue validado explícitamente esta zona
        archivo_estado = obtener_archivo_estado_validacion(current_user.usuario, zona)
        if archivo_estado.exists():
            try:
                with open(archivo_estado, 'r', encoding='utf-8') as f:
                    estado = json.load(f)
                    return {
                        "success": True,
                        "message": f"Zona {zona} ya validada previamente",
                        "usuario": current_user.usuario,
                        "zona": zona,
                        "fecha": estado.get('fecha_validacion', 'fecha desconocida')
                    }
            except:
                pass
        
        # Si no existe archivo individual, copiar el base
        if not archivo_jefe.exists():
            if not EXCEL_ORIGINAL_PATH.exists():
                raise HTTPException(status_code=404, detail="Archivo base no encontrado")
            
            shutil.copy2(EXCEL_ORIGINAL_PATH, archivo_jefe)
            print(f" Archivo creado para {current_user.usuario}: {archivo_jefe}")
        else:
            print(f" Archivo ya existe para {current_user.usuario}, validando zona {zona}")
        
        # IMPORTANTE: Marcar explícitamente como validado solo esta zona
        marcar_como_validado(current_user.usuario, zona)
        
        return {
            "success": True,
            "message": f"Zona {zona} validada exitosamente",
            "usuario": current_user.usuario,
            "zona": zona,
            "archivo": archivo_jefe.name,
            "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al validar estructura: {str(e)}"
        )


@router.get("/historial-validacion")
async def obtener_historial_validacion(current_user: UserInfo = Depends(get_current_user)):
    """
    Obtiene el historial de validación de todos los jefes de venta.
    Muestra si han completado o no la validación mensual de la estructura de ventas.
    Ahora muestra el estado por zona independiente.
    """
    try:
        # Verificar que solo admin puede ver esto
        if current_user.cargo.upper() != 'ADMIN':
            raise HTTPException(status_code=403, detail="Solo administradores pueden ver el historial")
        
        if not USUARIOS_SISTEMA_PATH.exists():
            raise HTTPException(status_code=404, detail="Archivo de usuarios no encontrado")
        
        # Cargar usuarios
        df_usuarios = excel_cache.read_excel(USUARIOS_SISTEMA_PATH, sheet_name=0)
        
        # Filtrar solo jefes de venta
        jefes_venta = df_usuarios[df_usuarios['Cargo'].str.upper() == 'JEFE DE VENTA'].copy()
        
        historial = []
        
        for _, jefe in jefes_venta.iterrows():
            usuario = jefe['Usuario']
            nombre = jefe['Nombre']
            
            # Obtener zonas asignadas
            zonas = obtener_zonas_usuario(usuario)
            
            if not zonas:
                # Si no tiene zonas asignadas, crear una entrada sin zona
                historial.append({
                    "usuario": usuario,
                    "nombre": nombre,
                    "zonas": "Sin zonas asignadas",
                    "estado": "Pendiente",
                    "fecha_validacion": None
                })
            else:
                # Crear una entrada por cada zona
                for zona in zonas:
                    # Verificar si esta zona específica ha sido validada
                    tiene_validacion = verificar_validacion_jefe_por_archivo(usuario, zona)
                    
                    fecha_validacion = None
                    
                    if tiene_validacion:
                        # Leer fecha de validación del archivo de estado
                        archivo_estado = obtener_archivo_estado_validacion(usuario, zona)
                        try:
                            with open(archivo_estado, 'r', encoding='utf-8') as f:
                                estado = json.load(f)
                                fecha_validacion = estado.get('fecha_validacion')
                        except:
                            # Si no se puede leer, usar fecha de modificación del archivo
                            fecha_validacion = datetime.fromtimestamp(
                                archivo_estado.stat().st_mtime
                            ).strftime("%Y-%m-%d %H:%M:%S")
                    
                    historial.append({
                        "usuario": usuario,
                        "nombre": nombre,
                        "zonas": zona,
                        "estado": "Completado" if tiene_validacion else "Pendiente",
                        "fecha_validacion": fecha_validacion
                    })
        
        return {
            "historial": historial,
            "total_jefes": len(jefes_venta)
        }
    
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al obtener historial de validación: {str(e)}"
        )


@router.get("/estados-jefes")
async def obtener_estados_jefes(current_user: UserInfo = Depends(get_current_user)):
    """
    Obtiene el estado de completitud de todos los jefes de venta
    Solo accesible para administradores
    """
    try:
        # Verificar que sea admin
        if current_user.cargo.upper() != 'ADMIN':
            raise HTTPException(status_code=403, detail="Acceso denegado. Solo administradores.")
        
        if not USUARIOS_SISTEMA_PATH.exists():
            raise HTTPException(status_code=404, detail=f"Archivo de usuarios no encontrado en: {USUARIOS_SISTEMA_PATH}")
        
        # Leer usuarios del sistema
        df_usuarios = excel_cache.read_excel(USUARIOS_SISTEMA_PATH, sheet_name=0)
        
        print(f"DEBUG: Columnas del archivo: {df_usuarios.columns.tolist()}")
        print(f"DEBUG: Total usuarios: {len(df_usuarios)}")
        
        if 'Cargo' in df_usuarios.columns:
            print(f"DEBUG: Cargos únicos: {df_usuarios['Cargo'].unique()}")
        
        # Filtrar solo Jefes de Venta (case insensitive)
        jefes_venta = df_usuarios[df_usuarios['Cargo'].str.upper().str.contains('JEFE', na=False)].copy()
        
        print(f"DEBUG: Jefes encontrados: {len(jefes_venta)}")
        
        estados = []
        
        for _, jefe in jefes_venta.iterrows():
            usuario = jefe['Usuario']
            nombre = jefe['Nombre']
            cargo = jefe['Cargo']
            
            print(f"DEBUG: Procesando jefe: {usuario} - {nombre}")
            
            # Obtener zonas asignadas (códigos)
            zonas_codigos = obtener_zonas_usuario(usuario)
            
            # Obtener nombres completos de las zonas
            zonas_completas = obtener_nombres_zonas(zonas_codigos)
            
            print(f"DEBUG: Zonas para {usuario}: {zonas_completas}")
            
            # Obtener archivo individual del jefe
            archivo_jefe = obtener_archivo_jefe(usuario)
            ha_validado = verificar_validacion_jefe_por_archivo(usuario)  # Verificar archivo de estado de validación
            completado = ha_validado
            total_vendedores = 0
            fecha_ultima_actualizacion = None
            
            if ha_validado and archivo_jefe.exists() and len(zonas_codigos) > 0:
                try:
                    # Leer el archivo individual del jefe
                    df_work = pd.read_excel(archivo_jefe, sheet_name='Dotacion')
                    
                    # Contar vendedores en las zonas del jefe (para información)
                    if 'CodDistrito' in df_work.columns:
                        vendedores_jefe = df_work[df_work['CodDistrito'].isin(zonas_codigos)]
                        total_vendedores = len(vendedores_jefe)
                        
                        # Obtener fecha de última modificación del archivo
                        fecha_ultima_actualizacion = datetime.fromtimestamp(
                            archivo_jefe.stat().st_mtime
                        ).strftime("%Y-%m-%d %H:%M:%S")
                        print(f" {usuario} COMPLETADO (archivo desde {fecha_ultima_actualizacion})")
                
                except Exception as e:
                    print(f"Error al leer datos del jefe {usuario}: {str(e)}")
                    import traceback
                    traceback.print_exc()
            else:
                if not ha_validado:
                    print(f"DEBUG: {usuario} está PENDIENTE (no ha validado)")
                if len(zonas_codigos) == 0:
                    print(f"DEBUG: {usuario} no tiene zonas asignadas")
            
            estados.append({
                "usuario": usuario,
                "nombre": nombre,
                "cargo": cargo,
                "zonas": [z["nombre"] for z in zonas_completas],  # Solo nombres
                "zonas_codigos": zonas_codigos,  # Códigos para referencia
                "zonas_str": ", ".join([z["nombre"] for z in zonas_completas]) if zonas_completas else "Sin zonas",
                "completado": completado,
                "fecha_ultima_actualizacion": fecha_ultima_actualizacion,
                "total_vendedores": total_vendedores
            })
        
        # Ordenar: primero pendientes, luego completados
        estados.sort(key=lambda x: (x['completado'], x['nombre']))
        
        return {
            "estados": estados,
            "total_jefes": len(estados),
            "completados": sum(1 for e in estados if e['completado']),
            "pendientes": sum(1 for e in estados if not e['completado'])
        }
    
    except HTTPException:
        raise
    except Exception as e:
        print(f"ERROR en obtener_estados_jefes: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error al obtener estados de jefes: {str(e)}"
        )


@router.post("/consolidar")
async def consolidar_archivos(current_user: UserInfo = Depends(get_current_user)):
    """
    Consolida todos los archivos individuales de los jefes en un único archivo.
    Solo accesible para administradores.
    Compara con el archivo original para determinar estatus: Nuevo, Eliminado, Modificado, Antiguo
    """
    try:
        # Verificar que sea admin
        if current_user.cargo.upper() != 'ADMIN':
            raise HTTPException(status_code=403, detail="Solo administradores pueden consolidar archivos")
        
        # Buscar todos los archivos individuales de jefes
        archivos_jefes = list(CAMBIOS_DIR.glob("TBL_EstructuraVenta_*.xlsx"))
        
        if not archivos_jefes:
            raise HTTPException(status_code=404, detail="No hay archivos individuales para consolidar")
        
        print(f"📋 Encontrados {len(archivos_jefes)} archivos para consolidar")
        
        # Cargar el archivo original (Respaldo) para comparar
        if not EXCEL_ORIGINAL_PATH.exists():
            raise HTTPException(status_code=404, detail="Archivo original no encontrado para comparación")
        
        df_original = pd.read_excel(EXCEL_ORIGINAL_PATH, sheet_name='Dotacion')
        print(f"📄 Archivo original cargado: {len(df_original)} registros")
        
        # Usar el primer archivo como base (contiene todas las hojas)
        archivo_base = archivos_jefes[0]
        wb = openpyxl.load_workbook(archivo_base)
        
        # Leer la hoja Dotacion del primer archivo
        df_consolidado = pd.read_excel(archivo_base, sheet_name='Dotacion')
        print(f"✓ Cargado archivo base: {archivo_base.name} ({len(df_consolidado)} registros)")
        
        # Leer y consolidar el resto de archivos (solo hoja Dotacion)
        for archivo in archivos_jefes[1:]:
            try:
                df_jefe = pd.read_excel(archivo, sheet_name='Dotacion')
                
                # Obtener usuario del nombre del archivo
                usuario = archivo.stem.replace("TBL_EstructuraVenta_", "")
                
                # Obtener zonas del jefe
                zonas_jefe = obtener_zonas_usuario(usuario)
                
                if zonas_jefe and 'CodDistrito' in df_jefe.columns:
                    # Filtrar solo registros de las zonas del jefe
                    registros_jefe = df_jefe[df_jefe['CodDistrito'].isin(zonas_jefe)]
                    
                    # DEBUG: Ver estatus de los registros del jefe
                    if 'Estatus' in df_jefe.columns:
                        estatus_jefe = registros_jefe['Estatus'].value_counts().to_dict()
                        print(f"   📊 Estatus en archivo de {usuario}:")
                        for est, cnt in estatus_jefe.items():
                            print(f"      - {est}: {cnt}")
                    
                    # Eliminar registros de esas zonas del consolidado
                    df_consolidado = df_consolidado[~df_consolidado['CodDistrito'].isin(zonas_jefe)]
                    
                    # Agregar registros actualizados del jefe
                    df_consolidado = pd.concat([df_consolidado, registros_jefe], ignore_index=True)
                    
                    print(f"✓ Consolidado {archivo.name}: {len(registros_jefe)} registros de zonas {zonas_jefe}")
                else:
                    print(f"⚠ No se pudo consolidar {archivo.name}: sin zonas asignadas")
            
            except Exception as e:
                print(f"⚠ Error al procesar {archivo.name}: {str(e)}")
                continue
        
        # Mantener el estatus que ya viene en los archivos individuales
        print(f"\n🔍 Procesando estatus de registros...")
        
        # Asegurar que existe la columna Estatus y que todos los registros tengan un valor válido
        if 'Estatus' not in df_consolidado.columns:
            df_consolidado['Estatus'] = 'Antiguo'
        else:
            # Si la columna existe pero tiene valores nulos, asignar 'Antiguo'
            df_consolidado['Estatus'] = df_consolidado['Estatus'].fillna('Antiguo')
            # Si tiene valores vacíos, reemplazarlos por 'Antiguo'
            df_consolidado.loc[df_consolidado['Estatus'] == '', 'Estatus'] = 'Antiguo'
        
        # Crear una clave única para cada registro (CodVenta + Rut)
        df_original['_clave'] = df_original['CodVenta'].astype(str) + '_' + df_original['Rut'].astype(str)
        df_consolidado['_clave'] = df_consolidado['CodVenta'].astype(str) + '_' + df_consolidado['Rut'].astype(str)
        
        claves_originales = set(df_original['_clave'].tolist())
        
        contador_estatus = {'Nuevo': 0, 'Eliminado': 0, 'Modificado': 0, 'Antiguo': 0}
        
        # Contar los estatus que ya vienen en los archivos individuales
        for idx, row in df_consolidado.iterrows():
            estatus_actual = row.get('Estatus', 'Antiguo')
            
            # Validar que el estatus sea uno de los valores permitidos
            if estatus_actual not in contador_estatus:
                estatus_actual = 'Antiguo'
                df_consolidado.at[idx, 'Estatus'] = 'Antiguo'
            
            # Mantener el estatus que ya viene en el archivo individual
            # Los jefes ya asignaron el estatus correcto cuando guardaron sus cambios
            contador_estatus[estatus_actual] += 1
        
        # Agregar registros eliminados del original que no están en el consolidado
        # SOLO para las zonas que fueron validadas por los jefes
        print(f"\n🔍 Buscando registros eliminados...")
        
        # Obtener las zonas que fueron validadas (presentes en los archivos individuales)
        zonas_validadas = set()
        for archivo in archivos_jefes:
            usuario = archivo.stem.replace("TBL_EstructuraVenta_", "")
            zonas_jefe = obtener_zonas_usuario(usuario)
            if zonas_jefe:
                zonas_validadas.update(zonas_jefe)
        
        print(f"📍 Zonas validadas: {zonas_validadas}")
        
        claves_consolidado = set(df_consolidado['_clave'].tolist())
        registros_eliminados = []
        
        for clave in claves_originales:
            if clave not in claves_consolidado:
                # Este registro existía en el original pero ya no está
                registro_original = df_original[df_original['_clave'] == clave].iloc[0]
                
                # SOLO agregarlo como eliminado si pertenece a una zona validada
                if registro_original.get('CodDistrito') in zonas_validadas:
                    registro_dict = registro_original.to_dict()
                    registro_dict['Estatus'] = 'Eliminado'
                    registros_eliminados.append(registro_dict)
        
        if registros_eliminados:
            df_eliminados = pd.DataFrame(registros_eliminados)
            df_consolidado = pd.concat([df_consolidado, df_eliminados], ignore_index=True)
            contador_estatus['Eliminado'] += len(registros_eliminados)
            print(f"✓ Se agregaron {len(registros_eliminados)} registros eliminados de zonas validadas")
        
        # Eliminar columna temporal de clave
        df_consolidado = df_consolidado.drop(columns=['_clave'])
        
        # Ordenar por CodDistrito y NombreCompleto (si existe)
        if 'CodDistrito' in df_consolidado.columns:
            sort_columns = ['CodDistrito']
            if 'NombreCompleto' in df_consolidado.columns:
                sort_columns.append('NombreCompleto')
            elif 'Nombre' in df_consolidado.columns:
                sort_columns.append('Nombre')
            df_consolidado = df_consolidado.sort_values(sort_columns)
        
        # Crear nombre del archivo consolidado con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archivo_consolidado = CAMBIOS_DIR / f"TBL_EstructuraVentaCM_Consolidado_{timestamp}.xlsx"
        
        # Limpiar columna Nombre: si contiene los apellidos, dejar solo el nombre de pila
        if all(col in df_consolidado.columns for col in ['Nombre', 'APaterno', 'AMaterno']):
            _n  = df_consolidado['Nombre'].fillna('').astype(str).str.strip()
            _ap = df_consolidado['APaterno'].fillna('').astype(str).str.strip()
            _am = df_consolidado['AMaterno'].fillna('').astype(str).str.strip()

            def _limpiar_nombre_consolidado(n, ap, am):
                original = n
                if ap and ap in n:
                    n = n.replace(ap, '').strip()
                if am and am in n:
                    n = n.replace(am, '').strip()
                # Seguridad: si queda vacío, mantener el valor original
                return n if n else original

            df_consolidado['Nombre'] = np.vectorize(_limpiar_nombre_consolidado)(
                _n.values, _ap.values, _am.values
            )
            print(f"✓ Columna 'Nombre' limpiada: se removieron apellidos duplicados")

        # Guardar la hoja Dotacion actualizada
        with pd.ExcelWriter(archivo_consolidado, engine='openpyxl') as writer:
            # Copiar todas las hojas del workbook base
            for sheet_name in wb.sheetnames:
                if sheet_name == 'Dotacion':
                    # Escribir el DataFrame consolidado
                    df_consolidado.to_excel(writer, sheet_name='Dotacion', index=False)
                else:
                    # Copiar otras hojas sin modificar
                    df_sheet = pd.read_excel(archivo_base, sheet_name=sheet_name)
                    df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"\n✅ Archivo consolidado creado: {archivo_consolidado.name}")
        print(f"📊 Resumen de estatus:")
        print(f"   • Nuevos: {contador_estatus['Nuevo']}")
        print(f"   • Modificados: {contador_estatus['Modificado']}")
        print(f"   • Eliminados: {contador_estatus['Eliminado']}")
        print(f"   • Antiguos (sin cambios): {contador_estatus['Antiguo']}")
        print(f"   • TOTAL registros: {sum(contador_estatus.values())}")
        
        return {
            "success": True,
            "message": "Archivos consolidados exitosamente",
            "archivo": archivo_consolidado.name,
            "total_archivos_procesados": len(archivos_jefes),
            "total_registros": len(df_consolidado),
            "estatus": contador_estatus,
            "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    
    except HTTPException:
        raise
    except Exception as e:
        print(f"ERROR en consolidar: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error al consolidar archivos: {str(e)}"
        )


@router.get("/descargar-consolidado")
async def descargar_consolidado(current_user: UserInfo = Depends(get_current_user)):
    """
    Descarga el archivo consolidado más reciente.
    Solo accesible para administradores.
    """
    try:
        # Verificar que sea admin
        if current_user.cargo.upper() != 'ADMIN':
            raise HTTPException(status_code=403, detail="Solo administradores pueden descargar el consolidado")
        
        # Buscar el archivo consolidado más reciente
        archivos_consolidados = list(CAMBIOS_DIR.glob("TBL_EstructuraVentaCM_Consolidado_*.xlsx"))
        
        if not archivos_consolidados:
            raise HTTPException(status_code=404, detail="No hay archivo consolidado disponible. Debe consolidar primero.")
        
        # Ordenar por fecha de modificación y tomar el más reciente
        archivo_mas_reciente = max(archivos_consolidados, key=lambda x: x.stat().st_mtime)
        
        print(f" Descargando archivo consolidado: {archivo_mas_reciente.name}")
        
        return FileResponse(
            path=str(archivo_mas_reciente),
            filename=archivo_mas_reciente.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except HTTPException:
        raise
    except Exception as e:
        print(f"ERROR en descargar-consolidado: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error al descargar archivo consolidado: {str(e)}"
        )


@router.post("/limpiar-validaciones")
async def limpiar_validaciones(current_user: UserInfo = Depends(get_current_user)):
    """
    Elimina todos los archivos de la carpeta Cambios y backups antiguos de Respaldo
    para iniciar un nuevo ciclo mensual. Solo accesible para administradores.
    """
    try:
        # Verificar que sea admin
        if current_user.cargo.upper() != 'ADMIN':
            raise HTTPException(status_code=403, detail="Solo administradores pueden limpiar validaciones")
        
        # Limpiar TODOS los archivos de Cambios
        archivos_eliminados = limpiar_carpeta_cambios_estructura()
        
        # Limpiar backups antiguos de Respaldo (conserva el archivo base)
        backups_eliminados = limpiar_backups_respaldo_estructura()
        archivos_eliminados.extend(backups_eliminados)
        
        if not archivos_eliminados:
            return {
                "success": True,
                "message": "No hay archivos para limpiar",
                "archivos_eliminados": 0
            }
        
        return {
            "success": True,
            "message": f"Se eliminaron {len(archivos_eliminados)} archivos exitosamente. Nuevo ciclo iniciado.",
            "archivos_eliminados": len(archivos_eliminados),
            "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    
    except HTTPException:
        raise
    except Exception as e:
        print(f"ERROR en limpiar_validaciones: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error al limpiar validaciones: {str(e)}"
        )


@router.post("/upload-base-file")
async def upload_base_file(
    file: UploadFile = File(...),
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Endpoint para subir y reemplazar el archivo base de Estructura de Ventas.
    Solo accesible para administradores.
    Usado al inicio de cada ciclo mensual.
    
    IMPORTANTE: Al subir un nuevo archivo base, se eliminan automáticamente TODOS los archivos
    individuales de los jefes de venta, reiniciando el proceso de validación desde cero.
    """
    try:
        # Verificar permisos de administrador
        if current_user.cargo.upper() != 'ADMIN':
            raise HTTPException(status_code=403, detail="Solo administradores pueden subir el archivo base")
        
        # Validar extensión del archivo
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=400,
                detail="El archivo debe ser un archivo Excel (.xlsx o .xls)"
            )
        
        # Verificar que el archivo tenga contenido
        contents = await file.read()
        if len(contents) == 0:
            raise HTTPException(
                status_code=400,
                detail="El archivo está vacío"
            )
        
        # Hacer backup del archivo anterior si existe
        if EXCEL_ORIGINAL_PATH.exists():
            backup_path = EXCEL_ORIGINAL_PATH.parent / f"TBL EstructuraVentaCM_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            shutil.copy2(EXCEL_ORIGINAL_PATH, backup_path)
            print(f"✓ Backup creado: {backup_path.name}")
        
        # Guardar el nuevo archivo
        with open(EXCEL_ORIGINAL_PATH, 'wb') as f:
            f.write(contents)
        
        # Validar que el archivo se puede leer correctamente
        try:
            wb = openpyxl.load_workbook(EXCEL_ORIGINAL_PATH)
            sheet_names = wb.sheetnames
            wb.close()
            print(f"✓ Archivo cargado exitosamente. Hojas encontradas: {sheet_names}")
        except Exception as e:
            # Si hay error al leer, restaurar el backup
            if EXCEL_ORIGINAL_PATH.exists():
                EXCEL_ORIGINAL_PATH.unlink()
            raise HTTPException(
                status_code=400,
                detail=f"El archivo no es un Excel válido: {str(e)}"
            )
        
        # Validar columnas obligatorias (buscar en cualquier hoja, no solo 'Dotacion')
        try:
            columnas_obligatorias = [
                'Año', 'Mes', 'CodDistrito', 'DesDistrito', 'CodOficina', 'DesOficina',
                'CodVenta', 'Cargo', 'Rut', 'Nombre', 'APaterno', 'AMaterno',
                'Telefono', 'Correo', 'ZonaEstival', 'Genero', 'TallaPantalon', 'TallaCamisa'
            ]
            
            hoja_dotacion = None
            xls = pd.ExcelFile(EXCEL_ORIGINAL_PATH)
            reporte_hojas = {}
            
            for nombre_hoja in xls.sheet_names:
                df_test = pd.read_excel(xls, sheet_name=nombre_hoja, nrows=0)
                columnas_hoja = list(df_test.columns)
                faltantes = [col for col in columnas_obligatorias if col not in columnas_hoja]
                encontradas = [col for col in columnas_obligatorias if col in columnas_hoja]
                reporte_hojas[nombre_hoja] = {
                    "columnas_encontradas": encontradas,
                    "columnas_faltantes": faltantes,
                    "columnas_archivo": columnas_hoja
                }
                if not faltantes and hoja_dotacion is None:
                    hoja_dotacion = nombre_hoja
                    if nombre_hoja != 'Dotacion':
                        print(f"ℹ Hoja '{nombre_hoja}' detectada con las columnas correctas, se renombrará a 'Dotacion'")
            
            xls.close()
            
            if hoja_dotacion is None:
                if EXCEL_ORIGINAL_PATH.exists():
                    EXCEL_ORIGINAL_PATH.unlink()
                # Construir mensaje simplificado por hoja
                detalle_hojas = []
                for nombre_hoja, info in reporte_hojas.items():
                    if info['columnas_faltantes']:
                        faltantes_str = ', '.join(info['columnas_faltantes'])
                        detalle_hojas.append(
                            f"Hoja '{nombre_hoja}': faltan las columnas: {faltantes_str}"
                        )
                raise HTTPException(
                    status_code=400,
                    detail={
                        "mensaje": f"Ninguna hoja del archivo tiene todas las columnas requeridas.",
                        "columnas_faltantes_por_hoja": detalle_hojas,
                        "sugerencia": f"El archivo debe tener una hoja con todas estas columnas: {', '.join(columnas_obligatorias)}"
                    }
                )
            
            columnas_hoja_encontrada = reporte_hojas[hoja_dotacion]['columnas_archivo']
            columnas_extra = [col for col in columnas_hoja_encontrada if col not in columnas_obligatorias]
            if columnas_extra:
                print(f"ℹ Columnas adicionales detectadas (se preservarán): {columnas_extra}")
            
            # Si la hoja no se llama 'Dotacion', renombrarla para compatibilidad interna
            if hoja_dotacion != 'Dotacion':
                wb_rename = openpyxl.load_workbook(EXCEL_ORIGINAL_PATH)
                wb_rename[hoja_dotacion].title = 'Dotacion'
                wb_rename.save(EXCEL_ORIGINAL_PATH)
                wb_rename.close()
                print(f"✓ Hoja '{hoja_dotacion}' renombrada a 'Dotacion'")
                # Actualizar sheet_names para el response
                sheet_names = [('Dotacion' if s == hoja_dotacion else s) for s in sheet_names]
                
        except HTTPException:
            raise
        except Exception as e:
            if EXCEL_ORIGINAL_PATH.exists():
                EXCEL_ORIGINAL_PATH.unlink()
            raise HTTPException(
                status_code=400,
                detail=f"Error al validar el archivo: {str(e)}"
            )
        
        # PASO CRÍTICO: Eliminar todos los archivos de Cambios y backups de Respaldo
        # Esto fuerza a todos a comenzar desde el nuevo archivo base
        print(f"\n🔄 INICIANDO LIMPIEZA DE ARCHIVOS...")
        
        archivos_eliminados = limpiar_carpeta_cambios_estructura()
        backups_eliminados = limpiar_backups_respaldo_estructura()
        archivos_eliminados.extend(backups_eliminados)
        
        print(f"✓ Limpieza completada: {len(archivos_eliminados)} archivos eliminados")
        
        mensaje_limpieza = f"Archivo base actualizado y {len(archivos_eliminados)} archivos anteriores eliminados."
        
        return {
            "success": True,
            "message": mensaje_limpieza,
            "filename": file.filename,
            "size": len(contents),
            "sheets": sheet_names,
            "archivos_jefes_eliminados": len(archivos_eliminados),
            "archivos_eliminados": len(archivos_eliminados),
            "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "info": "Todos los jefes de venta deberán validar nuevamente su estructura desde el archivo base actualizado."
        }
    
    except HTTPException:
        raise
    except Exception as e:
        print(f"ERROR en upload_base_file: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error al subir archivo: {str(e)}"
        )


@router.get("/estados-jefes-combinados")
async def obtener_estados_jefes_combinados(current_user: UserInfo = Depends(get_current_user)):
    """
    Obtiene el estado de validación combinado de Estructura de Venta y Carteras
    para todos los jefes de venta. Solo accesible para administradores.
    """
    try:
        # Verificar que sea admin
        if current_user.cargo.upper() != 'ADMIN':
            raise HTTPException(status_code=403, detail="Solo administradores pueden ver estados")
        
        # Verificar que existe el archivo de usuarios
        if not USUARIOS_SISTEMA_PATH.exists():
            raise HTTPException(status_code=404, detail=f"Archivo de usuarios no encontrado")
        
        # Leer usuarios del sistema
        df_usuarios = pd.read_excel(USUARIOS_SISTEMA_PATH)
        
        # Filtrar solo Jefes de Venta
        jefes_venta = df_usuarios[df_usuarios['Cargo'].str.upper().str.contains('JEFE', na=False)].copy()
        
        print(f"DEBUG: Jefes encontrados: {len(jefes_venta)}")
        
        # Paths para Carteras
        CARTERAS_CAMBIOS_DIR = Path(__file__).parent.parent.parent / "data" / "carga_cm" / "Cartera" / "Cambios"
        
        estados = []
        
        for _, jefe in jefes_venta.iterrows():
            usuario = jefe['Usuario']
            nombre = jefe['Nombre']
            cargo = jefe['Cargo']
            
            print(f"DEBUG: Procesando jefe: {usuario} - {nombre}")
            
            # Obtener zonas asignadas
            zonas_codigos = obtener_zonas_usuario(usuario)
            zonas_completas = obtener_nombres_zonas(zonas_codigos)
            
            # Verificar validación de Estructura de Venta por cada zona
            zonas_estructura_validada = {}
            for codigo_zona in zonas_codigos:
                zonas_estructura_validada[codigo_zona] = verificar_validacion_jefe_por_archivo(usuario, codigo_zona)
            
            # estructura_validada = True solo si TODAS las zonas están validadas
            estructura_validada = bool(zonas_codigos) and all(zonas_estructura_validada.values())
            
            # Verificar validación de Carteras (por zona, igual que Estructura de Venta)
            if zonas_codigos:
                zonas_cartera_validada = {}
                for codigo_zona in zonas_codigos:
                    archivo_estado_cartera = CARTERAS_CAMBIOS_DIR / f"VALIDADO_{usuario}_{codigo_zona}.json"
                    if archivo_estado_cartera.exists():
                        try:
                            with open(archivo_estado_cartera, 'r', encoding='utf-8') as f:
                                estado_c = json.load(f)
                            zonas_cartera_validada[codigo_zona] = estado_c.get('validado', False)
                        except Exception:
                            zonas_cartera_validada[codigo_zona] = False
                    else:
                        zonas_cartera_validada[codigo_zona] = False
                cartera_validada = all(zonas_cartera_validada.values())
            else:
                zonas_cartera_validada = {}
                cartera_validada = False
            
            estados.append({
                "usuario": usuario,
                "nombre": nombre,
                "cargo": cargo,
                "zonas": [z["nombre"] for z in zonas_completas],
                "zonas_codigos": zonas_codigos,
                "zonas_estructura_validada": zonas_estructura_validada,
                "zonas_cartera_validada": zonas_cartera_validada,
                "zonas_str": ", ".join([z["nombre"] for z in zonas_completas]) if zonas_completas else "Sin zonas",
                "estructura_validada": estructura_validada,
                "cartera_validada": cartera_validada,
                "ambos_validados": estructura_validada and cartera_validada
            })
        
        # Ordenar: primero los que no han validado ambos, luego los que sí
        estados.sort(key=lambda x: (x['ambos_validados'], x['nombre']))
        
        return {
            "estados": estados,
            "total_jefes": len(estados),
            "estructura_completados": sum(1 for e in estados if e['estructura_validada']),
            "cartera_completados": sum(1 for e in estados if e['cartera_validada']),
            "ambos_completados": sum(1 for e in estados if e['ambos_validados']),
            "estructura_pendientes": sum(1 for e in estados if not e['estructura_validada']),
            "cartera_pendientes": sum(1 for e in estados if not e['cartera_validada'])
        }
    
    except HTTPException:
        raise
    except Exception as e:
        print(f"ERROR en obtener_estados_jefes_combinados: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error al obtener estados combinados: {str(e)}"
        )


@router.post("/consolidar-combinado")
async def consolidar_combinado(current_user: UserInfo = Depends(get_current_user)):
    """
    Consolida los archivos de Estructura de Venta y Carteras en un único Excel
    con dos hojas: 'Estructura de Venta' y 'Carteras'.
    Solo accesible para administradores.
    """
    try:
        # Verificar que sea admin
        if current_user.cargo.upper() != 'ADMIN':
            raise HTTPException(status_code=403, detail="Solo administradores pueden consolidar archivos")
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        archivo_consolidado = CAMBIOS_DIR / f"Consolidado_Completo_{timestamp}.xlsx"
        
        # Paths para Carteras
        CARTERAS_CAMBIOS_DIR = Path(__file__).parent.parent.parent / "data" / "carga_cm" / "Cartera" / "Cambios"
        CARTERAS_BASE_FILE = Path(__file__).parent.parent.parent / "data" / "carga_cm" / "Cartera" / "Respaldo" / "TBL_Cartera.xlsx"
        
        # ===========================================
        # CONSOLIDAR ESTRUCTURA DE VENTA
        # ===========================================
        
        # Buscar todos los archivos individuales de jefes
        archivos_estructura = list(CAMBIOS_DIR.glob("TBL_EstructuraVenta_*.xlsx"))
        
        if not archivos_estructura:
            raise HTTPException(status_code=404, detail="No hay archivos de Estructura de Venta para consolidar")
        
        print(f"📋 Encontrados {len(archivos_estructura)} archivos de Estructura de Venta")
        
        # Cargar el archivo original para comparar
        if not EXCEL_ORIGINAL_PATH.exists():
            raise HTTPException(status_code=404, detail="Archivo original de Estructura de Venta no encontrado")
        
        df_original = pd.read_excel(EXCEL_ORIGINAL_PATH, sheet_name='Dotacion')
        print(f"📄 Archivo original de Estructura de Venta cargado: {len(df_original)} registros")
        
        # Usar el primer archivo como base
        archivo_base = archivos_estructura[0]
        df_estructura_consolidado = pd.read_excel(archivo_base, sheet_name='Dotacion')
        print(f"✓ Cargado archivo base de Estructura de Venta: {archivo_base.name} ({len(df_estructura_consolidado)} registros)")
        
        # Consolidar el resto de archivos
        for archivo in archivos_estructura[1:]:
            try:
                df_jefe = pd.read_excel(archivo, sheet_name='Dotacion')
                usuario = archivo.stem.replace("TBL_EstructuraVenta_", "")
                zonas_jefe = obtener_zonas_usuario(usuario)
                
                if zonas_jefe and 'CodDistrito' in df_jefe.columns:
                    registros_jefe = df_jefe[df_jefe['CodDistrito'].isin(zonas_jefe)]
                    df_estructura_consolidado = df_estructura_consolidado[~df_estructura_consolidado['CodDistrito'].isin(zonas_jefe)]
                    df_estructura_consolidado = pd.concat([df_estructura_consolidado, registros_jefe], ignore_index=True)
                    print(f"✓ Consolidado {archivo.name}: {len(registros_jefe)} registros")
            except Exception as e:
                print(f"⚠ Error al procesar {archivo.name}: {str(e)}")
                continue
        
        # Asegurar que existe la columna Estatus
        if 'Estatus' not in df_estructura_consolidado.columns:
            df_estructura_consolidado['Estatus'] = 'Antiguo'
        else:
            df_estructura_consolidado['Estatus'] = df_estructura_consolidado['Estatus'].fillna('Antiguo')
            df_estructura_consolidado.loc[df_estructura_consolidado['Estatus'] == '', 'Estatus'] = 'Antiguo'
        
        # ===========================================
        # CONSOLIDAR CARTERAS
        # ===========================================
        
        # CARGAR ARCHIVO BASE (MAESTRO) CON TODOS LOS CLIENTES DE TODAS LAS ZONAS
        if not CARTERAS_BASE_FILE.exists():
            raise HTTPException(status_code=404, detail="Archivo base de Carteras no encontrado")
        
        df_carteras_maestro = pd.read_excel(CARTERAS_BASE_FILE)
        print(f"📄 Archivo maestro de Carteras cargado: {len(df_carteras_maestro)} registros")
        
        # Filtrar solo las 3 columnas requeridas: CodDistrito, CodVend, CodCliente
        columnas_requeridas = ['CodDistrito', 'CodVend', 'CodCliente']
        columnas_disponibles = [col for col in columnas_requeridas if col in df_carteras_maestro.columns]
        
        if len(columnas_disponibles) != len(columnas_requeridas):
            columnas_faltantes = set(columnas_requeridas) - set(columnas_disponibles)
            raise HTTPException(
                status_code=400, 
                detail=f"Faltan columnas requeridas en archivo maestro de Carteras: {columnas_faltantes}"
            )
        
        # Crear DataFrame consolidado SOLO con las 3 columnas (todos los registros del maestro)
        df_carteras_consolidado = df_carteras_maestro[columnas_requeridas].copy()
        print(f"✓ Base de Carteras creada con {len(df_carteras_consolidado)} registros de todas las zonas")
        
        # Buscar archivos validados de jefes de venta
        archivos_carteras = list(CARTERAS_CAMBIOS_DIR.glob("TBL_Cartera_*.xlsx"))
        
        if archivos_carteras:
            print(f"📋 Encontrados {len(archivos_carteras)} archivos validados de Carteras")
            
            # Actualizar con los datos validados de cada jefe
            for archivo in archivos_carteras:
                try:
                    df_jefe_cartera = pd.read_excel(archivo)
                    
                    # Verificar que tenga las columnas necesarias
                    if not all(col in df_jefe_cartera.columns for col in columnas_requeridas):
                        print(f"⚠ Archivo {archivo.name} no tiene todas las columnas requeridas, omitiendo")
                        continue
                    
                    # Obtener zonas directamente del archivo (ya fue validado durante la carga)
                    zonas_archivo = df_jefe_cartera['CodDistrito'].dropna().unique().tolist()
                    
                    if zonas_archivo:
                        # Filtrar solo las 3 columnas del archivo validado
                        df_jefe_filtrado = df_jefe_cartera[columnas_requeridas].copy()
                        
                        # Eliminar registros anteriores de esas zonas
                        df_carteras_consolidado = df_carteras_consolidado[~df_carteras_consolidado['CodDistrito'].isin(zonas_archivo)]
                        
                        # Agregar los registros validados
                        df_carteras_consolidado = pd.concat([df_carteras_consolidado, df_jefe_filtrado], ignore_index=True)
                        print(f"✓ Actualizado con {archivo.name}: {len(df_jefe_filtrado)} registros validados de zonas {zonas_archivo}")
                except Exception as e:
                    print(f"⚠ Error al procesar {archivo.name}: {str(e)}")
                    continue
        else:
            print("ℹ No hay archivos validados de Carteras, usando datos del maestro")
        
        # ===========================================
        # GUARDAR ARCHIVO CONSOLIDADO CON DOS HOJAS
        # ===========================================
        
        with pd.ExcelWriter(archivo_consolidado, engine='openpyxl') as writer:
            df_estructura_consolidado.to_excel(writer, sheet_name='Estructura de Venta', index=False)
            if not df_carteras_consolidado.empty:
                df_carteras_consolidado.to_excel(writer, sheet_name='Carteras', index=False)
        
        print(f"✅ Archivo consolidado completo creado: {archivo_consolidado.name}")
        print(f"   Hoja 'Estructura de Venta': {len(df_estructura_consolidado)} registros")
        print(f"   Hoja 'Carteras': {len(df_carteras_consolidado)} registros")
        
        return {
            "success": True,
            "message": f"Consolidación exitosa",
            "archivo": archivo_consolidado.name,
            "total_archivos_estructura": len(archivos_estructura),
            "total_archivos_carteras": len(archivos_carteras),
            "registros_estructura": len(df_estructura_consolidado),
            "registros_carteras": len(df_carteras_consolidado),
            "vendedores_estructura": int(df_estructura_consolidado['CodVenta'].nunique()) if 'CodVenta' in df_estructura_consolidado.columns else 0,
            "clientes_carteras": int(df_carteras_consolidado['CodCliente'].nunique()) if not df_carteras_consolidado.empty and 'CodCliente' in df_carteras_consolidado.columns else 0
        }
    
    except HTTPException:
        raise
    except Exception as e:
        print(f"ERROR en consolidar_combinado: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Error al consolidar archivos: {str(e)}"
        )


@router.get("/descargar-consolidado-combinado")
async def descargar_consolidado_combinado(current_user: UserInfo = Depends(get_current_user)):
    """
    Descarga el último archivo consolidado combinado (con ambas hojas).
    Solo accesible para administradores.
    """
    try:
        # Verificar que sea admin
        if current_user.cargo.upper() != 'ADMIN':
            raise HTTPException(status_code=403, detail="Solo administradores pueden descargar el consolidado")
        
        # Buscar el archivo consolidado más reciente
        archivos_consolidados = list(CAMBIOS_DIR.glob("Consolidado_Completo_*.xlsx"))
        
        if not archivos_consolidados:
            raise HTTPException(status_code=404, detail="No hay archivos consolidados disponibles")
        
        # Obtener el más reciente
        archivo_consolidado = max(archivos_consolidados, key=lambda p: p.stat().st_mtime)
        
        return FileResponse(
            path=str(archivo_consolidado),
            filename=f"Consolidado_Completo_{datetime.now().strftime('%Y%m%d')}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except HTTPException:
        raise
    except Exception as e:
        print(f"ERROR en descargar_consolidado_combinado: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al descargar archivo consolidado: {str(e)}"
        )
