"""
Rutas de gestión de rutas y optimización
Endpoints: /api/routes/*
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse, JSONResponse, Response
from app.models.user import UserInfo
from app.utils.dependencies import get_current_user
from app.services import route_service
from app.services.user_service import get_all_users, get_distritos_from_maestro, get_jefe_venta_from_maestro
from app.services.pdf_service import pdf_service
from app.routes.config import load_config
import json
import asyncio
import shutil
import logging
import httpx
from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo
import os

# Logger
logger = logging.getLogger(__name__)

# Crear router de rutas
router = APIRouter()

@router.post("/calculate-routes")
async def calculate_routes(current_user: UserInfo = Depends(get_current_user)):
    """
    Calcular rutas desde el archivo Maestro APP.csv usando OSRM con streaming de progreso
    Solo accesible para usuarios Admin
    """
    # Verificar que sea Admin
    if current_user.cargo.lower() != "admin":
        raise HTTPException(status_code=403, detail="Solo usuarios Admin pueden calcular rutas")
    
    async def event_generator():
        """Generador de eventos SSE para el progreso"""
        try:
            # Variable para almacenar resultados
            results = None
            
            def progress_callback(progress_data):
                """Callback que se llama para cada distrito procesado"""
                nonlocal results
                # Enviar evento de progreso
                event_data = {
                    "type": "progress",
                    "data": progress_data
                }
                return json.dumps(event_data)
            
            # Procesar rutas con callback de progreso
            def process_with_callback():
                nonlocal results
                progress_events = []
                
                def callback(data):
                    progress_events.append(data)
                
                results = route_service.process_maestro_csv(progress_callback=callback)
                return progress_events
            
            # Ejecutar en thread pool para no bloquear
            loop = asyncio.get_event_loop()
            progress_events = await loop.run_in_executor(None, process_with_callback)
            
            # Enviar eventos de progreso
            for event in progress_events:
                yield f"data: {json.dumps({'type': 'progress', 'data': event})}\n\n"
                await asyncio.sleep(0.1)  # Pequeña pausa para visualizar progreso
            
            # Enviar evento final con resultados
            yield f"data: {json.dumps({'type': 'complete', 'data': results})}\n\n"
            
        except FileNotFoundError as e:
            yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'type': 'error', 'message': f'Error al calcular rutas: {str(e)}'})}\n\n"
    
    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )

@router.post("/upload-routes")
async def upload_routes(
    file: UploadFile = File(...),
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Subir archivo de rutas (CSV) al servidor
    Solo accesible para usuarios Admin
    Guarda el archivo en backend/data/rutas/
    """
    # Verificar que sea Admin
    if current_user.cargo.lower() != "admin":
        raise HTTPException(status_code=403, detail="Solo usuarios Admin pueden cargar archivos")
    
    # Validar extensión del archivo (solo CSV)
    file_ext = Path(file.filename).suffix.lower()
    
    if file_ext != ".csv":
        raise HTTPException(
            status_code=400,
            detail="Formato no permitido. Solo se acepta: .csv"
        )
    
    try:
        # Definir ruta de destino
        upload_dir = Path(__file__).parent.parent.parent / "data" / "rutas"
        upload_dir.mkdir(parents=True, exist_ok=True)
        
        file_path = upload_dir / file.filename
        
        # Eliminar versiones anteriores de Maestro APP solo si no es el archivo destino
        if "Maestro APP" in file.filename:
            for old_file in upload_dir.glob("Maestro APP.*"):
                # No intentar eliminar si es el mismo archivo (evitar error de permisos)
                if old_file.exists() and old_file != file_path:
                    try:
                        old_file.unlink()
                    except Exception as e:
                        print(f"Warning: No se pudo eliminar {old_file}: {e}")
        
        # Guardar archivo (sobrescribirá si ya existe)
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        return {
            "success": True,
            "message": f"Archivo '{file.filename}' cargado exitosamente",
            "filename": file.filename,
            "path": str(file_path),
            "size_bytes": file_path.stat().st_size
        }
    
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al guardar archivo: {str(e)}"
        )
    finally:
        file.file.close()

@router.get("/ranking-vendedores")
async def get_ranking_vendedores(current_user: UserInfo = Depends(get_current_user)):
    """
    Obtener ranking de vendedores con comparación entre rutas originales y optimizadas
    Filtra por distritos permitidos del usuario
    
    Returns:
        Ranking con métricas por vendedor y totales generales
    """
    try:
        routes_dir = Path(__file__).parent.parent.parent / "data" / "routes_json"
        
        if not routes_dir.exists():
            return {
                "vendedores": [],
                "metricas_totales": {
                    "total_vendedores": 0,
                    "km_sap_total": 0,
                    "km_optimizado_total": 0,
                    "ahorro_km": 0,
                    "ahorro_porcentaje": 0
                }
            }
        
        # Cargar umbral de porcentaje desde config.json
        config_file = Path(__file__).parent.parent.parent / "data" / "config.json"
        umbral_porcentaje = 10.0  # Valor por defecto
        if config_file.exists():
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config_data = json.load(f)
                    umbral_porcentaje = config_data.get("umbral_porcentaje", 10.0)
            except Exception as e:
                logger.warning(f"Error al cargar umbral desde config.json: {e}")
        
        vendedores_dict = {}
        distritos_permitidos = current_user.distritos_permitidos
        
        # Iterar sobre todas las carpetas de distritos
        for distrito_dir in routes_dir.iterdir():
            if not distrito_dir.is_dir():
                continue
            
            codigo_distrito = distrito_dir.name
            
            # Filtrar por distritos permitidos (si no es admin)
            if current_user.cargo.lower() != "admin" and codigo_distrito not in distritos_permitidos:
                continue
            
            # Leer archivos JSON
            ruta_original_path = distrito_dir / "ruta_original.json"
            ruta_optimizada_path = distrito_dir / "ruta_optimizada.json"
            ruta_validada_path = distrito_dir / "ruta_validada.json"
            
            if not ruta_original_path.exists() or not ruta_optimizada_path.exists():
                continue
            
            with open(ruta_original_path, 'r', encoding='utf-8') as f:
                data_original = json.load(f)
            
            with open(ruta_optimizada_path, 'r', encoding='utf-8') as f:
                data_optimizada = json.load(f)
            
            # Leer rutas validadas si existen
            data_validada = None
            if ruta_validada_path.exists():
                with open(ruta_validada_path, 'r', encoding='utf-8') as f:
                    data_validada = json.load(f)
            
            nombre_distrito = data_original.get("distrito", codigo_distrito)
            
            # Procesar cada ruta del día
            for ruta_orig, ruta_opt in zip(data_original.get("rutas_por_dia", []), 
                                           data_optimizada.get("rutas_por_dia", [])):
                vendedor = ruta_orig.get("vendedor", {})
                cod_vend = str(vendedor.get("codigo", ""))
                nombre_vend = vendedor.get("nombre", "Sin nombre")
                dia = ruta_orig.get("dia", "")
                
                # Buscar si existe ruta validada para este día y vendedor
                ruta_a_usar = ruta_opt  # Por defecto usar optimizada
                if data_validada:
                    for ruta_val in data_validada.get("rutas_por_dia", []):
                        vendedor_val = ruta_val.get("vendedor", {})
                        if (str(vendedor_val.get("codigo", "")) == cod_vend and 
                            ruta_val.get("dia", "") == dia):
                            ruta_a_usar = ruta_val  # Usar validada si existe
                            break
                
                # Crear clave única por vendedor
                if cod_vend not in vendedores_dict:
                    # Verificar si existe ruta guardada para este vendedor específico en este distrito
                    ruta_guardada_path = routes_dir / codigo_distrito / "ruta_guardada.json"
                    tiene_guardado = False
                    rutas_guardadas_vendedor = {}  # Diccionario {dia: km} de las rutas guardadas
                    
                    if ruta_guardada_path.exists():
                        try:
                            with open(ruta_guardada_path, 'r', encoding='utf-8') as f:
                                data_guardada = json.load(f)
                                # Verificar si el vendedor está en el archivo guardado
                                # El archivo guardado usa "rutas" en lugar de "rutas_por_dia"
                                for ruta_guardada in data_guardada.get("rutas", []):
                                    if str(ruta_guardada.get("vendedor", {}).get("codigo", "")) == cod_vend:
                                        tiene_guardado = True
                                        # Guardar km de cada día para comparar después
                                        dia_guardado = ruta_guardada.get("dia", "")
                                        km_guardado = ruta_guardada.get("distancia_km", 0)
                                        rutas_guardadas_vendedor[dia_guardado] = km_guardado
                        except Exception as e:
                            logger.warning(f"Error al verificar ruta guardada para vendedor {cod_vend}: {e}")
                    
                    vendedores_dict[cod_vend] = {
                        "cod_vend": cod_vend,
                        "nombre_vendedor": nombre_vend,
                        "distrito": nombre_distrito,
                        "codigo_distrito": codigo_distrito,
                        "km_sap_total": 0,
                        "km_optimizado_total": 0,
                        "km_ruta_total": 0,
                        "km_holgura_total": 0,  # Acumular holgura individual
                        "km_validado_total": 0,  # Acumular validado individual
                        "dias": set(),
                        "rutas_porcentajes": [],  # Para verificar si todas están bajo umbral
                        "rutas_diferencias": [],  # Para verificar si hay diferencias negativas
                        "tiene_guardado": tiene_guardado,
                        "rutas_guardadas": rutas_guardadas_vendedor,  # Para comparar después
                        "tiene_cambios_pendientes": False  # Se actualizará después
                    }
                
                # Calcular porcentaje de esta ruta individual
                km_sap_ruta = ruta_orig.get("distancia_km", 0)
                km_ruta_ruta = ruta_a_usar.get("distancia_km", 0)
                
                # Calcular holgura y validado para cada ruta
                km_holgura_ruta = round(km_ruta_ruta * (umbral_porcentaje / 100), 2)
                km_validado_ruta = round(km_ruta_ruta + km_holgura_ruta, 2)
                diferencia_km_ruta = round(km_validado_ruta - km_sap_ruta, 2)
                
                porcentaje_ruta = round((diferencia_km_ruta / km_sap_ruta * 100), 2) if km_sap_ruta > 0 else 0
                
                # Acumular kilómetros
                vendedores_dict[cod_vend]["km_sap_total"] += ruta_orig.get("distancia_km", 0)
                vendedores_dict[cod_vend]["km_optimizado_total"] += ruta_opt.get("distancia_km", 0)
                vendedores_dict[cod_vend]["km_ruta_total"] += ruta_a_usar.get("distancia_km", 0)
                vendedores_dict[cod_vend]["km_holgura_total"] += km_holgura_ruta  # Acumular holgura
                vendedores_dict[cod_vend]["km_validado_total"] += km_validado_ruta  # Acumular validado
                vendedores_dict[cod_vend]["dias"].add(dia)
                vendedores_dict[cod_vend]["rutas_porcentajes"].append(porcentaje_ruta)
                vendedores_dict[cod_vend]["rutas_diferencias"].append(diferencia_km_ruta)
                
                # Verificar si hay cambios respecto a la ruta guardada
                if vendedores_dict[cod_vend]["tiene_guardado"] and dia in vendedores_dict[cod_vend]["rutas_guardadas"]:
                    km_guardado = vendedores_dict[cod_vend]["rutas_guardadas"][dia]
                    # Comparar con tolerancia de 0.01 km por redondeos
                    if abs(km_ruta_ruta - km_guardado) > 0.01:
                        vendedores_dict[cod_vend]["tiene_cambios_pendientes"] = True
        
        # Convertir a lista y calcular métricas
        vendedores_list = []
        for vend_data in vendedores_dict.values():
            km_sap = round(vend_data["km_sap_total"], 2)
            km_opt = round(vend_data["km_optimizado_total"], 2)
            km_ruta = round(vend_data["km_ruta_total"], 2)
            
            # Usar los valores acumulados de holgura y validado (ya redondeados individualmente)
            km_holgura = round(vend_data["km_holgura_total"], 2)
            km_validado = round(vend_data["km_validado_total"], 2)
            
            # Diferencia calculada con km_validado - km_sap (negativo = ahorro)
            diferencia_km = round(km_validado - km_sap, 2)
            # Porcentaje: (km_validado - km_sap) / km_sap * 100 (negativo = ahorro)
            porcentaje_dif = round((diferencia_km / km_sap * 100), 2) if km_sap > 0 else 0
            
            # Verificar si todas las diferencias son >= 0 (Verde) o al menos una es negativa (Rojo)
            todas_bajo_umbral = all(
                dif >= 0
                for dif in vend_data["rutas_diferencias"]
            ) if vend_data["rutas_diferencias"] else False
            
            vendedores_list.append({
                "distrito": vend_data["distrito"],
                "codigo_distrito": vend_data["codigo_distrito"],
                "cod_vend": vend_data["cod_vend"],
                "vendedor": vend_data["nombre_vendedor"],
                "km_sap": km_sap,
                "km_optimizado": km_opt,
                "km_ruta": km_ruta,
                "km_holgura": km_holgura,
                "km_validado": km_validado,
                "diferencia_km": diferencia_km,
                "porcentaje_diferencia": porcentaje_dif,
                "dias": len(vend_data["dias"]),
                "todas_rutas_bajo_umbral": todas_bajo_umbral,
                "tiene_guardado": vend_data.get("tiene_guardado", False),
                "tiene_cambios_pendientes": vend_data.get("tiene_cambios_pendientes", False)
            })
        
        # Ordenar por mayor diferencia en km (negativo = mayor ahorro)
        vendedores_list.sort(key=lambda x: x["diferencia_km"], reverse=False)
        
        # Calcular métricas totales
        total_km_sap = sum(v["km_sap"] for v in vendedores_list)
        total_km_opt = sum(v["km_optimizado"] for v in vendedores_list)
        total_km_val = sum(v["km_ruta"] for v in vendedores_list)
        # Ahorro calculado con km_ruta - km_sap (negativo = ahorro)
        ahorro_total = total_km_val - total_km_sap
        # Porcentaje: (km_ruta - km_sap) / km_sap * 100 (negativo = ahorro)
        ahorro_porcentaje = round((ahorro_total / total_km_sap * 100), 2) if total_km_sap > 0 else 0
        
        return {
            "vendedores": vendedores_list,
            "metricas_totales": {
                "total_vendedores": len(vendedores_list),
                "km_sap_total": round(total_km_sap, 2),
                "km_optimizado_total": round(total_km_opt, 2),
                "km_ruta_total": round(total_km_val, 2),
                "ahorro_km": round(ahorro_total, 2),
                "ahorro_porcentaje": ahorro_porcentaje
            }
        }
    
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al calcular ranking: {str(e)}"
        )


@router.get("/vendedor/{codigo_vendedor}")
async def get_vendedor_detalles(
    codigo_vendedor: str,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Obtener detalles completos de un vendedor específico
    
    Args:
        codigo_vendedor: Código del vendedor
        current_user: Usuario actual
        
    Returns:
        Información detallada del vendedor con todas sus rutas por distrito y día
    """
    try:
        routes_dir = Path(__file__).parent.parent.parent / "data" / "routes_json"
        
        if not routes_dir.exists():
            raise HTTPException(
                status_code=404,
                detail="No se encontraron datos de rutas"
            )
        
        # Cargar umbral de porcentaje (holgura) desde config.json
        config_file = Path(__file__).parent.parent.parent / "data" / "config.json"
        umbral_porcentaje = 10.0  # Valor por defecto
        if config_file.exists():
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config_data = json.load(f)
                    umbral_porcentaje = config_data.get("umbral_porcentaje", 10.0)
            except Exception as e:
                logger.warning(f"Error al cargar umbral desde config.json: {e}")
        
        vendedor_info = None
        rutas_detalle = []
        distritos_permitidos = current_user.distritos_permitidos
        
        # Iterar sobre todas las carpetas de distritos
        for distrito_dir in routes_dir.iterdir():
            if not distrito_dir.is_dir():
                continue
            
            codigo_distrito = distrito_dir.name
            
            # Filtrar por distritos permitidos (si no es admin)
            if current_user.cargo.lower() != "admin" and codigo_distrito not in distritos_permitidos:
                continue
            
            # Leer archivos JSON
            ruta_original_path = distrito_dir / "ruta_original.json"
            ruta_optimizada_path = distrito_dir / "ruta_optimizada.json"
            ruta_validada_path = distrito_dir / "ruta_validada.json"
            
            if not ruta_original_path.exists() or not ruta_optimizada_path.exists():
                continue
            
            with open(ruta_original_path, 'r', encoding='utf-8') as f:
                data_original = json.load(f)
            
            with open(ruta_optimizada_path, 'r', encoding='utf-8') as f:
                data_optimizada = json.load(f)
            
            # Cargar data validada si existe
            data_validada = None
            if ruta_validada_path.exists():
                with open(ruta_validada_path, 'r', encoding='utf-8') as f:
                    data_validada = json.load(f)
            
            nombre_distrito = data_original.get("distrito", codigo_distrito)
            
            # Procesar cada ruta original
            for ruta_orig in data_original.get("rutas_por_dia", []):
                vendedor = ruta_orig.get("vendedor", {})
                cod_vend = str(vendedor.get("codigo", ""))
                dia = ruta_orig.get("dia", "")
                
                # Buscar rutas de este vendedor específico
                if cod_vend == codigo_vendedor:
                    # Guardar info del vendedor (solo una vez)
                    if not vendedor_info:
                        vendedor_info = {
                            "codigo": cod_vend,
                            "nombre": vendedor.get("nombre", "Sin nombre")
                        }
                    
                    # Buscar la ruta correspondiente en validada (si existe) o en optimizada
                    ruta_comparacion = None
                    
                    # Primero buscar en validada
                    if data_validada:
                        for ruta_val in data_validada.get("rutas_por_dia", []):
                            if (str(ruta_val.get("vendedor", {}).get("codigo", "")) == cod_vend and 
                                ruta_val.get("dia", "") == dia):
                                ruta_comparacion = ruta_val
                                break
                    
                    # Si no está en validada, buscar en optimizada
                    if not ruta_comparacion:
                        for ruta_opt in data_optimizada.get("rutas_por_dia", []):
                            if (str(ruta_opt.get("vendedor", {}).get("codigo", "")) == cod_vend and 
                                ruta_opt.get("dia", "") == dia):
                                ruta_comparacion = ruta_opt
                                break
                    
                    # Si encontramos la ruta de comparación, agregarla al detalle
                    if ruta_comparacion:
                        km_sap = round(ruta_orig.get("distancia_km", 0), 2)
                        km_opt = round(ruta_comparacion.get("distancia_km", 0) if ruta_comparacion else 0, 2)
                        km_ruta = round(ruta_comparacion.get("distancia_km", 0), 2)
                        
                        # Calcular KM Holgura y KM Validado
                        km_holgura = round(km_ruta * (umbral_porcentaje / 100), 2)
                        km_validado = round(km_ruta + km_holgura, 2)
                        
                        # Diferencia calculada con km_validado - km_sap (negativo = ahorro)
                        diferencia_km = round(km_validado - km_sap, 2)
                        # Porcentaje: (km_validado - km_sap) / km_sap * 100 (negativo = ahorro)
                        porcentaje_dif = round((diferencia_km / km_sap * 100), 2) if km_sap > 0 else 0
                        
                        # Buscar la ruta optimizada para obtener km_optimizado
                        km_opt_real = 0
                        for ruta_opt in data_optimizada.get("rutas_por_dia", []):
                            if (str(ruta_opt.get("vendedor", {}).get("codigo", "")) == cod_vend and 
                                ruta_opt.get("dia", "") == dia):
                                km_opt_real = round(ruta_opt.get("distancia_km", 0), 2)
                                break
                        
                        rutas_detalle.append({
                            "distrito": nombre_distrito,
                            "codigo_distrito": codigo_distrito,
                            "dia": ruta_orig.get("dia", ""),
                            "fecha": ruta_orig.get("fecha", ""),
                            "km_sap": km_sap,
                            "km_optimizado": km_opt_real,
                            "km_ruta": km_ruta,
                            "km_holgura": km_holgura,
                            "km_validado": km_validado,
                            "diferencia_km": diferencia_km,
                            "porcentaje_diferencia": porcentaje_dif,
                            "total_puntos": ruta_orig.get("total_puntos", 0)
                        })
        
        # Verificar que se encontró el vendedor
        if not vendedor_info:
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró el vendedor con código '{codigo_vendedor}'"
            )
        
        # Cargar validaciones para incluir km_ruta_validada desde routes_json/[distrito]/ruta_validada.json
        estado_vendedor = "sin_validar"  # por defecto
        
        # Verificar si hay rutas guardadas para este vendedor
        guardado_encontrado = False
        for ruta in rutas_detalle:
            distrito_dir = routes_dir / ruta["codigo_distrito"]
            guardado_path = distrito_dir / "ruta_guardada.json"
            
            if guardado_path.exists():
                with open(guardado_path, 'r', encoding='utf-8') as f:
                    data_guardada = json.load(f)
                    
                    # Verificar si este vendedor tiene rutas guardadas
                    for ruta_guardada in data_guardada.get("rutas", []):
                        if str(ruta_guardada.get("vendedor", {}).get("codigo", "")) == codigo_vendedor:
                            guardado_encontrado = True
                            break
                
                if guardado_encontrado:
                    estado_vendedor = "guardado"
                    break
        
        # Recorrer todos los distritos para buscar validaciones del vendedor
        for ruta in rutas_detalle:
            distrito_dir = routes_dir / ruta["codigo_distrito"]
            validaciones_path = distrito_dir / "ruta_validada.json"
            
            if validaciones_path.exists():
                with open(validaciones_path, 'r', encoding='utf-8') as f:
                    validaciones_distrito = json.load(f)
                    
                    # Buscar la ruta validada que coincida con vendedor y día
                    for ruta_validada in validaciones_distrito.get("rutas_por_dia", []):
                        if (str(ruta_validada.get("vendedor", {}).get("codigo", "")) == codigo_vendedor and 
                            ruta_validada.get("dia", "") == ruta["dia"]):
                            ruta["km_ruta_validada"] = ruta_validada.get("distancia_km", 0)
                            break
            
            # Si no se encontró validación, asignar 0
            if "km_ruta_validada" not in ruta:
                ruta["km_ruta_validada"] = 0
        
        # Calcular totales sumando valores individuales ya redondeados
        total_km_sap = sum(r["km_sap"] for r in rutas_detalle)
        total_km_opt = sum(r["km_optimizado"] for r in rutas_detalle)
        total_km_ruta = sum(r["km_ruta"] for r in rutas_detalle)
        total_km_holgura = sum(r["km_holgura"] for r in rutas_detalle)
        total_km_validado = sum(r["km_validado"] for r in rutas_detalle)
        diferencia_total = sum(r["diferencia_km"] for r in rutas_detalle)  # Sumar diferencias individuales ya redondeadas
        # Porcentaje: (diferencia_total) / km_sap * 100 (negativo = ahorro)
        diferencia_porcentaje = round((diferencia_total / total_km_sap * 100), 2) if total_km_sap > 0 else 0
        
        return {
            "vendedor": vendedor_info,
            "rutas": rutas_detalle,
            "estado": estado_vendedor,  # Agregar estado
            "totales": {
                "total_km_sap": round(total_km_sap, 2),
                "total_km_optimizado": round(total_km_opt, 2),
                "total_km_ruta": round(total_km_ruta, 2),
                "total_km_holgura": round(total_km_holgura, 2),
                "total_km_validado": round(total_km_validado, 2),
                "ahorro_potencial": round(diferencia_total, 2),
                "ahorro_porcentaje": diferencia_porcentaje,
                "total_rutas": len(rutas_detalle)
            }
        }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al obtener detalles del vendedor: {str(e)}"
        )

@router.get("/ruta/{cod_vendedor}/{cod_distrito}/{dia}")
async def get_ruta_comparacion(
    cod_vendedor: str,
    cod_distrito: str,
    dia: str,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Obtener comparación de ruta SAP vs Optimizada para un vendedor, distrito y día específico
    
    Args:
        cod_vendedor: Código del vendedor
        cod_distrito: Código del distrito
        dia: Día de la semana (LU, MA, MI, JU, VI, SA, DO)
        current_user: Usuario actual
        
    Returns:
        Comparación con ruta_original, ruta_optimizada, ahorro_km y ahorro_porcentaje
    """
    try:
        routes_dir = Path(__file__).parent.parent.parent / "data" / "routes_json"
        distrito_dir = routes_dir / cod_distrito
        
        if not distrito_dir.exists():
            raise HTTPException(
                status_code=404,
                detail=f"No se encontraron rutas para el distrito '{cod_distrito}'"
            )
        
        # Leer JSON de ruta original
        json_original_path = distrito_dir / "ruta_original.json"
        if not json_original_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró ruta original para el distrito '{cod_distrito}'"
            )
        
        # Leer JSON de ruta optimizada o validada (si existe)
        json_validada_path = distrito_dir / "ruta_validada.json"
        json_optimizada_path = distrito_dir / "ruta_optimizada.json"
        
        if not json_optimizada_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró ruta optimizada para el distrito '{cod_distrito}'"
            )
        
        # Cargar JSONs
        with open(json_original_path, 'r', encoding='utf-8') as f:
            data_original = json.load(f)
        
        with open(json_optimizada_path, 'r', encoding='utf-8') as f:
            data_optimizada = json.load(f)
        
        # Cargar ruta validada si existe
        data_validada = None
        if json_validada_path.exists():
            with open(json_validada_path, 'r', encoding='utf-8') as f:
                data_validada = json.load(f)
        
        # Buscar la ruta específica del vendedor y día
        ruta_original = None
        ruta_optimizada = None
        es_ruta_editada = False
        
        # Buscar en ruta original
        for ruta in data_original.get("rutas_por_dia", []):
            if ruta["dia"] == dia and str(ruta["vendedor"]["codigo"]) == str(cod_vendedor):
                ruta_original = ruta
                break
        
        # Buscar en ruta optimizada original (sin editar)
        km_optimizada_original = 0
        for ruta in data_optimizada.get("rutas_por_dia", []):
            if ruta["dia"] == dia and str(ruta["vendedor"]["codigo"]) == str(cod_vendedor):
                km_optimizada_original = ruta["distancia_km"]
                break
        
        # Buscar en ruta validada primero (si existe)
        if data_validada:
            for ruta in data_validada.get("rutas_por_dia", []):
                if ruta["dia"] == dia and str(ruta["vendedor"]["codigo"]) == str(cod_vendedor):
                    ruta_optimizada = ruta
                    es_ruta_editada = True
                    break
        
        # Si no está en validada, buscar en optimizada
        if not ruta_optimizada:
            for ruta in data_optimizada.get("rutas_por_dia", []):
                if ruta["dia"] == dia and str(ruta["vendedor"]["codigo"]) == str(cod_vendedor):
                    ruta_optimizada = ruta
                    break
        
        if not ruta_original:
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró ruta original para vendedor {cod_vendedor}, distrito {cod_distrito}, día {dia}"
            )
        
        if not ruta_optimizada:
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró ruta optimizada para vendedor {cod_vendedor}, distrito {cod_distrito}, día {dia}"
            )
        
        # Añadir información del distrito a ambas rutas
        ruta_original["distrito"] = data_original.get("distrito", "")
        ruta_original["codigo_distrito"] = cod_distrito
        ruta_optimizada["distrito"] = data_optimizada.get("distrito", "")
        ruta_optimizada["codigo_distrito"] = cod_distrito
        
        # Verificar si existe ruta validada en routes_json/[distrito]/ruta_validada.json
        validaciones_path = distrito_dir / "ruta_validada.json"
        ruta_validada_km = 0
        km_para_comparar = ruta_optimizada["distancia_km"]  # Por defecto usar optimizada
        
        if validaciones_path.exists():
            with open(validaciones_path, 'r', encoding='utf-8') as f:
                validaciones = json.load(f)
                
                # Buscar en rutas_por_dia la ruta que coincida con vendedor y día
                for ruta_val in validaciones.get("rutas_por_dia", []):
                    if (str(ruta_val.get("vendedor", {}).get("codigo", "")) == cod_vendedor and 
                        ruta_val.get("dia", "") == dia):
                        ruta_validada_km = ruta_val.get("distancia_km", 0)
                        km_para_comparar = ruta_validada_km  # Usar validada si existe
                        break
        
        # Calcular ahorro usando km_ruta si existe, sino km_optimizado
        km_original = ruta_original["distancia_km"]
        ahorro_km = km_para_comparar - km_original  # Positivo si ruta > SAP, negativo si ruta < SAP
        ahorro_porcentaje = (ahorro_km / km_original * 100) if km_original > 0 else 0
        
        return {
            "ruta_original": ruta_original,
            "ruta_optimizada": ruta_optimizada,
            "ahorro_km": round(ahorro_km, 2),
            "ahorro_porcentaje": round(ahorro_porcentaje, 2),
            "es_ruta_editada": es_ruta_editada,
            "ruta_validada_km": ruta_validada_km,
            "km_optimizada_original": round(km_optimizada_original, 2)
        }
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al obtener comparación de rutas: {str(e)}"
        )


@router.post("/recalcular-temporal")
async def recalcular_ruta_temporal(
    datos_ruta: dict,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Recalcula una ruta temporalmente sin guardar cambios.
    Usado para simulación de ediciones del usuario.
    
    Args:
        datos_ruta: Diccionario con coordenadas y puntos editados
        current_user: Usuario actual
        
    Returns:
        Ruta recalculada con OSRM
    """
    try:
        coordenadas = datos_ruta.get('coordenadas', [])
        puntos = datos_ruta.get('puntos', [])
        vendedor = datos_ruta.get('vendedor', {})
        distrito = datos_ruta.get('distrito', '')
        codigo_distrito = datos_ruta.get('codigo_distrito', '')
        dia = datos_ruta.get('dia', '')
        fecha = datos_ruta.get('fecha', '')
        mantener_orden = datos_ruta.get('mantener_orden', False)  # No reoptimizar si es True
        
        if not coordenadas or len(coordenadas) < 2:
            raise HTTPException(
                status_code=400,
                detail="Se requieren al menos 2 coordenadas"
            )
        
        # Llamar a OSRM usando el método correcto: Table + NN + Route
        import httpx
        
        # Separar punto inicial y final (oficina) de los puntos intermedios
        coord_inicio = coordenadas[0]  # Oficina inicio [lon, lat]
        coord_fin = coordenadas[-1]   # Oficina fin [lon, lat]
        coords_clientes = coordenadas[1:-1]  # Clientes intermedios
        
        if len(coords_clientes) == 0:
            # Sin clientes, solo inicio-fin
            raise HTTPException(
                status_code=400,
                detail="No hay clientes intermedios para optimizar"
            )
        
        async with httpx.AsyncClient(timeout=60.0) as client:
            # Decidir si aplicar NN o mantener orden
            if mantener_orden:
                # Mantener el orden original (ya está optimizado)
                route_indices = list(range(len(coordenadas)))
            else:
                # PASO 1: Obtener matriz de distancias con OSRM Table
                coords_str = ';'.join([f"{lon},{lat}" for lon, lat in coordenadas])
                table_url = f"http://osrm:5000/table/v1/driving/{coords_str}"
                
                table_response = await client.get(table_url)
                
                if table_response.status_code != 200:
                    raise HTTPException(
                        status_code=500,
                        detail="Error al obtener matriz de distancias de OSRM"
                    )
                
                table_data = table_response.json()
                
                if table_data.get('code') != 'Ok':
                    raise HTTPException(
                        status_code=500,
                        detail=f"OSRM Table error: {table_data.get('message', 'Unknown error')}"
                    )
                
                # Matriz de duraciones (en segundos) - usamos esto como métrica
                duration_matrix = table_data['durations']
                
                # PASO 2: Algoritmo Nearest Neighbor
                # Empezamos siempre en índice 0 (oficina inicio)
                n_points = len(coordenadas)
                current_idx = 0
                visited = [False] * n_points
                visited[0] = True  # Oficina inicio
                visited[-1] = True  # Oficina fin no se visita en el loop
                
                route_indices = [0]  # Empezamos en oficina
                
                # Visitar todos los clientes (índices 1 hasta n-2)
                for _ in range(len(coords_clientes)):
                    min_duration = float('inf')
                    next_idx = -1
                    
                    # Buscar el cliente más cercano no visitado
                    for j in range(1, n_points - 1):  # Solo clientes (no oficinas)
                        if not visited[j]:
                            duration = duration_matrix[current_idx][j]
                            if duration < min_duration:
                                min_duration = duration
                                next_idx = j
                    
                    if next_idx == -1:
                        break
                    
                    visited[next_idx] = True
                    route_indices.append(next_idx)
                    current_idx = next_idx
                
                # Agregar oficina final
                route_indices.append(n_points - 1)
            
            # PASO 3: Obtener geometría con OSRM Route usando el orden optimizado
            coords_optimizadas = [coordenadas[i] for i in route_indices]
            coords_str_opt = ';'.join([f"{lon},{lat}" for lon, lat in coords_optimizadas])
            route_url = f"http://osrm:5000/route/v1/driving/{coords_str_opt}"
            route_params = {
                'overview': 'full',
                'geometries': 'geojson',
                'steps': 'true'
            }
            
            route_response = await client.get(route_url, params=route_params)
            
            if route_response.status_code != 200:
                raise HTTPException(
                    status_code=500,
                    detail="Error al calcular geometría de ruta con OSRM"
                )
            
            route_data = route_response.json()
        
        if route_data.get('code') != 'Ok':
            raise HTTPException(
                status_code=500,
                detail=f"OSRM Route error: {route_data.get('message', 'Unknown error')}"
            )
        
        route = route_data['routes'][0]
        
        # Reordenar puntos según el orden optimizado por NN
        puntos_reordenados = [puntos[i] for i in route_indices]
        
        # Actualizar secuencias optimizadas
        for idx, punto in enumerate(puntos_reordenados):
            if idx == 0:
                punto['secuencia_optimizada'] = 0
            elif idx == len(puntos_reordenados) - 1:
                punto['secuencia_optimizada'] = 1000
            else:
                punto['secuencia_optimizada'] = idx * 5
        
        # Construir respuesta
        ruta_recalculada = {
            'distrito': distrito,
            'codigo_distrito': codigo_distrito,
            'dia': dia,
            'fecha': fecha,
            'vendedor': vendedor,
            'distancia_km': route['distance'] / 1000,
            'duracion_minutos': route['duration'] / 60,
            'total_puntos': len(puntos_reordenados),
            'geometria': route['geometry'],
            'puntos': puntos_reordenados,
            'legs': route.get('legs', [])
        }
        
        return JSONResponse(content=ruta_recalculada)
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al recalcular ruta: {str(e)}"
        )


@router.post("/guardar-ruta-editada")
async def guardar_ruta_editada(
    datos_ruta: dict,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Guarda una ruta editada en ruta_validada.json.
    NO modifica el Maestro APP ni los archivos originales/optimizados.
    
    Args:
        datos_ruta: Diccionario con los datos de la ruta editada
        current_user: Usuario actual
        
    Returns:
        Mensaje de confirmación
    """
    try:
        # Verificar que la aplicación esté activa
        from app.routes.config import load_config
        config = load_config()
        if not config.get("app_activa", True):
            raise HTTPException(
                status_code=403,
                detail="El sistema está desactivado. No se pueden guardar rutas en este momento."
            )
        
        # Obtener datos de la ruta
        vendedor = datos_ruta.get('vendedor', {})
        distrito = datos_ruta.get('distrito', '')
        codigo_distrito = datos_ruta.get('codigo_distrito', '')
        dia = datos_ruta.get('dia', '')
        fecha = datos_ruta.get('fecha', '')
        puntos = datos_ruta.get('puntos', [])
        distancia_km = datos_ruta.get('distancia_km', 0)
        duracion_minutos = datos_ruta.get('duracion_minutos', 0)
        geometria = datos_ruta.get('geometria', {})
        legs = datos_ruta.get('legs', [])
        
        # Verificar permisos: admin puede editar cualquier distrito, 
        # usuarios normales solo sus distritos asignados
        if current_user.cargo.lower() != "admin":
            if codigo_distrito not in current_user.distritos_permitidos:
                raise HTTPException(
                    status_code=403,
                    detail=f"No tienes permisos para modificar rutas del distrito {distrito}"
                )
        
        if not puntos or len(puntos) < 2:
            raise HTTPException(
                status_code=400,
                detail="Se requieren al menos 2 puntos"
            )
        
        # Ruta al directorio de rutas JSON
        ruta_dir = Path(f"data/routes_json/{codigo_distrito}")
        
        if not ruta_dir.exists():
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró el directorio de rutas para el distrito {codigo_distrito}"
            )
        
        # Leer el archivo de ruta optimizada para obtener estructura base y km original
        ruta_optimizada_path = ruta_dir / "ruta_optimizada.json"
        
        if not ruta_optimizada_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró el archivo de ruta optimizada para el distrito {codigo_distrito}"
            )
        
        # Cargar el archivo de ruta optimizada existente
        with open(ruta_optimizada_path, 'r', encoding='utf-8') as f:
            data_optimizada = json.load(f)
        
        # Buscar la ruta específica del vendedor y día para obtener km_original
        ruta_encontrada = False
        km_original = 0
        jefe_venta_info = ""
        for ruta in data_optimizada.get("rutas_por_dia", []):
            if (ruta.get("vendedor", {}).get("codigo") == vendedor['codigo'] and 
                ruta.get("dia") == dia):
                km_original = ruta.get("distancia_km", 0)
                jefe_venta_info = ruta.get("jefe_venta", "")
                ruta_encontrada = True
                break
        
        if not ruta_encontrada:
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró la ruta del vendedor {vendedor['codigo']} para el día {dia}"
            )
        
        # Calcular distancias para cada punto usando legs de OSRM
        distancia_acumulada = 0
        puntos_con_distancias = []
        
        for idx, punto in enumerate(puntos):
            punto_con_distancia = punto.copy()
            
            # Calcular distancia desde el punto anterior usando legs
            distancia_desde_anterior = 0
            if idx > 0 and legs and len(legs) > idx - 1:
                # Usar distancia del leg de OSRM (distancia real por carretera)
                leg = legs[idx - 1]
                distancia_desde_anterior = leg.get('distance', 0) / 1000  # Convertir de metros a km
                distancia_acumulada += distancia_desde_anterior
            
            punto_con_distancia['distancia_desde_anterior'] = round(distancia_desde_anterior, 2)
            punto_con_distancia['distancia_acumulada'] = round(distancia_acumulada, 2)
            puntos_con_distancias.append(punto_con_distancia)
        
        # Crear la nueva ruta validada
        nueva_ruta_validada = {
            "dia": dia,
            "fecha": fecha,
            "vendedor": {
                "codigo": vendedor['codigo'],
                "nombre": vendedor['nombre']
            },
            "jefe_venta": jefe_venta_info,
            "total_puntos": len(puntos_con_distancias),
            "distancia_km": distancia_km,
            "duracion_minutos": duracion_minutos,
            "geometria": geometria,
            "legs": legs,
            "puntos": puntos_con_distancias
        }
        
        # Verificar si ya existe un archivo ruta_validada.json
        ruta_validada_path = ruta_dir / "ruta_validada.json"
        
        if ruta_validada_path.exists():
            # Cargar el archivo existente y actualizar la ruta
            with open(ruta_validada_path, 'r', encoding='utf-8') as f:
                distrito_json_validada = json.load(f)
            
            # Buscar y actualizar la ruta específica
            ruta_actualizada = False
            for idx, ruta_val in enumerate(distrito_json_validada.get("rutas_por_dia", [])):
                if (ruta_val.get("vendedor", {}).get("codigo") == vendedor['codigo'] and 
                    ruta_val.get("dia") == dia):
                    distrito_json_validada["rutas_por_dia"][idx] = nueva_ruta_validada
                    ruta_actualizada = True
                    break
            
            # Si no se encontró, agregarla
            if not ruta_actualizada:
                distrito_json_validada["rutas_por_dia"].append(nueva_ruta_validada)
        else:
            # Crear nuevo archivo basado en la estructura de ruta_optimizada
            distrito_json_validada = {
                "distrito": data_optimizada.get("distrito", distrito),
                "codigo_distrito": codigo_distrito,
                "tipo_ruta": "validada",
                "rutas_por_dia": [nueva_ruta_validada]
            }
        
        # Guardar el archivo ruta_validada.json
        with open(ruta_validada_path, 'w', encoding='utf-8') as f:
            json.dump(distrito_json_validada, f, ensure_ascii=False, indent=2)
        
        return JSONResponse(content={
            "message": "Ruta validada guardada exitosamente. Los cambios se verán reflejados en el sistema.",
            "vendedor": vendedor['nombre'],
            "distrito": distrito,
            "dia": dia,
            "total_puntos": len(puntos),
            "km_nuevo": round(distancia_km, 2),
            "km_original": round(km_original, 2)
        })
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error guardando ruta editada: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al guardar ruta: {str(e)}"
        )


@router.delete("/restablecer-ruta-editada/{cod_distrito}/{cod_vendedor}/{dia}")
async def restablecer_ruta_editada(
    cod_distrito: str,
    cod_vendedor: str,
    dia: str,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Restablecer una ruta validada a la versión optimizada original
    Esto hace que se vuelva a mostrar la ruta optimizada original
    
    Args:
        cod_distrito: Código del distrito
        cod_vendedor: Código del vendedor
        dia: Día de la semana (LU, MA, MI, JU, VI, SA, DO)
        current_user: Usuario actual
        
    Returns:
        Mensaje de confirmación
    """
    try:
        # Verificar si la aplicación permite guardar/restablecer rutas
        import os
        config_file = "data/config.json"
        app_activa = True  # Por defecto activa
        
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    app_activa = config.get("app_activa", True)
            except Exception:
                pass  # Si hay error, asumir que está activa
        
        # Si la app está desactivada, no permitir restablecer rutas
        if not app_activa:
            raise HTTPException(
                status_code=403,
                detail="El guardado de rutas está temporalmente desactivado por el administrador"
            )
        
        # Verificar permisos: admin puede restablecer cualquier distrito, 
        # usuarios normales solo sus distritos asignados
        if current_user.cargo.lower() != "admin":
            if cod_distrito not in current_user.distritos_permitidos:
                raise HTTPException(
                    status_code=403,
                    detail=f"No tienes permisos para restablecer rutas del distrito {cod_distrito}"
                )
        
        # Ruta al directorio de rutas JSON
        ruta_dir = Path(f"data/routes_json/{cod_distrito}")
        
        if not ruta_dir.exists():
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró el directorio de rutas para el distrito {cod_distrito}"
            )
        
        # Verificar si existe el archivo ruta_validada.json
        ruta_validada_path = ruta_dir / "ruta_validada.json"
        
        if not ruta_validada_path.exists():
            return JSONResponse(content={
                "message": "No hay ruta validada para restablecer",
                "distrito": cod_distrito,
                "vendedor": cod_vendedor,
                "dia": dia
            })
        
        # Cargar el archivo para verificar si contiene la ruta específica
        with open(ruta_validada_path, 'r', encoding='utf-8') as f:
            distrito_json_validada = json.load(f)
        
        # Buscar la ruta específica
        rutas_validadas = distrito_json_validada.get("rutas_por_dia", [])
        ruta_encontrada = False
        nuevas_rutas = []
        
        for ruta in rutas_validadas:
            if (str(ruta.get("vendedor", {}).get("codigo")) == str(cod_vendedor) and 
                ruta.get("dia") == dia):
                ruta_encontrada = True
                # No agregar esta ruta a la lista nueva (eliminarla)
            else:
                nuevas_rutas.append(ruta)
        
        if not ruta_encontrada:
            return JSONResponse(content={
                "message": "No se encontró ruta validada para este vendedor y día",
                "distrito": cod_distrito,
                "vendedor": cod_vendedor,
                "dia": dia
            })
        
        # Leer la ruta optimizada original para reemplazar
        ruta_optimizada_path = ruta_dir / "ruta_optimizada.json"
        with open(ruta_optimizada_path, 'r', encoding='utf-8') as f:
            data_optimizada = json.load(f)
        
        # Buscar la ruta optimizada correspondiente
        for ruta_opt in data_optimizada.get("rutas_por_dia", []):
            if (str(ruta_opt.get("vendedor", {}).get("codigo")) == str(cod_vendedor) and 
                ruta_opt.get("dia") == dia):
                # Agregar la ruta optimizada de vuelta
                nuevas_rutas.append(ruta_opt)
                break
        
        # Actualizar el archivo con las rutas
        distrito_json_validada["rutas_por_dia"] = nuevas_rutas
        with open(ruta_validada_path, 'w', encoding='utf-8') as f:
            json.dump(distrito_json_validada, f, ensure_ascii=False, indent=2)
        mensaje = "Ruta restablecida exitosamente a versión optimizada original"
        
        return JSONResponse(content={
            "message": mensaje,
            "distrito": cod_distrito,
            "vendedor": cod_vendedor,
            "dia": dia
        })
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error restableciendo ruta: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al restablecer ruta: {str(e)}"
        )


@router.post("/validar-ruta")
async def validar_ruta(
    datos_validacion: dict,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Registra la validación de una ruta en routes_json/[distrito]/ruta_validada.json
    con toda la información necesaria para exportar a Excel.
    
    Args:
        datos_validacion: Diccionario con los datos de la validación
        current_user: Usuario actual
        
    Returns:
        Mensaje de confirmación
    """
    try:
        from geopy.distance import geodesic
        
        # Extraer datos de validación
        vendedor_codigo = str(datos_validacion.get('vendedor_codigo'))
        vendedor_nombre = datos_validacion.get('vendedor_nombre', '')
        codigo_distrito = datos_validacion.get('codigo_distrito', '')
        distrito = datos_validacion.get('distrito', '')
        dia = datos_validacion.get('dia', '')
        km_ruta = datos_validacion.get('km_ruta', 0)
        km_sap = datos_validacion.get('km_sap', 0)
        tipo_ruta = datos_validacion.get('tipo_ruta', 'optimizada')
        
        # Verificar permisos
        if current_user.cargo.lower() != "admin":
            if codigo_distrito not in current_user.distritos_permitidos:
                raise HTTPException(
                    status_code=403,
                    detail=f"No tienes permisos para validar rutas del distrito {distrito}"
                )
        
        # Directorio del distrito
        distrito_dir = Path(f"data/routes_json/{codigo_distrito}")
        if not distrito_dir.exists():
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró el directorio para el distrito {codigo_distrito}"
            )
        
        # Cargar o crear JSON de validaciones del distrito
        validaciones_path = distrito_dir / "ruta_validada.json"
        if validaciones_path.exists():
            with open(validaciones_path, 'r', encoding='utf-8') as f:
                validaciones = json.load(f)
        else:
            validaciones = {
                "distrito": distrito,
                "codigo_distrito": codigo_distrito,
                "vendedores": {}
            }
        
        # Cargar ruta validada u optimizada para obtener puntos completos
        ruta_validada_path = distrito_dir / "ruta_validada.json"
        
        if ruta_validada_path.exists():
            with open(ruta_validada_path, 'r', encoding='utf-8') as f:
                ruta_data = json.load(f)
        else:
            raise HTTPException(
                status_code=404,
                detail=f"No se encontraron datos de ruta para el distrito {codigo_distrito}"
            )
        
        # Buscar la ruta del vendedor y día
        ruta_encontrada = None
        for ruta in ruta_data.get("rutas_por_dia", []):
            if str(ruta.get("vendedor", {}).get("codigo")) == str(vendedor_codigo) and ruta.get("dia") == dia:
                ruta_encontrada = ruta
                break
        
        if not ruta_encontrada:
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró la ruta del vendedor {vendedor_codigo} para el día {dia}"
            )
        
        # Calcular distancias y crear detalle de clientes
        puntos = ruta_encontrada.get("puntos", [])
        clientes_detalle = []
        distancia_acumulada = 0
        
        for i, punto in enumerate(puntos):
            if i > 0:
                coord_anterior = (puntos[i-1]['coordenadas']['lat'], puntos[i-1]['coordenadas']['lon'])
                coord_actual = (punto['coordenadas']['lat'], punto['coordenadas']['lon'])
                distancia_desde_anterior = geodesic(coord_anterior, coord_actual).kilometers
                distancia_acumulada += distancia_desde_anterior
            else:
                distancia_desde_anterior = 0
            
            cliente = {
                'cod_cliente': punto.get('cod_cliente', ''),
                'razon_social': punto.get('razon_social', ''),
                'tipo_negocio': punto.get('tipo_negocio', ''),
                'relevancia': '',
                'secuencia_visita': punto.get('secuencia_optimizada', punto.get('secuencia_original', i + 1)),
                'ritmo_visita': '',
                'distancia_desde_anterior': round(distancia_desde_anterior, 2),
                'distancia_acumulada': round(distancia_acumulada, 2),
                'latitud': punto['coordenadas']['lat'],
                'longitud': punto['coordenadas']['lon']
            }
            clientes_detalle.append(cliente)
        
        # Hora actual en Chile
        hora_chile = datetime.now(ZoneInfo('America/Santiago'))
        
        # Inicializar estructura del vendedor si no existe
        if vendedor_codigo not in validaciones["vendedores"]:
            validaciones["vendedores"][vendedor_codigo] = {
                "estado": "en_progreso",
                "vendedor_codigo": vendedor_codigo,
                "vendedor_nombre": vendedor_nombre,
                "fecha_validacion": hora_chile.strftime('%Y-%m-%d %H:%M:%S'),
                "validado_por": current_user.nombre,
                "rutas": []
            }
        
        # Eliminar ruta anterior del mismo día si existe
        validaciones["vendedores"][vendedor_codigo]["rutas"] = [
            r for r in validaciones["vendedores"][vendedor_codigo]["rutas"]
            if r["dia"] != dia
        ]
        
        # Agregar nueva ruta validada
        ruta_validada = {
            "dia": dia,
            "km_sap": round(km_sap, 2),
            "km_ruta": round(km_ruta, 2),
            "tipo_ruta": tipo_ruta,
            "clientes": clientes_detalle
        }
        validaciones["vendedores"][vendedor_codigo]["rutas"].append(ruta_validada)
        
        # Actualizar fecha de última validación
        validaciones["vendedores"][vendedor_codigo]["fecha_validacion"] = hora_chile.strftime('%Y-%m-%d %H:%M:%S')
        validaciones["vendedores"][vendedor_codigo]["validado_por"] = current_user.nombre
        
        # Guardar JSON actualizado
        with open(validaciones_path, 'w', encoding='utf-8') as f:
            json.dump(validaciones, f, ensure_ascii=False, indent=2)
        
        return JSONResponse(content={
            "message": "Ruta validada exitosamente",
            "vendedor": vendedor_nombre,
            "distrito": distrito,
            "dia": dia,
            "km_ruta": round(km_ruta, 2),
            "tipo_ruta": tipo_ruta,
            "estado": validaciones["vendedores"][vendedor_codigo]["estado"]
        })
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error validando ruta: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al validar ruta: {str(e)}"
        )

@router.post("/guardar-rutas-vendedor")
async def guardar_rutas_vendedor(
    datos_vendedor: dict,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Guarda las rutas de un vendedor específico en el archivo ruta_guardada.json.
    Similar a guardar-ruta-vendedor pero pensado para la vista individual del vendedor.
    
    Args:
        datos_vendedor: Diccionario con código del vendedor, nombre y distrito
        current_user: Usuario actual
        
    Returns:
        Confirmación con detalles de las rutas guardadas
    """
    try:
        # Verificar si la aplicación permite guardar rutas
        import os
        config_file = "data/config.json"
        app_activa = True  # Por defecto activa
        
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    app_activa = config.get("app_activa", True)
            except Exception:
                pass  # Si hay error, asumir que está activa
        
        # Si la app está desactivada, no permitir guardar rutas
        if not app_activa:
            raise HTTPException(
                status_code=403,
                detail="El guardado de rutas está temporalmente desactivado por el administrador"
            )
        
        # Verificar permisos
        if current_user.cargo.lower() == "admin":
            raise HTTPException(
                status_code=403,
                detail="Los administradores no pueden guardar rutas. Esta función es para jefes de venta."
            )
        
        # Extraer datos
        vendedor_codigo = str(datos_vendedor.get('vendedor_codigo'))
        rutas = datos_vendedor.get('rutas', [])
        
        if not vendedor_codigo or not rutas:
            raise HTTPException(
                status_code=400,
                detail="Debe proporcionar código del vendedor y sus rutas"
            )
        
        # Obtener el distrito de la primera ruta
        if not rutas:
            raise HTTPException(
                status_code=400,
                detail="No hay rutas para guardar"
            )
        
        codigo_distrito = rutas[0].get('codigo_distrito')
        if not codigo_distrito:
            raise HTTPException(
                status_code=400,
                detail="No se pudo determinar el distrito"
            )
        
        # Verificar que el usuario tenga permiso para este distrito
        if codigo_distrito not in current_user.distritos_permitidos:
            raise HTTPException(
                status_code=403,
                detail=f"No tienes permisos para guardar rutas del distrito {codigo_distrito}"
            )
        
        # Ruta al directorio del distrito
        distrito_dir = Path(f"data/routes_json/{codigo_distrito}")
        
        if not distrito_dir.exists():
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró el directorio del distrito {codigo_distrito}"
            )
        
        # Leer ruta_validada.json o ruta_optimizada.json
        ruta_validada_path = distrito_dir / "ruta_validada.json"
        ruta_optimizada_path = distrito_dir / "ruta_optimizada.json"
        
        data_rutas = None
        if ruta_validada_path.exists():
            with open(ruta_validada_path, 'r', encoding='utf-8') as f:
                data_rutas = json.load(f)
        elif ruta_optimizada_path.exists():
            with open(ruta_optimizada_path, 'r', encoding='utf-8') as f:
                data_rutas = json.load(f)
        else:
            raise HTTPException(
                status_code=404,
                detail=f"No se encontraron rutas para guardar en el distrito {codigo_distrito}"
            )
        
        # Cargar umbral de porcentaje desde config.json
        config_file_path = Path(__file__).parent.parent.parent / "data" / "config.json"
        umbral_porcentaje = 10.0  # Valor por defecto
        if config_file_path.exists():
            try:
                with open(config_file_path, 'r', encoding='utf-8') as f:
                    config_data = json.load(f)
                    umbral_porcentaje = config_data.get("umbral_porcentaje", 10.0)
            except Exception as e:
                logger.warning(f"Error al cargar umbral desde config.json: {e}")
        
        # Filtrar solo las rutas del vendedor especificado
        rutas_vendedor = []
        for ruta in data_rutas.get("rutas_por_dia", []):
            vendedor_info = ruta.get("vendedor", {})
            cod_vend_ruta = vendedor_info.get("codigo")
            
            # Solo incluir rutas de este vendedor
            if str(cod_vend_ruta) == str(vendedor_codigo):
                ruta_limpia = ruta.copy()
                # Eliminar geometría y legs si existen
                if "geometria" in ruta_limpia:
                    del ruta_limpia["geometria"]
                if "legs" in ruta_limpia:
                    del ruta_limpia["legs"]
                # Eliminar tiempo de ruta si existe
                if "tiempo_ruta" in ruta_limpia:
                    del ruta_limpia["tiempo_ruta"]
                if "duracion_minutos" in ruta_limpia:
                    del ruta_limpia["duracion_minutos"]
                
                # Calcular y agregar km_holgura y km_validado
                km_ruta = ruta_limpia.get("distancia_km", 0)
                km_holgura = round(km_ruta * (umbral_porcentaje / 100), 2)
                km_validado = round(km_ruta + km_holgura, 2)
                
                ruta_limpia["km_holgura"] = km_holgura
                ruta_limpia["km_validado"] = km_validado
                
                rutas_vendedor.append(ruta_limpia)
        
        if not rutas_vendedor:
            raise HTTPException(
                status_code=404,
                detail=f"No se encontraron rutas para el vendedor {vendedor_codigo} en el distrito {codigo_distrito}"
            )
        
        # Verificar si ya existe un guardado previo y cargar rutas existentes
        guardado_path = distrito_dir / "ruta_guardada.json"
        rutas_existentes = []
        
        if guardado_path.exists():
            with open(guardado_path, 'r', encoding='utf-8') as f:
                data_guardada = json.load(f)
                # Filtrar rutas que NO sean del vendedor actual (mantener las demás)
                for ruta in data_guardada.get("rutas", []):
                    vendedor_info = ruta.get("vendedor", {})
                    if str(vendedor_info.get("codigo")) != str(vendedor_codigo):
                        rutas_existentes.append(ruta)
        
        # Combinar rutas: existentes de otros vendedores + nuevas del vendedor actual
        todas_rutas = rutas_existentes + rutas_vendedor
        
        # Hora actual en Chile
        hora_chile = datetime.now(ZoneInfo('America/Santiago'))
        
        # Crear estructura del consolidado guardado
        consolidado = {
            "distrito": data_rutas.get("distrito", ""),
            "codigo_distrito": codigo_distrito,
            "tipo_ruta": "guardada",
            "total_dias": len(rutas_vendedor),
            "usuario": current_user.usuario,
            "usuario_email": current_user.email or "",
            "fecha_guardado": hora_chile.isoformat(),
            "rutas": todas_rutas
        }
        
        # Guardar consolidado
        with open(guardado_path, 'w', encoding='utf-8') as f:
            json.dump(consolidado, f, ensure_ascii=False, indent=2)
        
        return JSONResponse(content={
            "message": "Rutas guardadas exitosamente",
            "vendedor": rutas_vendedor[0].get("vendedor", {}).get("nombre", ""),
            "total_rutas": len(rutas_vendedor),
            "distrito": codigo_distrito
        })
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error guardando rutas del vendedor: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al guardar rutas: {str(e)}"
        )

@router.get("/info-guardado/{distrito}/{cod_vendedor}")
async def obtener_info_guardado(
    distrito: str,
    cod_vendedor: str,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Obtiene información del guardado previo de un vendedor en un distrito.
    
    Args:
        distrito: Código del distrito
        cod_vendedor: Código del vendedor
        current_user: Usuario actual
        
    Returns:
        Información del guardado previo (fecha, km_validado_total) o None si no existe
    """
    try:
        distrito_dir = Path(f"data/routes_json/{distrito}")
        guardado_path = distrito_dir / "ruta_guardada.json"
        
        if not guardado_path.exists():
            return {
                "tiene_guardado": False
            }
        
        with open(guardado_path, 'r', encoding='utf-8') as f:
            data_guardada = json.load(f)
        
        # Buscar rutas del vendedor en el guardado
        km_validado_total = 0
        rutas_vendedor = []
        rutas_detalle = []
        fecha_guardado = data_guardada.get("fecha_guardado")
        
        for ruta in data_guardada.get("rutas", []):
            vendedor_info = ruta.get("vendedor", {})
            if str(vendedor_info.get("codigo")) == str(cod_vendedor):
                rutas_vendedor.append(ruta)
                km_validado_total += ruta.get("km_validado", 0)
                # Agregar detalle de cada ruta para comparación
                rutas_detalle.append({
                    "dia": ruta.get("dia"),
                    "km_ruta": ruta.get("km_ruta", 0),
                    "km_validado": ruta.get("km_validado", 0)
                })
        
        if not rutas_vendedor:
            return {
                "tiene_guardado": False
            }
        
        return {
            "tiene_guardado": True,
            "fecha_guardado": fecha_guardado,
            "km_validado_total": round(km_validado_total, 2),
            "total_rutas": len(rutas_vendedor),
            "rutas_detalle": rutas_detalle
        }
    
    except Exception as e:
        logger.error(f"Error obteniendo info de guardado: {str(e)}")
        return {
            "tiene_guardado": False
        }

@router.get("/exportar-pdf-vendedor/{distrito}/{cod_vendedor}")
async def exportar_pdf_vendedor(
    distrito: str,
    cod_vendedor: str,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Genera un PDF con las rutas guardadas de un vendedor.
    Envía el PDF por correo al usuario actual con copia al administrador.
    
    Args:
        distrito: Código del distrito
        cod_vendedor: Código del vendedor
        current_user: Usuario actual
        
    Returns:
        PDF con información del vendedor y tabla de rutas
    """
    # Generar el PDF
    response = pdf_service.generar_pdf_vendedor(distrito, cod_vendedor, current_user.nombre)
    
    # Enviar por correo si el usuario tiene email configurado
    if current_user.email and current_user.email.strip():
        try:
            from app.services.email_service import email_service
            from app.config import settings
            
            # Asunto: CodigoVend - Distrito (código del distrito, no nombre)
            asunto = f"{cod_vendedor} - {distrito}"
            
            # Cuerpo del correo vacío (solo archivo adjunto)
            cuerpo_html = ""
            
            # Enviar correo con el PDF adjunto
            pdf_content = response.body
            pdf_filename = f"{cod_vendedor}.pdf"
            cc_email = settings.EMAIL_ADMIN if settings.EMAIL_ADMIN else None
            
            email_service.send_pdf_email(
                destinatario=current_user.email,
                asunto=asunto,
                cuerpo_html=cuerpo_html,
                pdf_content=pdf_content,
                pdf_filename=pdf_filename,
                cc_email=cc_email
            )
        except Exception as e:
            # Si falla el envío de correo, simplemente registrar el error pero continuar
            print(f"⚠️ Error al enviar correo: {str(e)}")
    
    return response


@router.get("/resumen-rutas/{distrito}")
async def get_resumen_rutas(
    distrito: str,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Obtener resumen de rutas validadas antes de guardar
    
    Args:
        distrito: Código del distrito
        current_user: Usuario actual
        
    Returns:
        Resumen con KM SAP, Optimizado, Validado, Ahorro, y advertencias
    """
    try:
        # Verificar permisos
        if current_user.cargo.lower() != "admin" and distrito not in current_user.distritos_permitidos:
            raise HTTPException(
                status_code=403,
                detail=f"No tienes permisos para este distrito"
            )
        
        distrito_dir = Path(f"data/routes_json/{distrito}")
        
        if not distrito_dir.exists():
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró el directorio del distrito {distrito}"
            )
        
        # Leer ruta_validada.json
        ruta_validada_path = distrito_dir / "ruta_validada.json"
        
        if not ruta_validada_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"No hay rutas validadas en el distrito {distrito}"
            )
        
        with open(ruta_validada_path, 'r', encoding='utf-8') as f:
            data_validada = json.load(f)
        
        # Leer ruta_original.json para obtener km_sap
        ruta_original_path = distrito_dir / "ruta_original.json"
        data_original = None
        
        if ruta_original_path.exists():
            with open(ruta_original_path, 'r', encoding='utf-8') as f:
                data_original = json.load(f)
        
        # Verificar si ya existe guardado previo
        guardado_path = distrito_dir / "ruta_guardada.json"
        guardado_previo = None
        
        if guardado_path.exists():
            with open(guardado_path, 'r', encoding='utf-8') as f:
                guardado_previo = json.load(f)
        
        # Calcular totales desde rutas_por_dia
        km_sap = 0
        km_optimizado = 0
        km_ruta = 0
        
        # Crear un diccionario para buscar km_sap por vendedor y día
        km_sap_dict = {}
        if data_original:
            for ruta in data_original.get("rutas_por_dia", []):
                vendedor_info = ruta.get("vendedor", {})
                vendedor_codigo = vendedor_info.get("codigo", "") if isinstance(vendedor_info, dict) else ""
                dia = ruta.get("dia", "")
                key = f"{vendedor_codigo}_{dia}"
                km_sap_dict[key] = ruta.get("distancia_km", 0)
        
        for ruta in data_validada.get("rutas_por_dia", []):
            # distancia_km es la distancia de la ruta validada
            km_ruta += ruta.get("distancia_km", 0)
            km_optimizado += ruta.get("distancia_km", 0)
            
            # Buscar km_sap correspondiente en ruta_original
            vendedor_info = ruta.get("vendedor", {})
            vendedor_codigo = vendedor_info.get("codigo", "") if isinstance(vendedor_info, dict) else ""
            dia = ruta.get("dia", "")
            key = f"{vendedor_codigo}_{dia}"
            km_sap += km_sap_dict.get(key, ruta.get("distancia_km", 0))
        
        # Ahorro = km_ruta - km_sap (negativo = ahorro)
        ahorro_potencial = km_ruta - km_sap
        porcentaje_ahorro = ((km_ruta - km_sap) / km_sap * 100) if km_sap > 0 else 0
        
        # Obtener umbral configurado
        config_path = Path("data/config.json")
        umbral_porcentaje = -25  # default
        
        if config_path.exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
                umbral_porcentaje = config_data.get("umbral_porcentaje", -25)
        
        # Buscar vendedores que superen el umbral
        vendedores_alerta = []
        
        for ruta in data_validada.get("rutas_por_dia", []):
            # Buscar km_sap en el diccionario que ya construimos
            vendedor_info = ruta.get("vendedor", {})
            vendedor_codigo = vendedor_info.get("codigo", "") if isinstance(vendedor_info, dict) else ""
            vendedor_nombre = vendedor_info.get("nombre", "") if isinstance(vendedor_info, dict) else vendedor_info
            dia = ruta.get("dia", "")
            key = f"{vendedor_codigo}_{dia}"
            
            km_sap_ruta = km_sap_dict.get(key, 0)
            km_ruta_ruta = ruta.get("distancia_km", 0)
            
            if km_sap_ruta > 0:
                # Calcular diferencia porcentual: (ruta - sap) / sap * 100
                # Si es negativo = ahorro, si es positivo = aumento
                diferencia_porcentaje = ((km_ruta_ruta - km_sap_ruta) / km_sap_ruta * 100)
                
                # Alertar si el aumento supera el umbral positivo (ej: si umbral es -30, alertar si diferencia > 30)
                if diferencia_porcentaje > abs(umbral_porcentaje):
                    vendedores_alerta.append({
                        "codigo": vendedor_codigo,
                        "nombre": vendedor_nombre,
                        "dia": dia,
                        "porcentaje": round(diferencia_porcentaje, 2)
                    })
        
        resumen = {
            "distrito": data_validada.get("distrito", distrito),
            "codigo_distrito": distrito,
            "km_sap": round(km_sap, 2),
            "km_optimizado": round(km_optimizado, 2),
            "km_ruta": round(km_ruta, 2),
            "ahorro_potencial": round(ahorro_potencial, 2),
            "porcentaje_ahorro": round(porcentaje_ahorro, 2),
            "total_rutas": len(data_validada.get("rutas_por_dia", [])),
            "vendedores_alerta": vendedores_alerta,
            "umbral_porcentaje": umbral_porcentaje,
            "tiene_guardado_previo": guardado_previo is not None
        }
        
        if guardado_previo:
            # Calcular km_ruta del guardado previo
            km_ruta_previo = 0
            for ruta in guardado_previo.get("rutas", []):
                km_ruta_previo += ruta.get("distancia_km", 0)
            
            resumen["guardado_previo"] = {
                "fecha": guardado_previo.get("fecha_guardado"),
                "usuario": guardado_previo.get("usuario"),
                "total_rutas": len(guardado_previo.get("rutas", [])),
                "km_ruta": round(km_ruta_previo, 2)
            }
        
        return resumen
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error obteniendo resumen: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al obtener resumen: {str(e)}"
        )


@router.post("/guardar-ruta-vendedor")
async def guardar_ruta_vendedor(
    datos: dict,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Guardar consolidado de rutas para un distrito del jefe de venta
    
    Args:
        datos: Diccionario con distrito y rutas a guardar
        current_user: Usuario actual (debe ser jefe de venta)
        
    Returns:
        Confirmación del guardado
    """
    try:
        # Verificar si la aplicación permite guardar rutas
        import os
        config_file = "data/config.json"
        app_activa = True  # Por defecto activa
        
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    app_activa = config.get("app_activa", True)
            except Exception:
                pass  # Si hay error, asumir que está activa
        
        # Si la app está desactivada, no permitir guardar rutas
        if not app_activa:
            raise HTTPException(
                status_code=403,
                detail="El guardado de rutas está temporalmente desactivado por el administrador"
            )
        
        # Verificar permisos
        if current_user.cargo.lower() == "admin":
            raise HTTPException(
                status_code=403,
                detail="Los administradores no pueden guardar consolidados. Esta función es para jefes de venta."
            )
        
        distrito = datos.get("distrito")
        cod_vendedor = datos.get("cod_vendedor")
        
        if not distrito:
            raise HTTPException(
                status_code=400,
                detail="Se requiere el código del distrito"
            )
        
        if not cod_vendedor:
            raise HTTPException(
                status_code=400,
                detail="Se requiere el código del vendedor"
            )
        
        # Verificar que el usuario tenga permiso para este distrito
        if distrito not in current_user.distritos_permitidos:
            raise HTTPException(
                status_code=403,
                detail=f"No tienes permisos para guardar rutas del distrito {distrito}"
            )
        
        # Ruta al directorio del distrito
        distrito_dir = Path(f"data/routes_json/{distrito}")
        
        if not distrito_dir.exists():
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró el directorio del distrito {distrito}"
            )
        
        # Leer ruta_validada.json para obtener los datos actuales
        ruta_validada_path = distrito_dir / "ruta_validada.json"
        
        if not ruta_validada_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"No hay rutas validadas para guardar en el distrito {distrito}"
            )
        
        with open(ruta_validada_path, 'r', encoding='utf-8') as f:
            data_validada = json.load(f)
        
        # Cargar umbral de porcentaje desde config.json
        config_file = Path(__file__).parent.parent.parent / "data" / "config.json"
        umbral_porcentaje = 10.0  # Valor por defecto
        if config_file.exists():
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    config_data = json.load(f)
                    umbral_porcentaje = config_data.get("umbral_porcentaje", 10.0)
            except Exception as e:
                logger.warning(f"Error al cargar umbral desde config.json: {e}")
        
        # Filtrar solo las rutas del vendedor especificado
        rutas_vendedor = []
        for ruta in data_validada.get("rutas_por_dia", []):
            vendedor_info = ruta.get("vendedor", {})
            cod_vend_ruta = vendedor_info.get("codigo")
            
            # Solo incluir rutas de este vendedor (comparar ambos como string para manejar int/string)
            if str(cod_vend_ruta) == str(cod_vendedor):
                ruta_limpia = ruta.copy()
                # Eliminar geometría y legs si existen
                if "geometria" in ruta_limpia:
                    del ruta_limpia["geometria"]
                if "legs" in ruta_limpia:
                    del ruta_limpia["legs"]
                # Eliminar tiempo de ruta si existe
                if "tiempo_ruta" in ruta_limpia:
                    del ruta_limpia["tiempo_ruta"]
                if "duracion_minutos" in ruta_limpia:
                    del ruta_limpia["duracion_minutos"]
                
                # Calcular y agregar km_holgura y km_validado
                km_ruta = ruta_limpia.get("distancia_km", 0)
                km_holgura = round(km_ruta * (umbral_porcentaje / 100), 2)
                km_validado = round(km_ruta + km_holgura, 2)
                
                ruta_limpia["km_holgura"] = km_holgura
                ruta_limpia["km_validado"] = km_validado
                
                rutas_vendedor.append(ruta_limpia)
        
        if not rutas_vendedor:
            logger.error(f"No se encontraron rutas para vendedor {cod_vendedor} en distrito {distrito}")
            logger.error(f"Total rutas en validada: {len(data_validada.get('rutas_por_dia', []))}")
            if data_validada.get('rutas_por_dia'):
                vendedores_encontrados = set(str(r.get('vendedor', {}).get('codigo', 'N/A')) for r in data_validada.get('rutas_por_dia', []))
                logger.error(f"Vendedores disponibles: {vendedores_encontrados}")
            raise HTTPException(
                status_code=404,
                detail=f"No se encontraron rutas para el vendedor {cod_vendedor} en el distrito {distrito}"
            )
        
        # Verificar si ya existe un guardado previo y cargar rutas existentes
        guardado_path = distrito_dir / "ruta_guardada.json"
        rutas_existentes = []
        
        if guardado_path.exists():
            with open(guardado_path, 'r', encoding='utf-8') as f:
                data_guardada = json.load(f)
                # Filtrar rutas que NO sean del vendedor actual (mantener las demás)
                for ruta in data_guardada.get("rutas", []):
                    vendedor_info = ruta.get("vendedor", {})
                    # Comparar como string para manejar int/string
                    if str(vendedor_info.get("codigo")) != str(cod_vendedor):
                        rutas_existentes.append(ruta)
        
        # Combinar rutas: existentes de otros vendedores + nuevas del vendedor actual
        todas_rutas = rutas_existentes + rutas_vendedor
        
        # Fecha del último guardado (siempre la más reciente)
        fecha_actual = datetime.now(ZoneInfo("America/Santiago")).isoformat()
        
        # Crear estructura del consolidado guardado
        consolidado = {
            "distrito": data_validada.get("distrito", distrito),
            "codigo_distrito": distrito,
            "tipo_ruta": "guardada",
            "total_dias": len(set(r.get("dia", "") for r in todas_rutas)),
            "usuario": current_user.nombre,
            "usuario_email": current_user.email or current_user.usuario,
            "fecha_guardado": fecha_actual,
            "rutas": todas_rutas
        }
        
        # Guardar consolidado
        consolidado_path = distrito_dir / "ruta_guardada.json"
        
        with open(consolidado_path, 'w', encoding='utf-8') as f:
            json.dump(consolidado, f, ensure_ascii=False, indent=2)
        
        return {
            "message": "Rutas del vendedor guardadas exitosamente",
            "distrito": distrito,
            "vendedor": cod_vendedor,
            "usuario": current_user.nombre,
            "total_rutas_vendedor": len(rutas_vendedor),
            "total_rutas_archivo": len(consolidado["rutas"]),
            "fecha_guardado": consolidado["fecha_guardado"]
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error guardando consolidado: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al guardar consolidado: {str(e)}"
        )


@router.get("/estado-consolidados")
async def get_estado_consolidados(current_user: UserInfo = Depends(get_current_user)):
    """
    Obtener estado de guardado de consolidados por distrito
    Solo para administradores
    
    Returns:
        Lista de distritos con su estado de guardado
    """
    try:
        # Solo administradores pueden ver el estado
        if current_user.cargo.lower() != "admin":
            raise HTTPException(
                status_code=403,
                detail="Solo los administradores pueden ver el estado de consolidados"
            )
        
        routes_dir = Path("data/routes_json")
        
        if not routes_dir.exists():
            return {
                "distritos": [],
                "total_distritos": 0,
                "guardados": 0,
                "pendientes": 0
            }
        
        distritos_estado = []
        
        # Obtener lista de usuarios para mapear username a nombre completo
        usuarios = get_all_users()
        usuarios_map = {u.usuario: u.nombre for u in usuarios}
        
        # Iterar sobre todas las carpetas de distritos
        for distrito_dir in routes_dir.iterdir():
            if not distrito_dir.is_dir():
                continue
            
            codigo_distrito = distrito_dir.name
            consolidado_path = distrito_dir / "ruta_guardada.json"
            
            # Leer nombre del distrito desde ruta_original.json
            ruta_original_path = distrito_dir / "ruta_original.json"
            nombre_distrito = codigo_distrito
            
            if ruta_original_path.exists():
                with open(ruta_original_path, 'r', encoding='utf-8') as f:
                    data_original = json.load(f)
                    nombre_distrito = data_original.get("distrito", codigo_distrito)
            
            # Obtener umbral de porcentaje desde config
            config_data = load_config()
            umbral_porcentaje = config_data.get("umbral_porcentaje", 10)
            
            # Primero, obtener vendedores guardados y construir diccionario de rutas guardadas
            vendedores_guardados_set = set()
            rutas_guardadas_por_vendedor_dia = {}  # {(cod_vend, dia): km_ruta}
            total_rutas_guardadas = 0
            fecha_guardado = None
            usuario_username = None
            
            if consolidado_path.exists():
                with open(consolidado_path, 'r', encoding='utf-8') as f:
                    consolidado = json.load(f)
                    fecha_guardado = consolidado.get("fecha_guardado")
                    usuario_username = consolidado.get("usuario")
                    
                    # Obtener rutas guardadas por vendedor y día
                    rutas = consolidado.get("rutas", [])
                    total_rutas_guardadas = len(rutas)
                    
                    for ruta in rutas:
                        vendedor = ruta.get("vendedor", {})
                        dia = ruta.get("dia", "")
                        km_ruta = ruta.get("distancia_km", 0)
                        
                        codigo_vendedor = None
                        if isinstance(vendedor, dict):
                            codigo_vendedor = vendedor.get("codigo")
                        elif vendedor:
                            codigo_vendedor = vendedor
                        
                        if codigo_vendedor:
                            vendedores_guardados_set.add(codigo_vendedor)
                            rutas_guardadas_por_vendedor_dia[(codigo_vendedor, dia)] = km_ruta
            
            # Contar todos los vendedores únicos del distrito (para calcular tasa)
            vendedores_originales = set()
            vendedores_lista = []
            if ruta_original_path.exists():
                with open(ruta_original_path, 'r', encoding='utf-8') as f:
                    data_original = json.load(f)
                    seen_codigos = set()
                    vendedores_guardados_str = {str(v) for v in vendedores_guardados_set}
                    for ruta in data_original.get("rutas_por_dia", []):
                        vendedor = ruta.get("vendedor", {})
                        if isinstance(vendedor, dict):
                            codigo_vendedor = vendedor.get("codigo")
                            if codigo_vendedor:
                                vendedores_originales.add(codigo_vendedor)
                                if str(codigo_vendedor) not in seen_codigos:
                                    seen_codigos.add(str(codigo_vendedor))
                                    guardado_vend = str(codigo_vendedor) in vendedores_guardados_str
                                    vendedores_lista.append({
                                        "codigo": codigo_vendedor,
                                        "nombre": vendedor.get("nombre", "Sin nombre"),
                                        "guardado": guardado_vend
                                    })
                        elif vendedor:
                            vendedores_originales.add(vendedor)
                            if str(vendedor) not in seen_codigos:
                                seen_codigos.add(str(vendedor))
                                guardado_vend = str(vendedor) in vendedores_guardados_str
                                vendedores_lista.append({
                                    "codigo": vendedor,
                                    "nombre": str(vendedor),
                                    "guardado": guardado_vend
                                })
            vendedores_lista.sort(key=lambda x: (0 if x["guardado"] else 1, x["nombre"]))
            
            # Ahora calcular métricas SOLO de los vendedores guardados (igual que ranking-vendedores)
            km_sap_total = 0
            km_optimizado_total = 0
            km_ruta_total = 0
            km_holgura_total = 0
            km_validado_total = 0
            
            # Leer datos originales y sumar solo vendedores guardados
            if ruta_original_path.exists() and vendedores_guardados_set:
                with open(ruta_original_path, 'r', encoding='utf-8') as f:
                    data_original = json.load(f)
                    rutas_originales = data_original.get("rutas_por_dia", [])
                    
                    # Iterar sobre rutas originales
                    for ruta_orig in rutas_originales:
                        vendedor = ruta_orig.get("vendedor", {})
                        dia = ruta_orig.get("dia", "")
                        
                        codigo_vendedor = None
                        if isinstance(vendedor, dict):
                            codigo_vendedor = vendedor.get("codigo")
                        elif vendedor:
                            codigo_vendedor = vendedor
                        
                        # Solo procesar si el vendedor está guardado
                        if codigo_vendedor and codigo_vendedor in vendedores_guardados_set:
                            km_sap_ruta = ruta_orig.get("distancia_km", 0)
                            
                            # Buscar km_ruta guardado para este vendedor y día
                            clave = (codigo_vendedor, dia)
                            if clave in rutas_guardadas_por_vendedor_dia:
                                km_ruta_ruta = rutas_guardadas_por_vendedor_dia[clave]
                                
                                # Calcular holgura y validado POR RUTA (igual que ranking-vendedores)
                                km_holgura_ruta = round(km_ruta_ruta * (umbral_porcentaje / 100), 2)
                                km_validado_ruta = round(km_ruta_ruta + km_holgura_ruta, 2)
                                
                                # Acumular
                                km_sap_total += km_sap_ruta
                                km_ruta_total += km_ruta_ruta
                                km_holgura_total += km_holgura_ruta
                                km_validado_total += km_validado_ruta
            
            # Leer km_optimizado solo de vendedores guardados
            ruta_optimizada_path = distrito_dir / "ruta_optimizada.json"
            if ruta_optimizada_path.exists() and vendedores_guardados_set:
                with open(ruta_optimizada_path, 'r', encoding='utf-8') as f:
                    data_optimizada = json.load(f)
                    rutas_optimizadas = data_optimizada.get("rutas_por_dia", [])
                    
                    for ruta in rutas_optimizadas:
                        vendedor = ruta.get("vendedor", {})
                        codigo_vendedor = None
                        if isinstance(vendedor, dict):
                            codigo_vendedor = vendedor.get("codigo")
                        elif vendedor:
                            codigo_vendedor = vendedor
                        
                        # Solo sumar si el vendedor está guardado
                        if codigo_vendedor and codigo_vendedor in vendedores_guardados_set:
                            km_optimizado_total += ruta.get("distancia_km", 0)
            
            # Redondear totales (ya se redondearon individualmente pero aseguramos precisión)
            km_sap_total = round(km_sap_total, 2)
            km_optimizado_total = round(km_optimizado_total, 2)
            km_ruta_total = round(km_ruta_total, 2)
            km_holgura_total = round(km_holgura_total, 2)
            km_validado_total = round(km_validado_total, 2)
            
            # Calcular diferencia: km_validado - km_sap (igual que ranking-vendedores)
            diferencia_km = round(km_validado_total - km_sap_total, 2)
            
            estado = {
                "codigo_distrito": codigo_distrito,
                "nombre_distrito": nombre_distrito,
                "guardado": consolidado_path.exists(),
                "fecha_guardado": fecha_guardado,
                "usuario": usuario_username,
                "usuario_nombre": usuarios_map.get(usuario_username, usuario_username) if usuario_username else None,
                "total_rutas": total_rutas_guardadas,
                "km_sap": km_sap_total,
                "km_optimizado": km_optimizado_total,
                "km_ruta": km_ruta_total,
                "km_holgura": km_holgura_total,
                "km_validado": km_validado_total,
                "diferencia_km": diferencia_km,
                "total_vendedores": len(vendedores_originales),
                "vendedores_guardados": len(vendedores_guardados_set),
                "vendedores_lista": vendedores_lista
            }
            
            distritos_estado.append(estado)
        
        # Ordenar por código de distrito
        distritos_estado.sort(key=lambda x: x["codigo_distrito"])
        
        guardados = sum(1 for d in distritos_estado if d["guardado"])
        pendientes = len(distritos_estado) - guardados
        
        return {
            "distritos": distritos_estado,
            "total_distritos": len(distritos_estado),
            "guardados": guardados,
            "pendientes": pendientes
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error obteniendo estado de consolidados: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al obtener estado: {str(e)}"
        )

@router.get("/exportar-vendedores-xlsx")
async def exportar_vendedores_xlsx(current_user: UserInfo = Depends(get_current_user)):
    """
    Exporta un archivo XLSX con todos los vendedores, indicando:
    - Nombre del vendedor
    - Código del vendedor
    - Distrito
    - Estado (Validado: Sí/No)
    - KM Semana (KM Validado)
    - KM Mes (KM Validado × Factor Semanas)
    
    Solo accesible para usuarios Admin.
    """
    try:
        # Verificar que sea Admin
        if current_user.cargo.lower() != "admin":
            raise HTTPException(
                status_code=403,
                detail="Solo usuarios Admin pueden exportar este reporte"
            )
        
        # Leer factor_semanas de config.json
        config_path = Path("data/config.json")
        factor_semanas = 4.20  # Valor por defecto
        
        if config_path.exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                config_data = json.load(f)
                factor_semanas = config_data.get("factor_semanas", 4.20)
        
        # Preparar datos
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        vendedores_dict = {}  # Diccionario para agrupar por código de vendedor
        routes_dir = Path("data/routes_json")
        
        if not routes_dir.exists():
            raise HTTPException(
                status_code=404,
                detail="No se encontró el directorio de rutas"
            )
        
        # Recorrer todos los distritos
        for distrito_dir in routes_dir.iterdir():
            if not distrito_dir.is_dir():
                continue
            
            distrito_codigo = distrito_dir.name
            
            # 1. Leer TODOS los vendedores de ruta_original.json (SAP)
            original_path = distrito_dir / "ruta_original.json"
            
            if not original_path.exists():
                continue
            
            try:
                # Leer vendedores originales
                with open(original_path, 'r', encoding='utf-8') as f:
                    data_original = json.load(f)
                
                # Agregar todos los vendedores del distrito
                for ruta in data_original.get("rutas_por_dia", []):
                    vendedor = ruta.get("vendedor", {})
                    cod_vendedor = str(vendedor.get("codigo", ""))
                    
                    if not cod_vendedor:
                        continue
                    
                    # Si el vendedor no existe en el diccionario, crear entrada
                    if cod_vendedor not in vendedores_dict:
                        vendedores_dict[cod_vendedor] = {
                            "NombreVend": vendedor.get("nombre", "N/A"),
                            "CodVend": cod_vendedor,
                            "Distrito": distrito_codigo,
                            "Estado": "No Validado",
                            "KM_Semana": 0
                        }
                
                # 2. Verificar si existe ruta_guardada.json (es el que realmente validó)
                guardada_path = distrito_dir / "ruta_guardada.json"
                
                if guardada_path.exists():
                    with open(guardada_path, 'r', encoding='utf-8') as f:
                        data_guardada = json.load(f)
                    
                    # Procesar rutas guardadas/validadas
                    for ruta in data_guardada.get("rutas", []):
                        vendedor = ruta.get("vendedor", {})
                        cod_vendedor = str(vendedor.get("codigo", ""))
                        
                        if not cod_vendedor:
                            continue
                        
                        # Si el vendedor existe en el diccionario, actualizar
                        if cod_vendedor in vendedores_dict:
                            vendedores_dict[cod_vendedor]["Estado"] = "Validado"
                            # Sumar km_validado de este día
                            km_validado_dia = ruta.get("km_validado", 0)
                            vendedores_dict[cod_vendedor]["KM_Semana"] += km_validado_dia
                    
            except Exception as e:
                logger.warning(f"Error procesando distrito {distrito_codigo}: {e}")
                continue
        
        # Convertir diccionario a lista y calcular KM Mes
        vendedores_data = []
        for vendedor in vendedores_dict.values():
            km_semana = round(vendedor["KM_Semana"], 2)
            km_mes = round(km_semana * factor_semanas, 2)
            
            vendedores_data.append({
                "NombreVend": vendedor["NombreVend"],
                "CodVend": vendedor["CodVend"],
                "Distrito": vendedor["Distrito"],
                "Estado": vendedor["Estado"],
                "KM Semana": km_semana,
                "KM Mes": km_mes
            })
        
        # Crear XLSX
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendedores KM"
        
        # Estilos
        header_fill = PatternFill(start_color="2d7a3e", end_color="2d7a3e", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Encabezados
        headers = ["NombreVend", "CodVend", "Distrito", "Estado", "KM Semana", "KM Mes"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Datos
        for row_num, vendedor in enumerate(vendedores_data, 2):
            ws.cell(row=row_num, column=1, value=vendedor["NombreVend"])
            ws.cell(row=row_num, column=2, value=vendedor["CodVend"])
            ws.cell(row=row_num, column=3, value=vendedor["Distrito"])
            ws.cell(row=row_num, column=4, value=vendedor["Estado"])
            ws.cell(row=row_num, column=5, value=vendedor["KM Semana"])
            ws.cell(row=row_num, column=6, value=vendedor["KM Mes"])
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        
        # Guardar en memoria
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Obtener fecha actual para el nombre del archivo
        from datetime import datetime
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"VendedoresKM_{fecha_actual}.xlsx"
        
        return Response(
            content=excel_file.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exportando XLSX de vendedores: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al generar XLSX: {str(e)}"
        )


@router.get("/exportar-maestro-consolidado-csv")
async def exportar_maestro_consolidado_csv(current_user: UserInfo = Depends(get_current_user)):
    """
    Exporta un CSV con el mismo formato del Maestro APP pero:
    - Solo incluye los distritos que fueron subidos correctamente (tienen carpeta en routes_json/)
    - Agrega 3 columnas nuevas: ValidadaLatitud, ValidadaLongitud, ValidadaSec.Visita
      * Si el distrito tiene ruta_guardada.json y el cliente fue validado, se incluyen los valores validados
      * Si no, esas columnas quedan vacías

    Solo accesible para usuarios Admin.
    """
    try:
        # Verificar que sea Admin
        if current_user.cargo.lower() != "admin":
            raise HTTPException(
                status_code=403,
                detail="Solo usuarios Admin pueden exportar este reporte"
            )

        import csv
        import io

        maestro_path = Path("data/rutas/Maestro APP.csv")
        routes_dir = Path("data/routes_json")

        if not maestro_path.exists():
            raise HTTPException(
                status_code=404,
                detail="No se encontró el archivo Maestro APP.csv"
            )

        if not routes_dir.exists():
            raise HTTPException(
                status_code=404,
                detail="No se encontró el directorio de rutas"
            )

        # 1. Obtener distritos disponibles (que se subieron correctamente)
        distritos_disponibles = set()
        for d in routes_dir.iterdir():
            if d.is_dir() and (d / "ruta_original.json").exists():
                distritos_disponibles.add(d.name)

        # 2. Construir lookup de validados desde ruta_guardada.json
        # Clave: (cod_distrito, dia, cod_vend, cod_cliente) → (lat, lon, sec_opt)
        validados_lookup: dict = {}

        for distrito_dir in routes_dir.iterdir():
            if not distrito_dir.is_dir():
                continue
            guardada_path = distrito_dir / "ruta_guardada.json"
            if not guardada_path.exists():
                continue
            try:
                with open(guardada_path, 'r', encoding='utf-8') as f:
                    data_guardada = json.load(f)
                cod_distrito = data_guardada.get("codigo_distrito", distrito_dir.name)
                for ruta in data_guardada.get("rutas", []):
                    dia = str(ruta.get("dia", "")).upper()
                    cod_vend = str(ruta.get("vendedor", {}).get("codigo", ""))
                    for punto in ruta.get("puntos", []):
                        cod_cliente = str(punto.get("cod_cliente", ""))
                        if cod_cliente in ("0", ""):
                            continue
                        lat = punto.get("coordenadas", {}).get("lat", "")
                        lon = punto.get("coordenadas", {}).get("lon", "")
                        sec_opt = punto.get("secuencia_optimizada", "")
                        key = (cod_distrito, dia, cod_vend, cod_cliente)
                        validados_lookup[key] = (lat, lon, sec_opt)
            except Exception as e:
                logger.warning(f"Error leyendo ruta_guardada de {distrito_dir.name}: {e}")
                continue

        # 3. Leer Maestro APP.csv y generar CSV de salida
        output = io.StringIO()
        new_columns = ["ValidadaLatitud", "ValidadaLongitud", "ValidadaSec.Visita"]

        # Detectar encoding del archivo automáticamente
        _encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
        _detected_encoding = 'latin-1'
        for _enc in _encodings:
            try:
                with open(maestro_path, 'r', encoding=_enc) as _test:
                    _test.read()
                _detected_encoding = _enc
                break
            except (UnicodeDecodeError, LookupError):
                continue

        with open(maestro_path, 'r', encoding=_detected_encoding) as f:
            reader = csv.DictReader(f, delimiter=';')
            original_fieldnames = reader.fieldnames or []

            output_fieldnames = list(original_fieldnames) + new_columns
            writer = csv.DictWriter(
                output,
                fieldnames=output_fieldnames,
                delimiter=';',
                lineterminator='\n',
                extrasaction='ignore'
            )
            writer.writeheader()

            for row in reader:
                cod_distrito = str(row.get("CodDistrito", "")).strip()

                # Solo incluir filas de distritos que fueron subidos correctamente
                if cod_distrito not in distritos_disponibles:
                    continue

                # Buscar datos validados
                dia = str(row.get("DiaVisita", "")).strip().upper()
                cod_vend = str(row.get("CodVend", "")).strip()
                cod_cliente = str(row.get("CodCliente", "")).strip()

                # Para la oficina (CodCliente=0) las coordenadas no cambian al validar:
                # se copian directamente los valores originales en las columnas validadas
                if cod_cliente == "0":
                    row["ValidadaLatitud"] = str(row.get("Latitud", ""))
                    row["ValidadaLongitud"] = str(row.get("Longitud", ""))
                    row["ValidadaSec.Visita"] = str(row.get("Sec.Visita", ""))
                    writer.writerow(row)
                    continue

                key = (cod_distrito, dia, cod_vend, cod_cliente)
                validado = validados_lookup.get(key)

                if validado:
                    lat_val, lon_val, sec_val = validado
                    row["ValidadaLatitud"] = str(lat_val).replace('.', ',')
                    row["ValidadaLongitud"] = str(lon_val).replace('.', ',')
                    row["ValidadaSec.Visita"] = str(sec_val)
                else:
                    row["ValidadaLatitud"] = ""
                    row["ValidadaLongitud"] = ""
                    row["ValidadaSec.Visita"] = ""

                writer.writerow(row)

        csv_content = output.getvalue()
        output.close()

        from datetime import datetime
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"MaestroConsolidado_{fecha_actual}.csv"

        return Response(
            content=csv_content.encode('utf-8-sig'),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exportando Maestro Consolidado CSV: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al generar CSV consolidado: {str(e)}"
        )


# ==================== ENDPOINTS - CAMBIOS MASIVOS (Jefe de Venta) ====================

@router.post("/guardar-consolidado")
async def guardar_consolidado(
    datos: dict,
    current_user: UserInfo = Depends(get_current_user)
):
    """
    Guardar consolidado de rutas VALIDADAS para un distrito (función de Jefe de Venta).

    El jefe de venta acepta el km validado de su distrito y lo deja registrado
    como `ruta_guardada.json`, que luego el administrador puede ver en el estado
    de consolidados.

    Args:
        datos: Diccionario con `distrito` (código de distrito)
        current_user: Usuario actual (debe ser jefe de venta, no admin)

    Returns:
        Confirmación del guardado con totales
    """
    try:
        # Verificar si la aplicación Optimiza Rutas está activa
        config = load_config()
        app_activa = config.get("app_activa", True)

        if not app_activa:
            raise HTTPException(
                status_code=403,
                detail="El guardado de rutas está temporalmente desactivado por el administrador"
            )

        # Los administradores usan /guardar-rutas-vendedor, no este endpoint
        if current_user.cargo.lower() == "admin":
            raise HTTPException(
                status_code=403,
                detail="Los administradores no pueden guardar consolidados. Esta función es para jefes de venta."
            )

        distrito = datos.get("distrito")
        if not distrito:
            raise HTTPException(status_code=400, detail="Se requiere el código del distrito")

        # Verificar que el usuario tenga permiso para este distrito
        if distrito not in current_user.distritos_permitidos:
            raise HTTPException(
                status_code=403,
                detail=f"No tienes permisos para guardar rutas del distrito {distrito}"
            )

        # Rutas de archivos (absolutas)
        routes_dir = Path(__file__).parent.parent.parent / "data" / "routes_json"
        distrito_dir = routes_dir / distrito

        if not distrito_dir.exists():
            raise HTTPException(
                status_code=404,
                detail=f"No se encontró el directorio del distrito {distrito}"
            )

        # Leer ruta_validada.json para obtener los datos actuales
        ruta_validada_path = distrito_dir / "ruta_validada.json"
        if not ruta_validada_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"No hay rutas validadas para guardar en el distrito {distrito}"
            )

        with open(ruta_validada_path, 'r', encoding='utf-8') as f:
            data_validada = json.load(f)

        # Limpiar geometría y legs de las rutas antes de guardar (reducir tamaño del JSON)
        rutas_limpias = []
        for ruta in data_validada.get("rutas_por_dia", []):
            ruta_limpia = ruta.copy()
            ruta_limpia.pop("geometria", None)
            ruta_limpia.pop("legs", None)
            rutas_limpias.append(ruta_limpia)

        # Verificar si ya existe un guardado previo
        guardado_path = distrito_dir / "ruta_guardada.json"
        es_primer_guardado = not guardado_path.exists()

        # Crear estructura del consolidado guardado
        consolidado = {
            "distrito": data_validada.get("distrito", distrito),
            "codigo_distrito": distrito,
            "tipo_ruta": "guardada",
            "total_dias": data_validada.get("total_dias", len(rutas_limpias)),
            "usuario": current_user.nombre,
            "usuario_email": getattr(current_user, 'email', None) or current_user.usuario,
            "fecha_guardado": datetime.now(ZoneInfo("America/Santiago")).isoformat(),
            "es_primer_guardado": es_primer_guardado,
            "rutas": rutas_limpias
        }

        with open(guardado_path, 'w', encoding='utf-8') as f:
            json.dump(consolidado, f, ensure_ascii=False, indent=2)

        return {
            "message": "Consolidado guardado exitosamente",
            "distrito": distrito,
            "usuario": current_user.nombre,
            "total_rutas": len(consolidado["rutas"]),
            "fecha_guardado": consolidado["fecha_guardado"]
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error guardando consolidado: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error al guardar consolidado: {str(e)}"
        )
