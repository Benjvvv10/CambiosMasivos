"""
Servicio de gestión de usuarios
Maneja operaciones CRUD y lectura desde Excel
"""

import pandas as pd
import os
import secrets
import string
import logging
from openpyxl import load_workbook
from typing import Optional, List, Dict
from app.config import settings
from app.models.user import UserInDB
from pathlib import Path

logger = logging.getLogger(__name__)


def read_csv_auto_encoding(file_path: Path, sep: str = ';') -> pd.DataFrame:
    """
    Lee un CSV intentando automáticamente múltiples codificaciones.
    Evita errores de UnicodeDecodeError en archivos con caracteres especiales.
    """
    encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252', 'windows-1252']

    for encoding in encodings:
        try:
            df = pd.read_csv(file_path, sep=sep, encoding=encoding)
            logger.info(f"CSV leído exitosamente con codificación: {encoding}")
            return df
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception as e:
            raise e

    raise ValueError(f"No se pudo leer el archivo CSV con ninguna codificación estándar")


def get_distritos_from_maestro() -> Dict[str, str]:
    """
    Leer códigos y nombres de distrito del archivo Maestro APP.csv
    Retorna diccionario {código: nombre}
    """
    try:
        maestro_path = Path(settings.DATA_DIR) / "rutas" / "Maestro APP.csv"
        if not maestro_path.exists():
            return {}

        df = read_csv_auto_encoding(maestro_path, sep=';')

        # Obtener pares únicos de CodDistrito y Distrito
        if 'CodDistrito' in df.columns and 'Distrito' in df.columns:
            distritos_df = df[['CodDistrito', 'Distrito']].drop_duplicates()
            return dict(zip(distritos_df['CodDistrito'], distritos_df['Distrito']))
        return {}
    except Exception as e:
        print(f"Error al leer distritos del Maestro APP: {e}")
        return {}


def get_jefe_venta_from_maestro(distrito: str) -> str:
    """
    Obtener el nombre del jefe de venta para un distrito específico desde Maestro APP.csv.
    Retorna el nombre del jefe de venta (columna NombreJV).
    """
    try:
        maestro_path = Path(settings.DATA_DIR) / "rutas" / "Maestro APP.csv"
        if not maestro_path.exists():
            return ""

        df = read_csv_auto_encoding(maestro_path, sep=';')

        # Filtrar por código de distrito y obtener el primer NombreJV
        if 'CodDistrito' in df.columns and 'NombreJV' in df.columns:
            distrito_df = df[df['CodDistrito'] == distrito]
            if not distrito_df.empty:
                return str(distrito_df['NombreJV'].iloc[0])
        return ""
    except Exception as e:
        print(f"Error al leer jefe de venta del Maestro APP: {e}")
        return ""


class UserService:
    """Servicio para gestión de usuarios"""
    
    def __init__(self):
        """Inicializar servicio de usuarios"""
        self.users_file = os.path.join(settings.DATA_DIR, settings.USERS_FILE)
        self._users_cache = None  # Cache de usuarios en memoria
        
    def _load_users_from_excel(self) -> pd.DataFrame:
        """
        Cargar usuarios desde archivo Excel
        
        Returns:
            DataFrame con usuarios o DataFrame vacío si hay error
        """
        try:
            if not os.path.exists(self.users_file):
                print(f"Archivo de usuarios no encontrado: {self.users_file}")
                return pd.DataFrame()
            
            # Leer archivo Excel
            df = pd.read_excel(self.users_file)
            
            # Validar columnas requeridas
            required_columns = ['Usuario', 'Contraseña', 'Nombre', 'Cargo']
            for col in required_columns:
                if col not in df.columns:
                    print(f"Columna requerida '{col}' no encontrada en archivo de usuarios")
                    return pd.DataFrame()
            
            # Convertir NaN a 0 en columnas de distritos
            distritos_dict = get_distritos_from_maestro()
            for codigo in distritos_dict.keys():
                if codigo in df.columns:
                    df[codigo] = df[codigo].fillna(0).astype(int)
            
            # Asegurar que columnas Bloquear y Cambiar existan
            if 'Bloquear' not in df.columns:
                df['Bloquear'] = 0
            if 'Cambiar' not in df.columns:
                df['Cambiar'] = 0
                
            df['Bloquear'] = df['Bloquear'].fillna(0).astype(int)
            df['Cambiar'] = df['Cambiar'].fillna(0).astype(int)
            
            return df
            
        except Exception as e:
            print(f"Error al cargar usuarios desde Excel: {e}")
            return pd.DataFrame()
    
    def get_user_by_username(self, username: str) -> Optional[UserInDB]:
        """
        Obtener usuario por nombre de usuario
        
        Args:
            username: Nombre de usuario
            
        Returns:
            UserInDB si se encuentra, None en caso contrario
        """
        df = self._load_users_from_excel()
        
        if df.empty:
            return None
        
        # Buscar usuario
        user_row = df[df['Usuario'] == username]
        
        if user_row.empty:
            return None
        
        user_data = user_row.iloc[0]
        
        # Obtener distritos permitidos
        distritos = {}
        distritos_dict = get_distritos_from_maestro()
        for codigo in distritos_dict.keys():
            if codigo in user_data.index:
                distritos[codigo] = int(user_data[codigo]) if pd.notna(user_data[codigo]) else 0
        
        # Obtener email (verificar si existe y no es NaN)
        email_value = None
        if 'Email' in user_data.index or 'Correo' in user_data.index:
            email_value = user_data.get('Email') or user_data.get('Correo')
            if pd.notna(email_value) and email_value and str(email_value).strip():
                email_value = str(email_value).strip()
            else:
                email_value = None
        
        # Crear objeto UserInDB
        return UserInDB(
            usuario=user_data['Usuario'],
            nombre=user_data['Nombre'],
            contraseña=user_data['Contraseña'],
            cargo=user_data.get('Cargo', 'Usuario'),
            email=email_value,
            bloquear=int(user_data.get('Bloquear', 0)),
            cambiar=int(user_data.get('Cambiar', 0)),
            distritos=distritos
        )
    
    def get_user_by_email(self, email: str) -> Optional[UserInDB]:
        """
        Obtener usuario por correo electrónico
        
        Args:
            email: Correo electrónico del usuario
            
        Returns:
            UserInDB si se encuentra, None en caso contrario
        """
        df = self._load_users_from_excel()
        
        if df.empty:
            return None
        
        # Buscar por columna Email o Correo
        if 'Email' in df.columns:
            user_row = df[df['Email'].str.lower() == email.lower()]
        elif 'Correo' in df.columns:
            user_row = df[df['Correo'].str.lower() == email.lower()]
        else:
            return None
        
        if user_row.empty:
            return None
        
        # Usar el username para obtener el usuario completo
        username = user_row.iloc[0]['Usuario']
        return self.get_user_by_username(username)
    
    def get_user_allowed_districts(self, username: str) -> List[str]:
        """
        Obtener lista de distritos permitidos para un usuario
        
        Args:
            username: Nombre de usuario
            
        Returns:
            Lista de códigos de distritos permitidos (ej: ['S2', 'AR'])
        """
        user = self.get_user_by_username(username)
        
        if not user:
            return []
        
        # Filtrar distritos con valor 1 y retornar códigos
        allowed_districts = []
        for codigo, valor in user.distritos.items():
            if valor == 1:
                allowed_districts.append(codigo)
        
        return allowed_districts
    
    def update_user_password(self, username: str, new_password: str, reset_cambiar: bool = True) -> bool:
        """
        Actualizar contraseña de usuario
        
        Args:
            username: Nombre de usuario
            new_password: Nueva contraseña
            reset_cambiar: Si True, pone Cambiar=0
            
        Returns:
            True si se actualizó correctamente, False en caso contrario
        """
        try:
            if not os.path.exists(self.users_file):
                return False
            
            # Leer Excel
            df = pd.read_excel(self.users_file)
            
            # Buscar usuario
            idx = df[df['Usuario'] == username].index
            
            if len(idx) == 0:
                return False
            
            # Actualizar contraseña
            df.loc[idx, 'Contraseña'] = new_password
            
            # Actualizar campo Cambiar si se solicita
            if reset_cambiar:
                df.loc[idx, 'Cambiar'] = 0
            
            # Guardar archivo
            df.to_excel(self.users_file, index=False)
            
            # Limpiar cache
            self._users_cache = None
            
            return True
            
        except Exception as e:
            print(f"Error al actualizar contraseña: {e}")
            return False
    
    def is_user_blocked(self, username: str) -> bool:
        """
        Verificar si un usuario está bloqueado
        
        Args:
            username: Nombre de usuario
            
        Returns:
            True si está bloqueado, False en caso contrario
        """
        user = self.get_user_by_username(username)
        return user.bloquear == 1 if user else False
    
    def get_all_users(self) -> List[UserInDB]:
        """
        Obtener todos los usuarios del sistema
        
        Returns:
            Lista de objetos UserInDB
        """
        df = self._load_users_from_excel()
        
        if df.empty:
            return []
        
        users = []
        distritos_dict = get_distritos_from_maestro()
        
        for _, user_data in df.iterrows():
            # Obtener distritos permitidos
            distritos = {}
            for codigo in distritos_dict.keys():
                if codigo in user_data.index:
                    distritos[codigo] = int(user_data[codigo]) if pd.notna(user_data[codigo]) else 0
            
            # Obtener email (verificar si existe y no es NaN)
            email_value = None
            if 'Email' in user_data.index or 'Correo' in user_data.index:
                email_value = user_data.get('Email') or user_data.get('Correo')
                if pd.notna(email_value) and email_value and str(email_value).strip():
                    email_value = str(email_value).strip()
                else:
                    email_value = None
            
            # Crear objeto UserInDB
            user = UserInDB(
                usuario=user_data['Usuario'],
                nombre=user_data['Nombre'],
                contraseña=user_data['Contraseña'],
                cargo=user_data.get('Cargo', 'Usuario'),
                email=email_value,
                bloquear=int(user_data.get('Bloquear', 0)),
                cambiar=int(user_data.get('Cambiar', 0)),
                distritos=distritos
            )
            users.append(user)
        
        return users
    
    def generate_temporary_password(self) -> str:
        """
        Generar contraseña temporal aleatoria
        
        Returns:
            Contraseña temporal de 10 caracteres
        """
        caracteres = string.ascii_letters + string.digits + "!@#$%&*"
        return ''.join(secrets.choice(caracteres) for _ in range(10))
    
    def recover_password(self, username: str) -> tuple[bool, Optional[str], Optional[str]]:
        """
        Recuperar contraseña generando una temporal y enviando por correo
        
        Args:
            username: Nombre de usuario
            
        Returns:
            Tupla (éxito, contraseña_temporal, correo)
        """
        try:
            # Obtener usuario
            user = self.get_user_by_username(username)
            
            if not user:
                print(f"❌ Usuario {username} no encontrado")
                return False, None, None
            
            if not user.email:
                print(f"❌ Usuario {username} no tiene email configurado")
                return False, None, None
            
            print(f"✅ Usuario encontrado: {username}, Email: {user.email}")
            
            # Generar contraseña temporal
            password_temporal = self.generate_temporary_password()
            print(f"✅ Contraseña temporal generada: {password_temporal}")
            
            # Leer Excel
            df = self._load_users_from_excel()
            
            if df.empty:
                print("❌ Error al cargar archivo de usuarios")
                return False, None, None
            
            # Buscar usuario
            idx = df[df['Usuario'] == username].index
            
            if len(idx) == 0:
                print(f"❌ Usuario {username} no encontrado en Excel")
                return False, None, None
            
            print(f"✅ Actualizando Excel para usuario {username}")
            
            # Actualizar contraseña y poner Cambiar=1 usando openpyxl para evitar problemas de permisos
            try:
                # Cargar el workbook directamente
                wb = load_workbook(self.users_file)
                ws = wb.active
                
                # Encontrar columnas
                headers = [cell.value for cell in ws[1]]
                usuario_col = headers.index('Usuario') + 1
                password_col = headers.index('Contraseña') + 1
                cambiar_col = headers.index('Cambiar') + 1
                
                # Buscar fila del usuario y actualizar
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                    if row[usuario_col - 1].value == username:
                        ws.cell(row=row_idx, column=password_col, value=password_temporal)
                        ws.cell(row=row_idx, column=cambiar_col, value=1)
                        break
                
                # Guardar el archivo
                wb.save(self.users_file)
                wb.close()
                
                print(f"✅ Excel guardado exitosamente")
                
            except Exception as save_error:
                print(f"❌ Error al guardar Excel con openpyxl: {save_error}")
                # Intentar con pandas como fallback
                df.loc[idx, 'Contraseña'] = password_temporal
                df.loc[idx, 'Cambiar'] = 1
                df.to_excel(self.users_file, index=False)
                print(f"✅ Excel guardado con pandas (fallback)")
            
            # Limpiar cache
            self._users_cache = None
            
            return True, password_temporal, user.email
            
        except Exception as e:
            print(f"❌ Error al recuperar contraseña: {e}")
            import traceback
            traceback.print_exc()
            return False, None, None

# Función auxiliar para obtener todos los usuarios
def get_all_users() -> List[UserInDB]:
    """
    Obtener todos los usuarios del sistema
    
    Returns:
        Lista de objetos UserInDB
    """
    return user_service.get_all_users()


# Instancia global del servicio
user_service = UserService()
